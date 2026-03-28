#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Garimpagem de Lotes com Gemini Vision
======================================
Usa Google Gemini Flash (gratuito) para analisar fotos dos lotes.

Modos:
  desconhecidos  Identifica artista de lotes sem autoria (padrão)
  hoje           Analisa TODOS os lotes de hoje + amanhã (LBR + Tableau)

Uso:
  python garimpo.py                         # desconhecidos (LBR)
  python garimpo.py --modo hoje             # hoje+amanhã, todos os lotes
  python garimpo.py --modo hoje --max 30    # limita a 30 lotes (teste)
  python garimpo.py --casa spiti.art        # só uma casa
  python garimpo.py --key SUA_KEY           # passa API key inline
  python garimpo.py --so-texto              # análise básica sem visão
"""

import sys, os, json, re, time, unicodedata, argparse, io
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import requests
import numpy as np
from datetime import date as _date_cls, timedelta, datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Argumentos ────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--modo",  choices=["desconhecidos","hoje"], default="desconhecidos",
                    help="desconhecidos=só sem autoria (padrão) | hoje=todos de hoje+amanhã")
parser.add_argument("--max",   type=int, default=0,  help="Máximo de lotes a analisar")
parser.add_argument("--casa",  type=str, default="", help="Filtrar por casa")
parser.add_argument("--key",   type=str, default="", help="Google API key")
parser.add_argument("--so-texto", action="store_true", help="Analisa só descrição, sem visão")
args = parser.parse_args()

# ── API Key ───────────────────────────────────────────────────────────────────
GOOGLE_KEY = (
    args.key
    or os.environ.get("GOOGLE_API_KEY", "")
    or os.environ.get("GEMINI_API_KEY", "")
)
if not GOOGLE_KEY and not args.so_texto:
    cfg = os.path.join(_DIR, "config.json")
    if os.path.exists(cfg):
        with open(cfg, "r") as f:
            GOOGLE_KEY = json.load(f).get("google_api_key", "")

if not GOOGLE_KEY and not args.so_texto:
    print("=" * 65)
    print("  GOOGLE API KEY não encontrada.")
    print("  Opções:")
    print("  1) python garimpo.py --key SUA_KEY")
    print("  2) set GOOGLE_API_KEY=SUA_KEY")
    print("  3) crie config.json com {\"google_api_key\": \"SUA_KEY\"}")
    print("  4) python garimpo.py --so-texto  (sem API)")
    print()
    print("  Chave grátis em: https://aistudio.google.com/apikey")
    print("=" * 65)
    sys.exit(1)

# ── Gemini ────────────────────────────────────────────────────────────────────
USE_GEMINI = bool(GOOGLE_KEY)
MODELO = "gemini-1.5-flash"   # 15 RPM, 1500 req/dia no free tier
if USE_GEMINI:
    from google import genai
    from google.genai import types as gtypes
    gemini = genai.Client(api_key=GOOGLE_KEY)
    for _m in ["gemini-1.5-flash", "gemini-2.0-flash", "gemini-2.5-flash"]:
        try:
            gemini.models.generate_content(model=_m, contents="ping")
            MODELO = _m
            break
        except Exception:
            continue

DELAY_API = 6    # segundos entre chamadas (15 RPM = 1 a cada 4s, 6s com margem)

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(s):
    s = re.sub(r'\(.*?\)', '', s or '').strip()
    nfkd = unicodedata.normalize('NFKD', s.upper())
    s = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', re.sub(r'[^A-Z\s]', '', s)).strip()

def vb(v):
    try: return float(v or 0)
    except: return 0.0

def parse_dims(s):
    if not s: return 0.0
    s = str(s).lower().replace(' ','').replace('cm','').replace('mm','')
    m = re.match(r'([\d,\.]+)[x×]([\d,\.]+)', s)
    if m:
        try:
            w = float(m.group(1).replace(',','.'))
            h = float(m.group(2).replace(',','.'))
            if 0 < w < 1000 and 0 < h < 1000:
                return round(w * h, 1)
        except: pass
    return 0.0

def _parse_lote_date(s):
    """Extrai date de string de data (LBR ou Tableau)."""
    s = str(s or '').strip()
    # Tableau: "TERÇA FEIRA (24/03/2026) a partir de 20h"
    m = re.search(r'\((\d{1,2}/\d{2}/\d{4})\)', s)
    if m:
        parts = m.group(1).split('/')
    else:
        parts = s.split('/')
    if len(parts) == 3:
        try:
            return _date_cls(int(parts[2]), int(parts[1]), int(parts[0]))
        except: pass
    return None

# ── Carregar histórico de preços ───────────────────────────────────────────────
print("Carregando bases históricas...")
historico = {}      # artista_norm → [preços]
historico_cm2 = {}
catalogo_nomes = set()

def add_hist(artista, preco, dims=''):
    k = norm(artista)
    if not k: return
    catalogo_nomes.add(k)
    if preco <= 0: return
    historico.setdefault(k, []).append(preco)
    area = parse_dims(dims)
    if area > 0:
        historico_cm2.setdefault(k, []).append(round(preco / area, 4))

for arq in ["bolsadearte_db.json", "arrematearte_db.json", "cda_db.json"]:
    p = os.path.join(_DIR, arq)
    if not os.path.exists(p): continue
    with open(p, "r", encoding="utf-8") as f:
        db = json.load(f)
    for k, v in db.items():
        if isinstance(v, dict):
            add_hist(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

lbr_path = os.path.join(_DIR, "leiloesbr_db.json")
with open(lbr_path, "r", encoding="utf-8") as f:
    lbr = json.load(f)
for v in lbr.values():
    if isinstance(v, dict) and not v.get('_ignorado'):
        add_hist(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

print(f"  {len(catalogo_nomes)} artistas | {len(historico)} com preços | {sum(len(v) for v in historico.values())} registros")

# ── Checkpoint ────────────────────────────────────────────────────────────────
CHECKPOINT_FILE = os.path.join(_DIR, "garimpo_checkpoint.json")
checkpoint = {}
if os.path.exists(CHECKPOINT_FILE):
    try:
        with open(CHECKPOINT_FILE, 'r', encoding='utf-8') as f:
            checkpoint = json.load(f)
        print(f"  Checkpoint: {len(checkpoint)} lotes já analisados")
    except: pass

def save_checkpoint(key, data):
    checkpoint[key] = data
    try:
        with open(CHECKPOINT_FILE, 'w', encoding='utf-8') as f:
            json.dump(checkpoint, f, ensure_ascii=False, indent=2)
    except: pass

# ── Artistas desconhecidos ─────────────────────────────────────────────────────
DESCONHECIDO = {
    "autor desconhecido","desconhecido","?","","não identificado",
    "artista desconhecido","sem assinatura","sem identificação",
    "desconhecida","autor não identificado","artista não identificado",
}

def _is_desconhecido(artista):
    a = (artista or '').strip().lower()
    return not a or a in DESCONHECIDO or 'desconhec' in a or 'nao ident' in norm(a)

# ── Coletar lotes ─────────────────────────────────────────────────────────────
TODAY    = _date_cls.today()
TOMORROW = TODAY + timedelta(days=1)

print(f"\nModo: {args.modo.upper()}")
lotes_garimpo = []

if args.modo == 'desconhecidos':
    print("Filtrando lotes com autor desconhecido (LBR)...")
    for k, v in lbr.items():
        if not isinstance(v, dict) or v.get('_ignorado'): continue
        if not v.get('em_leilao', True): continue
        if not _is_desconhecido(v.get('artista','')): continue
        if args.casa and args.casa not in v.get('casa',''): continue

        foto = v.get('foto_url','') or ''
        ass  = v.get('assinatura','') or ''
        tec  = v.get('tecnica','') or ''
        prio = 0
        if foto: prio += 3
        if ass and ass != 'Não assinado': prio += 2
        if tec: prio += 1
        if v.get('dimensoes'): prio += 1
        if v.get('titulo'): prio += 1

        lotes_garimpo.append({
            'key': k, 'foto': foto,
            'artista_original': v.get('artista','') or 'Desconhecido',
            'desconhecido': True,
            'assinatura': ass, 'tecnica': tec,
            'dimensoes': v.get('dimensoes','') or '',
            'titulo': v.get('titulo','') or '',
            'ano': v.get('ano','') or '',
            'casa': v.get('casa',''),
            'preco': vb(v.get('lance_base')),
            'url': v.get('url_detalhe',''),
            'prio': prio,
            'data': v.get('data_leilao','') or v.get('data_coleta',''),
            'source': 'lbr',
        })

elif args.modo == 'hoje':
    print(f"Filtrando lotes de hoje ({TODAY.strftime('%d/%m')}) e amanhã ({TOMORROW.strftime('%d/%m')})...")

    # LBR
    for k, v in lbr.items():
        if not isinstance(v, dict) or v.get('_ignorado'): continue
        if not v.get('em_leilao', True): continue
        if args.casa and args.casa not in v.get('casa',''): continue

        dt = _parse_lote_date(v.get('data_leilao',''))
        if dt not in (TODAY, TOMORROW): continue

        foto = v.get('foto_url','') or ''
        art  = (v.get('artista','') or '').strip()
        is_desc = _is_desconhecido(art)
        prio = 20 if is_desc else 0   # desconhecidos têm prioridade
        if foto: prio += 3
        preco = vb(v.get('lance_base'))
        if preco > 0: prio += min(10, int(preco / 1000))  # preço maior = mais urgente

        lotes_garimpo.append({
            'key': k, 'foto': foto,
            'artista_original': art or 'Desconhecido',
            'desconhecido': is_desc,
            'assinatura': v.get('assinatura','') or '',
            'tecnica': v.get('tecnica','') or '',
            'dimensoes': v.get('dimensoes','') or '',
            'titulo': v.get('titulo','') or '',
            'ano': v.get('ano','') or '',
            'casa': v.get('casa',''),
            'preco': preco,
            'url': v.get('url_detalhe',''),
            'prio': prio,
            'data': v.get('data_leilao','') or v.get('data_coleta',''),
            'source': 'lbr',
        })

    # Tableau
    tab_path = os.path.join(_DIR, "tableau_db.json")
    if os.path.exists(tab_path):
        with open(tab_path, 'r', encoding='utf-8') as f:
            tab_db = json.load(f)
        tab_added = 0
        for item in tab_db:
            if not isinstance(item, dict): continue
            dt = _parse_lote_date(item.get('data_leilao',''))
            if dt not in (TODAY, TOMORROW): continue

            lote_num = item.get('lote_num', 0)
            if args.casa and 'tableau' not in args.casa: continue

            foto = item.get('img_grande','') or item.get('img_thumb','') or ''
            art_raw = (item.get('artista','') or '').strip()
            # Remove anos "(1922 - 2006)" do nome
            art = re.sub(r'\s*\(\d{4}\s*[-–]\s*\d{4}\)\s*$', '', art_raw).strip()
            is_desc = _is_desconhecido(art)
            preco = vb(item.get('valor_base'))
            prio = 20 if is_desc else 0
            if foto: prio += 3
            if preco > 0: prio += min(10, int(preco / 1000))

            key = f"tableau_{lote_num}"
            lotes_garimpo.append({
                'key': key, 'foto': foto,
                'artista_original': art or 'Desconhecido',
                'desconhecido': is_desc,
                'assinatura': item.get('assinado','') or '',
                'tecnica': item.get('tecnica','') or '',
                'dimensoes': item.get('medidas','') or '',
                'titulo': item.get('titulo','') or '',
                'ano': item.get('data_obra','') or '',
                'casa': 'tableau.com.br',
                'preco': preco,
                'url': item.get('url_lote',''),
                'prio': prio,
                'data': item.get('data_leilao',''),
                'source': 'tableau',
            })
            tab_added += 1
        print(f"  Tableau: {tab_added} lotes")

# Remove já analisados do checkpoint (mas mantém na lista para o Excel)
ja_analisados = {k for k in checkpoint if k in {l['key'] for l in lotes_garimpo}}
lotes_novos   = [l for l in lotes_garimpo if l['key'] not in ja_analisados]

lotes_garimpo.sort(key=lambda x: -x['prio'])
lotes_novos.sort(key=lambda x: -x['prio'])

if args.max:
    lotes_novos = lotes_novos[:args.max]

desc_count = sum(1 for l in lotes_garimpo if l.get('desconhecido'))
print(f"  Total: {len(lotes_garimpo)} lotes | Desconhecidos: {desc_count}")
print(f"  Já analisados: {len(ja_analisados)} | A analisar: {len(lotes_novos)}")
print(f"  Com foto: {sum(1 for l in lotes_novos if l['foto'])}")

# ── Prompts ───────────────────────────────────────────────────────────────────
PROMPT_IDENTIFICAR = """Você é um especialista em arte brasileira e leilões de arte.
Analise a imagem desta obra e responda em JSON com exatamente estas chaves:

{
  "assinatura_visivel": "texto exato da assinatura ou 'não visível'",
  "assinatura_localizacao": "canto inf. dir., inf. esq., sup. dir., etc. ou 'não visível'",
  "tecnica_confirmada": "técnica que você vê na imagem",
  "periodo_estimado": "ex: 1950-1970, século XIX, etc.",
  "escola_estilo": "ex: modernismo brasileiro, paisagismo carioca, arte naif, etc.",
  "tema_descricao": "descrição breve do que representa a obra",
  "artistas_sugeridos": ["Nome Artista 1", "Nome Artista 2"],
  "justificativa": "breve explicação dos motivos da sugestão",
  "confianca": "alta / média / baixa",
  "autenticidade": "confirmada / duvidosa / suspeita",
  "assinatura_confere": "sim / parcial / não / não visível",
  "coerencia_estilo": "sim / não / parcial",
  "red_flags": "problemas encontrados ou 'nenhum'",
  "valor_estimado_min": 0,
  "valor_estimado_max": 0
}

Para artistas_sugeridos: liste até 3 artistas brasileiros que poderiam ser o autor.
Se a assinatura for legível, use como prioridade máxima.
Para valor_estimado: estime em R$ considerando o mercado brasileiro atual."""

PROMPT_VALIDAR = """Você é um especialista em arte brasileira e leilões de arte.
Esta obra é atribuída a {artista}.
Analise a imagem e responda em JSON com exatamente estas chaves:

{{
  "assinatura_visivel": "texto exato da assinatura ou 'não visível'",
  "assinatura_localizacao": "localização ou 'não visível'",
  "assinatura_confere": "sim / parcial / não / não visível",
  "tecnica_confirmada": "técnica que você vê",
  "periodo_estimado": "ex: 1950-1970",
  "escola_estilo": "escola/estilo da obra",
  "tema_descricao": "descrição breve",
  "artistas_sugeridos": ["{artista}"],
  "coerencia_estilo": "sim / não / parcial",
  "autenticidade": "confirmada / duvidosa / suspeita",
  "red_flags": "problemas encontrados ou 'nenhum'",
  "justificativa": "análise resumida da autenticidade",
  "confianca": "alta / média / baixa",
  "valor_estimado_min": 0,
  "valor_estimado_max": 0
}}

Verifique:
1) A assinatura bate com a forma conhecida de {artista} assinar?
2) O estilo, técnica e período são coerentes com o trabalho de {artista}?
3) Há algo suspeito (tela nova, pintura sobre outra obra, inconsistências)?"""

PROMPT_TEXTO = """Você é um especialista em arte brasileira e leilões de arte.
Com base nesta descrição de lote (SEM imagem), responda em JSON:

DESCRIÇÃO: {descricao}

{
  "assinatura_visivel": "conforme descrição ou 'não informado'",
  "assinatura_localizacao": "",
  "tecnica_confirmada": "conforme descrição",
  "periodo_estimado": "estimativa baseada na técnica/título",
  "escola_estilo": "escola/estilo provável",
  "tema_descricao": "tema baseado no título",
  "artistas_sugeridos": ["Nome Artista 1"],
  "justificativa": "breve explicação",
  "confianca": "baixa",
  "autenticidade": "não avaliado",
  "assinatura_confere": "não avaliado",
  "coerencia_estilo": "não avaliado",
  "red_flags": "não avaliado",
  "valor_estimado_min": 0,
  "valor_estimado_max": 0
}"""

# ── Funções de análise ─────────────────────────────────────────────────────────
session = requests.Session()
session.headers.update({"User-Agent": "Mozilla/5.0"})

def baixar_imagem(url):
    try:
        r = session.get(url, timeout=15)
        if r.status_code == 200 and len(r.content) > 1000:
            return r.content
    except: pass
    return None

def parse_resposta(texto):
    texto = re.sub(r'```(?:json)?\s*', '', texto).strip('`').strip()
    m = re.search(r'\{.*\}', texto, re.DOTALL)
    if m:
        try: return json.loads(m.group())
        except: pass
    return None

def analisar_com_gemini(lote, tentativa=1):
    """Analisa lote. Retorna (dict, modo_str). Modo: visao/texto/offline/erro:..."""
    if not USE_GEMINI:
        return _analisar_offline(lote), "offline"

    descricao = f"{lote['titulo']} | {lote['tecnica']} | {lote['dimensoes']} | assinatura: {lote['assinatura']} | ano: {lote['ano']}"

    try:
        if not args.so_texto and lote['foto']:
            img_bytes = baixar_imagem(lote['foto'])
            if img_bytes:
                img = PILImage.open(io.BytesIO(img_bytes))
                img.thumbnail((800, 800), PILImage.LANCZOS)
                buf = io.BytesIO()
                img.save(buf, format='JPEG', quality=75)
                img_bytes = buf.getvalue()

                if lote['desconhecido']:
                    prompt_final = PROMPT_IDENTIFICAR + f"\n\nDescrição: {descricao}"
                else:
                    artista = lote['artista_original']
                    prompt_final = PROMPT_VALIDAR.format(artista=artista) + f"\n\nDescrição: {descricao}"

                resp = gemini.models.generate_content(
                    model=MODELO,
                    contents=[
                        gtypes.Part.from_bytes(data=img_bytes, mime_type="image/jpeg"),
                        prompt_final,
                    ]
                )
                return parse_resposta(resp.text), "visao"

        # Fallback: só texto
        if lote['desconhecido']:
            prompt = PROMPT_TEXTO.format(descricao=descricao)
        else:
            artista = lote['artista_original']
            prompt = PROMPT_VALIDAR.format(artista=artista) + f"\n\nDescrição: {descricao}\n(Sem imagem disponível)"

        resp = gemini.models.generate_content(model=MODELO, contents=prompt)
        return parse_resposta(resp.text), "texto"

    except Exception as e:
        err_str = str(e)
        # 429 quota exceeded — aguarda e tenta de novo
        if '429' in err_str and tentativa <= 3:
            delay = 60
            m_delay = re.search(r'retryDelay["\s:]+(\d+)', err_str)
            if m_delay: delay = int(m_delay.group(1))
            print(f"\n  429 quota — aguardando {delay}s (tentativa {tentativa}/3)...", end=" ")
            time.sleep(delay)
            return analisar_com_gemini(lote, tentativa + 1)
        return None, f"erro:{e}"

def _analisar_offline(lote):
    ass  = (lote['assinatura'] or '').lower()
    tec  = (lote['tecnica'] or '').lower()
    tit  = (lote['titulo'] or '').lower()
    nome_ass = ''
    for pat in [r'assinado[:\s]+([A-Za-zÀ-ú\s\.]+)', r'ass[\.:\s]+([A-Za-zÀ-ú\s\.]+)']:
        m = re.search(pat, ass, re.I)
        if m:
            nome_ass = m.group(1).strip()
            break
    periodo = ''
    if any(x in tec+tit for x in ['séc. xix','século xix','1800']): periodo = 'século XIX'
    elif any(x in tec+tit for x in ['1920','1930','1940','1950']): periodo = 'início séc. XX'
    escola = ''
    if 'naif' in tec+tit: escola = 'arte naif'
    elif 'abstract' in tec+tit: escola = 'abstrato'
    elif 'paisag' in tec+tit: escola = 'paisagismo'
    return {
        'assinatura_visivel': nome_ass or 'não informado',
        'assinatura_localizacao': '',
        'tecnica_confirmada': lote['tecnica'],
        'periodo_estimado': periodo, 'escola_estilo': escola,
        'tema_descricao': lote['titulo'],
        'artistas_sugeridos': [nome_ass] if nome_ass else [],
        'justificativa': 'análise textual local (sem IA)',
        'confianca': 'baixa',
        'autenticidade': 'não avaliado',
        'assinatura_confere': 'não avaliado',
        'coerencia_estilo': 'não avaliado',
        'red_flags': 'não avaliado',
        'valor_estimado_min': 0, 'valor_estimado_max': 0,
    }

def cruzar_com_historico(artistas_sugeridos):
    melhor = None
    melhor_score = 0
    for nome in (artistas_sugeridos or []):
        k = norm(nome)
        if not k: continue
        hist = historico.get(k, [])
        if hist:
            med = float(np.median(hist))
            score = len(hist) * 10 + med * 0.001
            if score > melhor_score:
                melhor_score = score
                melhor = {'nome': nome, 'norm': k, 'n_hist': len(hist),
                          'mediana': med, 'max': max(hist), 'so_catalogo': False}
        elif k in catalogo_nomes and melhor_score == 0:
            melhor = {'nome': nome, 'norm': k, 'n_hist': 0,
                      'mediana': 0, 'max': 0, 'so_catalogo': True}
        # Match parcial por sobrenome
        palavras = k.split()
        if len(palavras) >= 2:
            for art_k, precos in historico.items():
                if all(p in art_k for p in palavras[-2:]):
                    med = float(np.median(precos))
                    score = len(precos) * 8 + med * 0.001
                    if score > melhor_score:
                        melhor_score = score
                        melhor = {'nome': nome, 'norm': art_k, 'n_hist': len(precos),
                                  'mediana': med, 'max': max(precos), 'so_catalogo': False}
    return melhor

# ── Loop principal ─────────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print(f"  ANÁLISE {'GEMINI VISION' if not args.so_texto else 'TEXTUAL'} | modelo: {MODELO}")
print(f"  {len(lotes_novos)} lotes a analisar | {DELAY_API}s entre chamadas")
print(f"{'='*65}")

# Carrega resultados do checkpoint para incluir no Excel final
resultados_checkpoint = []
for k, data in checkpoint.items():
    lote_info = next((l for l in lotes_garimpo if l['key'] == k), None)
    if lote_info:
        resultados_checkpoint.append(data)

resultados_novos = []
erros = 0

for i, lote in enumerate(lotes_novos, 1):
    tipo = "DESCON" if lote['desconhecido'] else "VALID"
    print(f"\n[{i:03d}/{len(lotes_novos)}] {tipo} | {lote['casa'][:28]} | {lote['artista_original'][:30]}", end=" ", flush=True)

    resultado, modo = analisar_com_gemini(lote)

    if modo.startswith("erro"):
        print(f"  erro: {modo[5:60]}")
        erros += 1
        time.sleep(30)   # aguarda antes do próximo após 429 esgotado
        continue

    if resultado is None:
        print("  parse falhou")
        erros += 1
        continue

    sugeridos = resultado.get('artistas_sugeridos', [])
    # Para artistas conhecidos, garante que o artista original está na lista
    if not lote['desconhecido'] and lote['artista_original']:
        if lote['artista_original'] not in sugeridos:
            sugeridos = [lote['artista_original']] + sugeridos

    match_hist = cruzar_com_historico(sugeridos)
    if not match_hist and not lote['desconhecido']:
        match_hist = cruzar_com_historico([lote['artista_original']])

    # Score de garimpo (0–100)
    score = 0
    confianca = resultado.get('confianca', 'baixa')
    if confianca == 'alta':    score += 40
    elif confianca == 'média': score += 20
    else:                      score += 5

    ass_vis = resultado.get('assinatura_visivel', '')
    _ass_neg = {'não visível', 'não informado', 'nao visivel', 'nao informado', ''}
    if ass_vis and ass_vis.lower() not in _ass_neg and len(ass_vis) > 2:
        score += 20

    autent = resultado.get('autenticidade', '')
    if autent == 'confirmada':  score += 15
    elif autent == 'duvidosa':  score -= 10
    elif autent == 'suspeita':  score -= 25

    if match_hist:
        if match_hist.get('so_catalogo'):
            score += 10
        else:
            score += 20
            if match_hist['n_hist'] >= 5: score += 10
            mult = match_hist['mediana'] / lote['preco'] if lote['preco'] > 0 else 0
            if mult >= 3: score += 15
            elif mult >= 1.5: score += 7

    if modo == "visao": score += 5
    score = max(0, min(100, score))

    row = {
        'score':             score,
        'casa':              lote['casa'],
        'data':              lote['data'],
        'tecnica':           lote['tecnica'],
        'dimensoes':         lote['dimensoes'],
        'preco_base':        lote['preco'],
        'artista_original':  lote['artista_original'],
        'modo_analise':      'identificar' if lote['desconhecido'] else 'validar',
        'ass_original':      lote['assinatura'],
        'ass_detectada':     ass_vis,
        'ass_local':         resultado.get('assinatura_localizacao',''),
        'periodo':           resultado.get('periodo_estimado',''),
        'escola':            resultado.get('escola_estilo',''),
        'tema':              resultado.get('tema_descricao',''),
        'artistas_gem':      ', '.join(sugeridos[:3]),
        'confianca':         confianca,
        'justificativa':     resultado.get('justificativa',''),
        'val_min':           resultado.get('valor_estimado_min', 0),
        'val_max':           resultado.get('valor_estimado_max', 0),
        'hist_nome':         match_hist['nome']    if match_hist else '',
        'hist_mediana':      match_hist['mediana'] if match_hist else 0,
        'hist_max':          match_hist['max']     if match_hist else 0,
        'hist_n':            match_hist['n_hist']  if match_hist else 0,
        'mult_potencial':    round(match_hist['mediana'] / lote['preco'], 1) if (match_hist and lote['preco'] > 0) else 0,
        'autenticidade':     resultado.get('autenticidade', ''),
        'assinatura_confere':resultado.get('assinatura_confere', ''),
        'coerencia_estilo':  resultado.get('coerencia_estilo', ''),
        'red_flags':         resultado.get('red_flags', ''),
        'modo':              modo,
        'url':               lote['url'],
        'foto':              lote['foto'],
        'key':               lote['key'],
    }
    resultados_novos.append(row)
    save_checkpoint(lote['key'], row)

    ass_info = f"| ass: '{ass_vis[:18]}'" if ass_vis and ass_vis.lower() not in _ass_neg else ""
    aut_info = f"| {autent}" if autent and autent != 'não avaliado' else ""
    if match_hist and not match_hist.get('so_catalogo'):
        hist_info = f"| {match_hist['nome'][:18]} R${match_hist['mediana']:,.0f}"
    else:
        hist_info = ""
    print(f"  score={score} {ass_info} {aut_info} {hist_info}")

    if USE_GEMINI and i < len(lotes_novos):
        print(f"    aguardando {DELAY_API}s...", end="\r", flush=True)
        time.sleep(DELAY_API)

# Combina novos + checkpoint para o Excel
resultados = resultados_novos + resultados_checkpoint
resultados.sort(key=lambda x: -x['score'])

print(f"\n{'='*65}")
print(f"  Novos analisados: {len(resultados_novos)}  |  Erros: {erros}")
print(f"  Total no Excel: {len(resultados)} (incluindo checkpoint)")
print(f"  Score >= 50: {sum(1 for r in resultados if r['score'] >= 50)}")
print(f"{'='*65}")

# ── Excel ──────────────────────────────────────────────────────────────────────
print("\nGerando planilha...")

C_DARK = '0F0F1A'; C_CARD = '1A1A2E'; C_ALT = '161628'
C_GOLD = 'FFD700'; C_GREEN = '22C55E'; C_BLUE = '60A5FA'
C_RED = 'EF4444'; C_AMBER = 'F59E0B'; C_GRAY = '94A3B8'; C_TEXT = 'E2E8F0'
C_BORD = '2D2D4E'; C_PURPLE = 'A855F7'; C_ORANGE = 'F97316'

def fill(c): return PatternFill('solid', fgColor=c)
def brd():
    s = Side(style='thin', color=C_BORD)
    return Border(bottom=s, right=s)
def aln(h='left'): return Alignment(horizontal=h, vertical='center', wrap_text=True)

wb = Workbook()
ws = wb.active
ws.title = 'Garimpo'
ws2 = wb.create_sheet('Resumo')

HDRS = [
    'Foto', 'Score', 'Casa', 'Data', 'Técnica', 'Dim', 'Base R$',
    'Artista Original', 'Análise',
    'Ass. Original', 'Ass. Detectada', 'Localização',
    'Período', 'Escola/Estilo', 'Tema',
    'Artistas Sugeridos (Gemini)', 'Confiança', 'Justificativa',
    'Val. Min R$', 'Val. Max R$',
    'Match Histórico', 'Med Hist R$', 'Max Hist R$', 'N Hist',
    'Múltiplo Potencial',
    'Autenticidade', 'Ass. Confere', 'Coerência', 'Red Flags',
    'Modo', 'Ver Lote',
]
N = len(HDRS)
WIDTHS = [
    14, 7, 24, 11, 20, 11, 10,
    22, 10,
    12, 18, 12,
    12, 22, 25,
    30, 10, 40,
    12, 12,
    22, 12, 12, 8,
    12,
    14, 12, 12, 30,
    7, 35,
]

THUMB_H = 80
THUMB_W = 80
ROW_H_PX = 85

# Título
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
tc = ws.cell(row=1, column=1,
    value=f'GARIMPAGEM DE LOTES  |  {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  {len(resultados)} lotes | modo: {args.modo}')
tc.font = Font(name='Calibri', size=15, bold=True, color=C_GOLD)
tc.fill = fill(C_DARK); tc.alignment = aln('center')
ws.row_dimensions[1].height = 32

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
sc2 = ws.cell(row=2, column=1,
    value=f'Análise: {"Gemini Vision + Texto" if not args.so_texto else "Textual"} | modelo: {MODELO}')
sc2.font = Font(name='Calibri', size=10, color=C_GRAY)
sc2.fill = fill(C_DARK); sc2.alignment = aln('center')
ws.row_dimensions[2].height = 16

for col, h in enumerate(HDRS, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = Font(name='Calibri', size=9, bold=True, color=C_GOLD)
    c.fill = fill(C_CARD)
    c.alignment = aln('center')
    c.border = Border(bottom=Side(style='medium', color=C_GOLD))
ws.row_dimensions[3].height = 26

for i, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A4'

CONF_COLOR  = {'alta': C_GREEN, 'média': C_AMBER, 'baixa': C_GRAY}
AUTENT_COLOR = {'confirmada': C_GREEN, 'duvidosa': C_AMBER, 'suspeita': C_RED}

def _make_thumb(foto_url):
    if not foto_url: return None
    try:
        resp = requests.get(foto_url, timeout=8)
        if resp.status_code != 200: return None
        img = PILImage.open(io.BytesIO(resp.content)).convert('RGB')
        img.thumbnail((THUMB_W, THUMB_H), PILImage.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return buf
    except: return None

for ri, row in enumerate(resultados, 4):
    bg = C_ALT if ri % 2 == 0 else C_CARD
    sc = row['score']
    sc_color = C_GREEN if sc >= 70 else (C_BLUE if sc >= 50 else (C_AMBER if sc >= 30 else C_GRAY))

    def wc(col, val, fmt=None, h='left', bold=False, color=C_TEXT, bg_=None):
        cell = ws.cell(row=ri, column=col, value=val)
        cell.font = Font(name='Calibri', size=9, bold=bold, color=color)
        cell.fill = fill(bg_ if bg_ is not None else bg)
        cell.alignment = aln(h)
        cell.border = brd()
        if fmt: cell.number_format = fmt
        return cell

    # Col 1 — Miniatura
    ws.cell(row=ri, column=1).fill = fill(bg)
    ws.cell(row=ri, column=1).border = brd()
    thumb_buf = _make_thumb(row['foto'])
    if thumb_buf:
        xl_img = XLImage(thumb_buf)
        xl_img.width  = THUMB_W
        xl_img.height = THUMB_H
        xl_img.anchor = f'A{ri}'
        ws.add_image(xl_img)

    # Col 2 — Score
    scc = ws.cell(row=ri, column=2, value=sc)
    scc.font = Font(name='Calibri', size=11, bold=True, color=sc_color)
    scc.fill = fill(bg); scc.alignment = aln('center'); scc.border = brd()

    wc(3,  row['casa'])
    wc(4,  row['data'], h='center')
    wc(5,  row['tecnica'])
    wc(6,  row['dimensoes'], h='center')
    wc(7,  row['preco_base'], fmt='R$ #,##0', h='center', bold=True, color=C_GOLD)

    # Col 8 — Artista Original
    art_c = C_ORANGE if row['modo_analise'] == 'validar' else C_GRAY
    wc(8, row['artista_original'], bold=True, color=art_c)

    # Col 9 — Modo análise
    an_label = 'identificar' if row['modo_analise'] == 'identificar' else 'validar'
    wc(9, an_label, h='center', color=C_ORANGE if row['modo_analise'] == 'validar' else C_BLUE)

    wc(10, row['ass_original'], h='center')

    ass_det = row['ass_detectada']
    ass_color = C_GREEN if (ass_det and ass_det.lower() not in {'não visível','nao visivel','não informado',''}) else C_GRAY
    wc(11, ass_det, h='center', bold=bool(ass_color == C_GREEN), color=ass_color)
    wc(12, row['ass_local'], h='center')
    wc(13, row['periodo'], h='center')
    wc(14, row['escola'])
    wc(15, row['tema'])

    art_color = C_PURPLE if row['hist_nome'] else C_TEXT
    wc(16, row['artistas_gem'], color=art_color, bold=bool(row['hist_nome']))

    conf = row['confianca']
    cc = ws.cell(row=ri, column=17, value=conf)
    cc.font = Font(name='Calibri', size=9, bold=True, color=CONF_COLOR.get(conf, C_GRAY))
    cc.fill = fill(bg); cc.alignment = aln('center'); cc.border = brd()

    wc(18, row['justificativa'])
    wc(19, row['val_min'] or None, fmt='R$ #,##0', h='center')
    wc(20, row['val_max'] or None, fmt='R$ #,##0', h='center',
       bold=True, color=C_AMBER if row['val_max'] else C_GRAY)

    hist_color = C_GREEN if row['hist_nome'] else C_GRAY
    wc(21, row['hist_nome'], bold=bool(row['hist_nome']), color=hist_color)
    wc(22, row['hist_mediana'] or None, fmt='R$ #,##0', h='center',
       color=C_GREEN if row['hist_mediana'] else C_GRAY)
    wc(23, row['hist_max'] or None, fmt='R$ #,##0', h='center',
       color=C_AMBER if row['hist_max'] else C_GRAY)
    wc(24, row['hist_n'] or None, h='center')

    mult = row['mult_potencial']
    mult_color = C_GREEN if mult >= 3 else (C_AMBER if mult >= 1.5 else C_GRAY)
    wc(25, mult or None, fmt='0.0', h='center', bold=mult >= 3, color=mult_color)

    # Autenticidade
    autent = row['autenticidade']
    ac = ws.cell(row=ri, column=26, value=autent)
    ac.font = Font(name='Calibri', size=9, bold=True,
                   color=AUTENT_COLOR.get(autent, C_GRAY))
    ac.fill = fill(bg); ac.alignment = aln('center'); ac.border = brd()

    # Assinatura confere
    ass_cf = row['assinatura_confere']
    acc = ws.cell(row=ri, column=27, value=ass_cf)
    acc.font = Font(name='Calibri', size=9,
                    color=C_GREEN if ass_cf == 'sim' else (C_RED if ass_cf == 'não' else C_AMBER))
    acc.fill = fill(bg); acc.alignment = aln('center'); acc.border = brd()

    # Coerência
    coer = row['coerencia_estilo']
    coer_c = C_GREEN if coer == 'sim' else (C_RED if coer == 'não' else C_AMBER)
    wc(28, coer, h='center', color=coer_c)

    # Red flags
    rf = row['red_flags']
    rf_c = C_RED if (rf and rf.lower() not in {'nenhum','não avaliado',''}) else C_GRAY
    wc(29, rf, color=rf_c)

    modo_color = C_BLUE if row['modo'] == 'visao' else C_GRAY
    wc(30, 'visao' if row['modo'] == 'visao' else 'texto', h='center', color=modo_color)

    url = row['url']
    uc = ws.cell(row=ri, column=31, value='Ver lote' if url else '')
    if url:
        uc.hyperlink = url
        uc.font = Font(name='Calibri', size=9, color=C_BLUE, underline='single')
    else:
        uc.font = Font(name='Calibri', size=9, color=C_GRAY)
    uc.fill = fill(bg); uc.alignment = aln('center'); uc.border = brd()

    ws.row_dimensions[ri].height = ROW_H_PX

ultima = 3 + len(resultados)
ws.auto_filter.ref = f'A3:{get_column_letter(N)}{ultima}'
ws.sheet_properties.tabColor = C_GOLD

# ABA 2 — Resumo
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
th = ws2.cell(row=1, column=1, value='RESUMO POR SCORE')
th.font = Font(name='Calibri', size=13, bold=True, color=C_GOLD)
th.fill = fill(C_DARK); th.alignment = aln('center')

for col, h in enumerate(['Faixa','Qtd','Autenticidade OK','Com Ass. Detectada','Match Histórico','Artistas Top'], 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font = Font(name='Calibri', size=10, bold=True, color=C_GOLD)
    c.fill = fill(C_CARD); c.alignment = aln('center')
    c.border = Border(bottom=Side(style='medium', color=C_GOLD))

faixas = [("Alta (70-100)", 70, 100), ("Média (50-69)", 50, 69),
          ("Baixa (30-49)", 30, 49), ("Fraca (<30)", 0, 29)]
for ri2, (label, lo, hi) in enumerate(faixas, 3):
    grupo = [r for r in resultados if lo <= r['score'] <= hi]
    c_autent = sum(1 for r in grupo if r.get('autenticidade') == 'confirmada')
    c_ass = sum(1 for r in grupo if r['ass_detectada'] and
                r['ass_detectada'].lower() not in {'não visível','nao visivel','não informado',''})
    c_hist = sum(1 for r in grupo if r['hist_nome'])
    top_arts = ', '.join(dict.fromkeys(r['hist_nome'] for r in grupo if r['hist_nome']))[:60]
    bg = C_CARD if ri2 % 2 == 0 else C_ALT
    for col, val in enumerate([label, len(grupo), c_autent, c_ass, c_hist, top_arts or '-'], 1):
        c = ws2.cell(row=ri2, column=col, value=val)
        c.font = Font(name='Calibri', size=10, color=C_TEXT)
        c.fill = fill(bg); c.alignment = aln('center' if col > 1 else 'left')
        c.border = brd()

ws2.sheet_properties.tabColor = C_PURPLE
for i, w in enumerate([18, 8, 16, 18, 15, 55], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

OUT = os.path.join(_DIR, f'garimpo_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
wb.save(OUT)
print(f"\nPlanilha salva: {os.path.basename(OUT)}")
print(f"\nTop 10 garimpos:")
for r in resultados[:10]:
    ass = f"| '{r['ass_detectada'][:18]}'" if (r['ass_detectada'] and
          r['ass_detectada'].lower() not in {'não visível','nao visivel',''}) else ""
    aut = f"| {r['autenticidade']}" if r.get('autenticidade') else ""
    hist = f"| {r['hist_nome'][:18]} R${r['hist_mediana']:,.0f}" if r['hist_nome'] else ""
    print(f"  [{r['score']:>3}] {r['artista_original'][:25]:<25} {r['tecnica'][:18]:<18} {ass} {aut} {hist}")
