#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Catálogo de Pinturas — LeilõesBR
Coleta e acumula dados de pinturas em leilão em leiloesbr.com.br
Atualiza o catálogo a cada execução sem duplicar lotes já coletados.

Saída: catalogo_pinturas_leiloesbr.xlsx  (mesma pasta do script)
"""

import sys
import io

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
import os
import json
import time
from datetime import datetime
from urllib.parse import urljoin, urlparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
try:
    from PIL import Image as PILImage
    _PIL_OK = True
except ImportError:
    _PIL_OK = False

# ── Caminhos ──────────────────────────────────────────────────────────────────
_DIR        = os.path.dirname(os.path.abspath(__file__))
DB_FILE     = os.path.join(_DIR, "leiloesbr_db.json")
OUTPUT_XLSX = os.path.join(_DIR, "catalogo_pinturas_leiloesbr.xlsx")

# ── Configurações ─────────────────────────────────────────────────────────────
BASE_URL   = "https://www.leiloesbr.com.br"
DELAY      = 1.2      # segundos entre requisições
MAX_RETRY  = 3
ITEMS_PAGE = 126      # máximo por página (21/42/84/126)
MAX_PAGES  = 30       # limite por termo

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
    "Connection":      "keep-alive",
    "Referer":         "https://www.leiloesbr.com.br/",
}

# ── Palavras-chave ─────────────────────────────────────────────────────────────
PINTURA_KW = [
    # Óleos
    "óleo sobre tela", "oleo sobre tela",
    "óleo sobre madeira", "oleo sobre madeira",
    "óleo sobre cartão", "oleo sobre cartao",
    "óleo sobre papel", "oleo sobre papel",
    "óleo sobre eucatex", "oleo sobre eucatex",
    "óleo sobre compensado", "oleo sobre compensado",
    "óleo sobre hardboard", "oleo sobre hardboard",
    "óleo sobre mdf", "oleo sobre mdf",
    "óleo s/tela", "oleo s/tela",
    "óleo s/ tela", "oleo s/ tela",
    "óleo s/madeira", "oleo s/madeira",
    "tinta a óleo", "tinta a oleo",
    # Acrílicas
    "acrílica sobre tela", "acrilica sobre tela",
    "acrílico sobre tela", "acrilico sobre tela",
    "acrílica sobre madeira", "acrilica sobre madeira",
    "acrílica sobre papel", "acrilica sobre papel",
    "acrílica sobre eucatex", "acrilica sobre eucatex",
    "acrílico sobre papel", "acrilico sobre papel",
    "tinta acrílica", "tinta acrilica",
    # Aquarela / Guache
    "aquarela sobre", "aquarela",
    "guache sobre", "guache",
    "gouache",
    # Pastel
    "pastel sobre", "pastel seco", "pastel oleoso", "pastel a seco",
    # Têmpera
    "têmpera sobre", "tempera sobre",
    "têmpera", "tempera",
    # Técnica mista
    "técnica mista sobre tela", "tecnica mista sobre tela",
    "técnica mista sobre papel", "tecnica mista sobre papel",
    "técnica mista sobre madeira", "tecnica mista sobre madeira",
    "mista sobre tela", "mista sobre papel",
    # Outros suportes / técnicas pictóricas
    "encáustica", "encaustica",
    "afresco", "afresco sobre",
    "nanquim sobre", "tinta nanquim",
    "crayon sobre",
]

GRAVURA_KW = [
    "serigrafia", "gravura", "xilogravura", "litografia", "litogravura",
    "água-forte", "agua-forte", "água-tinta", "agua-tinta",
    "monotipia", "linoleogravura", "calcografia", "buril", "ponta-seca",
    "heliogravura", "fotogravura", "serigraph", "offset", "silk screen",
    "silkscreen", "impressão", "tiragem", "exemplar",
]


# ── Utilitários ───────────────────────────────────────────────────────────────

def detect_assinatura(text: str) -> str:
    """
    Detecta informação de assinatura no texto do lote.
    Retorna: 'Assinado', 'Não assinado', 'Monogramado', 'Rubricado' ou ''.
    """
    t = text.lower()

    # Não assinado — checar primeiro para não confundir com "assinado"
    nao_ass = [
        r"n[aã]o\s+assinado", r"n[aã]o\s+assinada",
        r"n/?ass\b", r"sem\s+assinatura",
    ]
    if any(re.search(p, t) for p in nao_ass):
        return "Não assinado"

    # Monogramado
    mono = [r"monogram[ao]", r"monogr\b", r"monograma"]
    if any(re.search(p, t) for p in mono):
        return "Monogramado"

    # Rubricado
    if re.search(r"rubric[ao]", t):
        return "Rubricado"

    # Assinado (inclui abreviações e indicações de canto)
    ass = [
        r"\bass(?:inado|inada|\.)\b",         # ass. / assinado / assinada
        r"\ba\.?\s*c\.?\s*[is]\.?\s*[de]\b",  # a.c.i.d. a.c.i.e. a.c.s.d. a.c.s.e.
        r"assinatura\s+(?:do\s+)?artista",
        r"assinado\s+(?:ao|no|na|em)",
        r"\b(?:canto|c\.?i\.?d\.?|c\.?i\.?e\.?|c\.?s\.?d\.?|c\.?s\.?e\.?)\b",  # CID CIE CSD CSE
    ]
    if any(re.search(p, t) for p in ass):
        return "Assinado"

    return ""


def clean_price(s: str) -> float:
    if not s:
        return 0.0
    s = re.sub(r"[R$\s\u00a0]", "", str(s))
    s = s.replace(".", "").replace(",", ".")
    try:
        v = float(s)
        return v if v > 0 else 0.0
    except Exception:
        return 0.0


_ABREV = [
    # Padrão: (regex, substituição)  — aplicado ao texto original antes da análise
    (re.compile(r"\bO\.?S\.?T\.?\b",  re.I), "óleo sobre tela"),
    (re.compile(r"\bO\.?S\.?M\.?\b",  re.I), "óleo sobre madeira"),
    (re.compile(r"\bO\.?S\.?P\.?\b",  re.I), "óleo sobre papel"),
    (re.compile(r"\bO\.?S\.?C\.?\b",  re.I), "óleo sobre cartão"),
    (re.compile(r"\bA\.?S\.?T\.?\b",  re.I), "acrílica sobre tela"),
    (re.compile(r"\bA\.?S\.?P\.?\b",  re.I), "acrílica sobre papel"),
    (re.compile(r"\bA\.?S\.?M\.?\b",  re.I), "acrílica sobre madeira"),
    (re.compile(r"\bT\.?M\.?\b(?!\s*\d)", re.I), "técnica mista"),  # TM mas não "TM 123"
    (re.compile(r"\bAQL?\b",           re.I), "aquarela"),
    (re.compile(r"\bGCH\b",            re.I), "guache"),
]

def expand_abbreviations(text: str) -> str:
    """Expande abreviações comuns de técnicas pictóricas no texto."""
    for pattern, replacement in _ABREV:
        text = pattern.sub(replacement, text)
    return text


def is_pintura(text: str) -> bool:
    t = expand_abbreviations(text).lower()
    if any(kw in t for kw in GRAVURA_KW):
        return False
    # Termos "fortes" — presença já confirma pintura
    FORTES = [kw for kw in PINTURA_KW if kw not in ("pastel sobre", "pastel", "nanquim sobre", "tinta nanquim", "crayon sobre")]
    if any(kw in t for kw in FORTES):
        return True
    # Termos ambíguos — precisam de qualificador pictórico
    QUALIFICADORES = ("sobre tela", "sobre papel", "sobre madeira", "sobre cartão", "sobre cartao",
                      "sobre eucatex", "sobre compensado", "seco", "oleoso", "a seco",
                      "s/tela", "s/papel", "s/ tela", "s/ papel")
    AMBIGUOS = ("pastel", "nanquim", "crayon")
    for amb in AMBIGUOS:
        if amb in t:
            if any(q in t for q in QUALIFICADORES):
                return True
    return False


def extract_year(text: str) -> str:
    matches = re.findall(r"\b(1[89]\d{2}|20[0-2]\d)\b", text)
    for y in matches:
        if 1800 <= int(y) <= 2030:
            return y
    return ""


def parse_card_text(text: str) -> dict:
    """
    Extrai artista, título, técnica e dimensões do texto do card.
    Formatos comuns:
      'Artista, Título, técnica, 40x50cm, ...'
      'ARTISTA NOME (1930-2000) - Título - técnica - 40x50cm'
      'ARTISTA "Título" técnica, ...'
    """
    text = expand_abbreviations(text)
    result = {"artista": "", "titulo": "", "tecnica": "", "dimensoes": ""}

    # Dimensões (ex: 40x50cm, 100 x 80cm)
    dm = re.search(r"(\d+(?:[.,]\d+)?\s*[xX×]\s*\d+(?:[.,]\d+)?\s*(?:cm|mm)?)", text)
    if dm:
        result["dimensoes"] = dm.group(1).strip()

    # Localiza a técnica no texto
    tecnica_pos = -1
    for kw in PINTURA_KW:
        idx = text.lower().find(kw)
        if idx >= 0:
            result["tecnica"] = text[idx:idx + 60].split(",")[0].strip()
            tecnica_pos = idx
            break

    if tecnica_pos > 0:
        # Tudo antes da técnica é "artista [sep] título"
        before = text[:tecnica_pos].rstrip(" ,-–")
        if "," in before:
            parts = [p.strip() for p in before.split(",", 1)]
            result["artista"] = parts[0]
            result["titulo"]  = parts[1].rstrip(" -") if len(parts) > 1 else ""
        elif " - " in before or " – " in before:
            sep = " - " if " - " in before else " – "
            parts = [p.strip() for p in before.split(sep, 1)]
            result["artista"] = parts[0]
            result["titulo"]  = parts[1] if len(parts) > 1 else ""
        elif '"' in before:
            # Ex: ARTISTA "Título"
            result["artista"] = before[:before.index('"')].strip()
            m_q = re.search(r'"([^"]+)"', before)
            result["titulo"]  = m_q.group(1).strip() if m_q else ""
        else:
            result["artista"] = before
    else:
        # Sem técnica identificada — divide pelo primeiro separador
        if "," in text:
            parts = [p.strip() for p in text.split(",", 3)]
            result["artista"] = parts[0]
            result["titulo"]  = parts[1] if len(parts) > 1 else ""
        elif " - " in text or " – " in text:
            sep = " - " if " - " in text else " – "
            parts = [p.strip() for p in text.split(sep, 2)]
            result["artista"] = parts[0]
            result["titulo"]  = parts[1] if len(parts) > 1 else ""
        else:
            result["artista"] = text[:80]

    # Limpa
    result["artista"] = re.sub(r"\s*\.\.\.$", "", result["artista"]).strip()[:100]
    result["titulo"]  = result["titulo"].strip()[:150]
    result["tecnica"] = result["tecnica"].strip()[:100]

    return result


# ── HTTP ──────────────────────────────────────────────────────────────────────

def get(session: requests.Session, url: str, params=None) -> requests.Response | None:
    for attempt in range(1, MAX_RETRY + 1):
        try:
            r = session.get(url, params=params, timeout=20, allow_redirects=True)
            r.encoding = "utf-8"
            return r
        except requests.exceptions.Timeout:
            print(f"    [timeout t{attempt}]", end=" ", flush=True)
            time.sleep(3 * attempt)
        except requests.exceptions.ConnectionError:
            print(f"    [conexão t{attempt}]", end=" ", flush=True)
            time.sleep(5)
        except Exception as e:
            print(f"    [erro {e} t{attempt}]", end=" ", flush=True)
            time.sleep(3)
    return None


# ── Extração de cards da listagem ────────────────────────────────────────────

def extract_cards(html: str) -> list[dict]:
    """
    Extrai cards de lotes da página de busca do leiloesbr.com.br.
    Retorna lista com: lot_id, url, card_text, preco_card
    """
    soup = BeautifulSoup(html, "lxml")
    seen_ids: set[str] = set()
    cards: list[dict] = []

    for a in soup.find_all("a", href=re.compile(r"abre_catalogo", re.I)):
        href = a.get("href", "").strip()
        txt  = a.get_text(separator=" ", strip=True)
        if not href or not txt:
            continue
        lot_id = href  # href único por lote
        if lot_id in seen_ids:
            continue
        seen_ids.add(lot_id)

        full_url = href if href.startswith("http") else urljoin(BASE_URL, href)

        # Preço base do card (raramente disponível aqui, mas tenta)
        price_match = re.search(r"R\$\s*([\d.,]+)", txt)
        preco_card  = clean_price(price_match.group(1)) if price_match else 0.0

        cards.append({
            "lot_id":    lot_id,
            "url":       full_url,
            "card_text": txt,
            "preco_card": preco_card,
        })

    return cards


def search_page(session: requests.Session, term: str, op: int, page: int) -> list[dict]:
    offset = page * ITEMS_PAGE
    params = {
        "pesquisa": term,
        "op":       str(op),
        "v":        str(ITEMS_PAGE),
        "b":        str(offset),
        "gbl":      "0",
        "tp":       "|",
    }
    r = get(session, f"{BASE_URL}/busca_andamento.asp", params=params)
    if not r or r.status_code != 200:
        return []
    return extract_cards(r.text)


# Termos para busca por PALAVRA-CHAVE no andamento (op=1)
# Captura lotes em qualquer categoria que contenha esses termos no título
# — complementa a busca por categoria "Quadros" que perde lotes em outras cats.
SEARCH_TERMS_ANDAMENTO = [
    "pintura",
    "aquarela",
    "óleo",
    "acrílica",
    "gravura",
    "desenho",
    "guache",
    "pastel",
    "têmpera",
    "nanquim",
    "serigrafia",
    "litografia",
    "xilogravura",
]

SEARCH_TERMS_FINALIZADOS = [
    "óleo sobre tela",
    "acrílica sobre tela",
    "óleo sobre madeira",
    "aquarela",
    "técnica mista sobre tela",
    "guache",
    "pastel sobre",
    "têmpera",
    "OST",
    "AST",
    "TM s/tela",
]

# Casas V8.9 conhecidas — varridas diretamente para capturar lotes
# não indexados na categoria "Quadros" do agregador
V89_HOUSES = [
    # Arte / pinturas (casas especializadas)
    "http://www.galeriaartemresilva.art.br",
    "http://www.inovaarteleiloes.com.br",
    "http://www.gavealeiloes.com.br",
    "http://www.spiti.art",
    "http://www.neogaleria.com.br",
    "http://www.seculoxxleiloes.com.br",
    "http://www.comitivaarteleiloes.com.br",
    "http://www.rfleiloes.com.br",
    "http://www.toquedeclasseleiloes.com.br",
    "http://www.vilmarama.com.br",
    "http://www.antiguera.com.br",
    "http://www.bortolanleiloes.com.br",
    "http://www.anamelloleiloeira.com.br",
    "http://www.casaamarelaleiloes.net.br",
    "http://www.evanioalvesleiloeiro.com.br",
    "http://www.leiloesbrunofrancesco.com.br",
    "http://www.leiloesportaldascolecoes.com.br",
    "http://www.oficinacenarioleiloes.com.br",
    "http://br.ernanileiloeiro.com.br",
    "http://www.bruceangeirasleiloeiro.com.br",
    # Casas adicionais descobertas no catalogo.asp
    "http://br.antonioferreira.lel.br",
    "http://www.alvuraleiloes.com.br",
    "http://www.antiguidadesdomestre.com.br",
    "http://www.awleiloes.com.br",
    "http://www.barradasleiloes.com.br",
    "http://www.bastosleiloes.com.br",
    "http://www.bsecbh.com.br",
    "http://www.cidamello.lel.br",
    "http://www.flaviacardososoaresleiloes.com.br",
    "http://www.fxmcolecoes.com.br",
    "http://www.jbiglesias.com.br",
    "http://www.leiloesfp.com.br",
    "http://www.lmenezesleiloes.com.br",
    "http://www.miltrekosleiloes.com.br",
    "http://www.missleiloes.com.br",
    "http://www.novasreliquiasleiloes.com.br",
    "http://www.olhodelibraleiloes.com.br",
    "http://www.oramundileiloes.com.br",
    "http://www.petrovickleiloes.com.br",
    "http://www.plusartedecorleiloes.com.br",
    "http://www.pontodocolecionadorleiloes.com.br",
    "http://www.premiumcolecionismo.com.br",
    "http://www.prochicleiloes.com.br",
    "http://www.pvleiloes.com.br",
    "http://www.rmgouvealeiloes.com.br",
    "http://www.tallonileiloes.com.br",
    "http://www.tpleiloes.com.br",
    # Casas adicionais — apenas as NÃO indexadas no agregador LeiloesBR
    # (casas indexadas chegam via buscas por keyword/categoria, sem duplicar)
    "http://www.gondololeiloes.lel.br",
    "http://br.canvasgaleriadearte.com.br",
    "http://www.centurysarteeleiloes.com.br",
    "http://www.harpyaleiloes.com.br",
    "http://www.lagemmeleiloes.com.br",
    "http://www.marisedomingues.com.br",
    "http://br.dagsaboya.com.br",
    "http://www.ccfgaleriadearte.com.br",
    "http://www.andreadiniz.com.br",
    "http://www.levyleiloeiro.com.br",
    "http://www.onzedinheiros.lel.br",
    "http://www.jonas.lel.br",
    "http://www.dargentleiloes.net.br",
    "http://www.castrolealleiloes.com.br",
    "http://www.leilaodesign.com.br",
    "http://www.tavaresleiloes.com.br",
    "http://www.danielbastosleiloeiro.com.br",
    # Casas adicionais — descobertas em março/2026
    "http://www.antiquariatoleiloes.com.br",    # Antiquariato Leilões
    "http://www.fundodobauleiloes.com.br",      # Antiquário Fundo do Baú
    "http://www.antiquarioibiza.com.br",        # Antiquário Ibiza
    "http://www.artemaiorleiloes.com.br",       # Arte Maior Leilões
    "http://www.artemodernaleiloes.com.br",     # Arte Moderna Leilões
    "http://www.bahialeiloes.lel.br",           # Bahia Leilões
    "http://www.brazcolecao.com.br",            # Braz Coleção
    "http://www.casabrasileira.lel.br",         # Casa Brasileira
    "http://www.casalumarialeiloes.com.br",     # Casa Lu Maria
    "http://www.cohenleiloes.com.br",           # Cohen Leilões
    "http://www.danielchaiebleiloeiro.com.br",  # Daniel Chaieb
    "http://www.galpaodosleiloes.lel.br",       # Galpão dos Leilões
    "http://www.leiloesnovaera.com.br",         # Leilões Nova Era
    "http://www.linhadotempoantiguidades.com.br", # Linha do Tempo Antiguidades
    "http://www.lipemobiliario.com.br",         # Lipe Mobiliário
    "http://www.marrakechantique.com.br",       # Marrakech Antique
    "http://www.mundoemartes.com.br",           # Mundo em Artes
    "http://www.narnialeiloes.com.br",          # Nárnia Leilões
    "http://www.rtleiloes.com.br",              # RT Leilões
    "http://www.recantodasartes.com.br",        # Recanto das Artes
    "http://www.riodejaneiroleiloes.com.br",    # Rio de Janeiro Leilões
    "http://www.tremdas7.com.br",               # Robson Gini (Trem das 7)
    "http://www.sergioaltitleiloes.com.br",     # Sergio Altit
    "http://www.saobentoleiloes.com.br",        # São Bento Leilões
    "http://www.tatypaschoal.com.br",           # Taty Paschoal
    "http://www.urcaleiloes.com.br",            # Urca Leilões
    "http://www.casapaladino.com.br",           # Casa Paladino
    "http://www.leilaodocacareco.com.br",       # Leilão do Cacareco
    "http://www.reginaldodacostaleiloes.com.br", # Reginaldo da Costa
    "http://www.suyaneserraleiloes.com.br",     # Suyane Serra
    "http://www.navovotem.com.br",              # Antiquário Na Vovó
    "http://www.brexodolucio.com.br",           # Antiquário Brexo do Lúcio
    "http://www.leilaoatenas.com.br",           # Atenas Antiquário
    "http://www.betoassef.com.br",              # Beto Assef Leilões
    "http://www.conradoleiloeiro.com.br",       # Felix Conrado Leiloeiro
    "http://www.jotamleiloes.com.br",           # Imperial JM Leilões
    "http://www.antiquariodutraqueimados.com.br", # Leilões Dutra Queimados
    "http://www.paulaoantiguidades.com.br",     # Leilões Paulão Antiguidades
    "http://www.miguelsalles.com.br",           # Miguel Salles Escritório de Artes
    "http://www.vmescritarteleiloes.com.br",    # VM Escritório de Arte
    "http://www.rachelnahon.com.br",            # Velho que Vale (Rachel Nahon)
    # ── Casas confirmadas em março/2026 — faltavam na lista ──────────────────
    "http://www.shoppingdosantiquarios.lel.br", # Shopping dos Antiquários
    "http://www.claudiofirminoleiloeiro.com.br",# Claudio Firmino Leiloeiro
    "http://www.leilaodeartebrasileira.com.br", # Leilão de Arte Brasileira
    "http://www.kenyapioleiloes.com.br",        # Kenya Pio Leilões de Arte
    "http://www.errolflynnleiloes.com.br",      # Errol Flynn - Leiloeiro Público
    "http://www.paulovasconcellosleiloes.com.br",# Paulo Vasconcellos Leilões
    "http://www.arsleiloes.com.br",             # Graphos Arte - Ars Escritório de Arte
    "http://www.geracaoleilao.com.br",          # Geração Leilões
    "http://www.robertohaddad.lel.br",          # Roberto Haddad - Leiloeiro Oficial
    "http://www.planaltogaleriadearte.com.br",  # Planalto Galeria de Arte
    "http://www.rmsleiloes.com.br",             # RMS Antiguidades e Decoração
    "http://www.ciclosleiloes.com.br",          # Ciclos Leilões
    "http://www.a2arteantiguidadeleilao.com.br",# A2 Arte Antiguidade
    "http://www.albertolopesleiloeiro.com.br",  # Alberto Lopes - Leiloeiro Público
    "http://www.marteauleiloes.com.br",         # Marteau Leilões
    "http://www.subdistritoleiloes.com.br",     # Subdistrito Leilões
    "http://www.benedictleiloes.com.br",        # Benedict Leilões
    "http://www.rrdeco.com.br",                 # RR DECO Antiguidades
    "http://www.brunameloleiloeira.com.br",     # Bruna Melo - Leiloeira Oficial
    "http://www.acervoleiloes.com.br",          # Acervo Leilões
    "http://www.leiloeslemos.com.br",           # Leilões Lemos
    "http://www.marciopinho.lel.br",            # Marcio Pinho - Leiloeiro Público
    "http://www.coisaantigaleiloes.com.br",     # Coisa Antiga Leilões
    "http://www.solmarelualeiloes.com.br",      # Sol Mar e Lua Leilões
    "http://www.jfleiloeira.com.br",            # JF Leiloeira
]


def _extract_house_lot_ids(html: str, house_url: str) -> list[dict]:
    """Extrai IDs de lotes do catálogo direto de uma casa V8.9."""
    cards = []
    seen = set()
    for m in re.finditer(r'peca\.asp\?Id=(\d+)', html, re.I):
        item_id = m.group(1)
        lot_id  = f"{house_url}/peca.asp?Id={item_id}"
        if lot_id in seen:
            continue
        seen.add(lot_id)
        # Extrai texto do card adjacente (se disponível)
        start  = max(0, m.start() - 300)
        snippet = html[start: m.end() + 200]
        txt    = BeautifulSoup(snippet, "lxml").get_text(separator=" ", strip=True)
        cards.append({
            "lot_id":    lot_id,
            "url":       lot_id,
            "card_text": txt,
            "preco_card": 0.0,
        })
    return cards


def collect_houses_direct(session: requests.Session) -> list[dict]:
    """Varre o catálogo direto de cada casa V8.9 conhecida."""
    all_cards = []
    seen_ids  = set()

    for house_url in V89_HOUSES:
        r = get(session, f"{house_url}/catalogo.asp")
        if not r or r.status_code != 200:
            continue
        cards = _extract_house_lot_ids(r.text, house_url)
        novos = 0
        for c in cards:
            if c["lot_id"] not in seen_ids:
                seen_ids.add(c["lot_id"])
                all_cards.append(c)
                novos += 1
        if novos:
            print(f"  [direto] {house_url.split('//')[1]}: {novos} lotes")
        time.sleep(DELAY * 0.5)

    return all_cards


def collect_lot_ids(session: requests.Session) -> list[dict]:
    """
    Coleta todos os IDs de lotes de pinturas.
    - Andamento  : categoria 'Quadros' (captura todas as técnicas)
    - Finalizados: busca por termos (categoria não funciona para op=2 no site)
    Filtro de técnica é aplicado localmente pelo is_pintura().
    """
    all_cards: list[dict] = []
    seen_ids:  set[str]   = set()

    def add_unique(cards):
        for c in cards:
            if c["lot_id"] not in seen_ids:
                seen_ids.add(c["lot_id"])
                all_cards.append(c)

    # ── 1) Andamento — categoria Quadros (cobre todas as técnicas) ────────────
    print(f"  [Quadros andamento]", end="", flush=True)
    for page in range(MAX_PAGES):
        offset = page * ITEMS_PAGE
        params = {
            "pesquisa": "",
            "op":       "1",
            "v":        str(ITEMS_PAGE),
            "b":        str(offset),
            "gbl":      "0",
            "tp":       "|51756164726F73|",  # hex de "Quadros"
        }
        r = get(session, f"{BASE_URL}/busca_andamento.asp", params=params)
        if not r or r.status_code != 200:
            break
        cards = extract_cards(r.text)
        for c in cards:
            c["em_leilao"] = True
        add_unique(cards)
        print(f" p{page+1}={len(cards)}", end="", flush=True)
        if len(cards) < ITEMS_PAGE * 0.3:
            break
        time.sleep(DELAY * 0.5)
    print()

    # ── 2) Andamento — busca por palavra-chave (cobre TODAS as categorias) ───
    # Complementa a busca por categoria "Quadros" capturando lotes que as casas
    # indexaram em outras categorias (Arte, Decoração, etc.)
    print(f"\n  [andamento keyword]")
    for term in SEARCH_TERMS_ANDAMENTO:
        print(f"  [andamento] '{term}'", end="", flush=True)
        for page in range(MAX_PAGES):
            cards = search_page(session, term, op=1, page=page)
            for c in cards:
                c["em_leilao"] = True
            antes = len(all_cards)
            add_unique(cards)
            novos_term = len(all_cards) - antes
            print(f" p{page+1}={len(cards)}(+{novos_term})", end="", flush=True)
            if len(cards) < ITEMS_PAGE * 0.3:
                break
            time.sleep(DELAY * 0.5)
        print()

    # ── 3) Andamento — varredura direta dos catálogos das casas V8.9 ─────────
    # Captura lotes não indexados no agregador (ex: casa usa plataforma própria)
    print(f"\n  [varredura direta das casas]")
    direct_cards = collect_houses_direct(session)
    for c in direct_cards:
        c["em_leilao"] = True
    add_unique(direct_cards)
    print(f"  [direto] total novos via catálogo direto: {len(direct_cards)}")

    # ── 4) Finalizados — busca por termos (categoria não funciona para op=2) ──
    for term in SEARCH_TERMS_FINALIZADOS:
        print(f"  [finalizado] '{term}'", end="", flush=True)
        for page in range(MAX_PAGES):
            cards = search_page(session, term, op=2, page=page)
            for c in cards:
                c["em_leilao"] = False
            add_unique(cards)
            print(f" p{page+1}={len(cards)}", end="", flush=True)
            if len(cards) < ITEMS_PAGE * 0.3:
                break
            time.sleep(DELAY * 0.5)
        print()

    return all_cards


# ── Parser da página de detalhe da casa de leilão ────────────────────────────

def parse_detail(html: str, final_url: str) -> dict:
    """
    Parser genérico para páginas de casas de leilão.
    Extrai: artista, titulo, tecnica, dimensoes, ano, lance_base, maior_lance,
            num_lances, data_leilao, casa, url_detalhe.

    Dois formatos principais:
      A) Plataforma LeilõesBR V8.9 — toda a descrição do lote fica em parts[0],
         preços carregados via JavaScript (não disponíveis no HTML estático).
      B) Casas com página estática (ex: galeriaartemresilva) — informações
         distribuídas em campos separados: Artista:, Valor Inicial:, etc.
    """
    soup = BeautifulSoup(html, "lxml")
    parts = [p.strip() for p in soup.get_text(separator="|", strip=True).split("|") if p.strip()]

    # ── Detecta plataforma V8.9 (descrição completa em parts[0]) ──────────────
    # Características: parts[0] é longa, contém técnica, e pouco depois aparece
    # "Powered by LeilõesBR"
    is_v89 = (
        len(parts) > 0
        and len(parts[0]) > 30
        and is_pintura(parts[0])
        and any("V8.9" in p or "v8.9" in p or "powered by" in p.lower() for p in parts[-15:])
    )
    if is_v89:
        # Extrai da descrição em parts[0] usando o mesmo parser de card_text
        parsed = parse_card_text(parts[0])
        d = {
            "artista":     parsed.get("artista", ""),
            "titulo":      parsed.get("titulo", ""),
            "tecnica":     parsed.get("tecnica", ""),
            "dimensoes":   parsed.get("dimensoes", ""),
            "ano":         extract_year(parts[0]),
            "lance_base":  0.0,  # JS-loaded, não disponível no HTML
            "maior_lance": 0.0,
            "num_lances":  0,
            "data_leilao": "",
            "casa":        urlparse(final_url).netloc.replace("www.", ""),
            "url_detalhe": final_url,
        }
        return d

    d = {
        "artista":      "",
        "titulo":       "",
        "tecnica":      "",
        "dimensoes":    "",
        "ano":          "",
        "lance_base":   0.0,
        "maior_lance":  0.0,
        "num_lances":   0,
        "data_leilao":  "",
        "casa":         urlparse(final_url).netloc.replace("www.", ""),
        "url_detalhe":  final_url,
    }

    i = 0
    last_label = ""

    while i < len(parts):
        seg   = parts[i]
        seg_l = seg.lower().strip(": ")

        # ── Labels → próximo segmento é o valor ────────────────────
        label_map = {
            "artista":        "artista",
            "autor":          "artista",
            "técnica":        "tecnica",
            "tecnica":        "tecnica",
            "técnica/suporte":"tecnica",
            "material":       "tecnica",
            "medidas":        "dimensoes",
            "dimensões":      "dimensoes",
            "dimensoes":      "dimensoes",
            "tamanho":        "dimensoes",
            "ano":            "ano",
            "data":           "data_leilao",
            "dia do leilão":  "data_leilao",
            "dia do leilao":  "data_leilao",
            "título":         "titulo",
            "titulo":         "titulo",
        }
        field = label_map.get(seg_l)
        if field and i + 1 < len(parts):
            val = parts[i + 1].strip()
            # Skip obvious non-values
            if val and val.lower() not in ("r$", "", "|") and len(val) < 250:
                if not d.get(field):
                    d[field] = val
            last_label = field
            i += 2
            continue

        # ── Artista em maiúsculas com parênteses ────────────────────
        # Ex: "ALDEMIR MARTINS (1922 - 2006)"
        if not d["artista"] and "(" in seg and ")" in seg and len(seg) < 100:
            before = seg[:seg.index("(")].strip()
            if (before and before == before.upper() and len(before) > 3
                    and not re.search(r"lote|leil|lance|dia|oferta|zoom|tipo", seg.lower())):
                d["artista"] = seg.strip()
                i += 1
                continue

        # ── Valor Inicial / Lance base ──────────────────────────────
        # Padrão: "Valor Inicial:" | "R$" | "350,00"  (ou "R$ 8500" no mesmo seg)
        # Só captura o PRIMEIRO match (lote atual — outros lotes aparecem mais abaixo)
        if not d["lance_base"] and re.search(r"valor\s+inicial|lance\s+m[ií]nimo|pre[çc]o\s+inicial", seg, re.I):
            # Tenta extrair do mesmo segmento
            m = re.search(r"R\$\s*([\d.,]+)", seg, re.I)
            if m:
                d["lance_base"] = clean_price(m.group(1))
            else:
                # Procura nos próximos 3 segmentos: "R$" | "350,00" ou "R$ 8500"
                for j in range(i + 1, min(i + 4, len(parts))):
                    m2 = re.search(r"R\$\s*([\d.,]+)|([\d.,]{3,})", parts[j])
                    if m2:
                        val = clean_price(m2.group(1) or m2.group(2))
                        if val > 0:
                            d["lance_base"] = val
                            break
            i += 1
            continue

        # ── Histórico de lances ─────────────────────────────────────
        m_lances = re.search(r"(\d+)\s+lance", seg, re.I)
        if m_lances:
            d["num_lances"] = int(m_lances.group(1))
            i += 1
            continue

        # ── Maior lance / lance atual ───────────────────────────────
        if re.search(r"maior\s+lance|lance\s+atual|oferta\s+de|arrematado\s+por", seg, re.I):
            m = re.search(r"R\$\s*([\d.,]+)", seg, re.I)
            if m:
                val = clean_price(m.group(1))
                if val > d["maior_lance"]:
                    d["maior_lance"] = val
            else:
                for j in range(i + 1, min(i + 4, len(parts))):
                    m2 = re.search(r"R\$\s*([\d.,]+)|([\d.,]{3,})", parts[j])
                    if m2:
                        val = clean_price(m2.group(1) or m2.group(2))
                        if val > d["maior_lance"]:
                            d["maior_lance"] = val
                        break
            i += 1
            continue

        # ── Fallback: qualquer R$ XX em contexto numérico ──────────
        # (usado como lance_base se ainda não temos)
        if not d["lance_base"]:
            m = re.search(r"R\$\s*([\d.,]+)", seg, re.I)
            if m:
                val = clean_price(m.group(1))
                if val > 0:
                    d["lance_base"] = val

        i += 1

    # ── Complementa dimensões se não encontradas via label ─────────
    if not d["dimensoes"]:
        full_text = " ".join(parts)
        dm = re.search(r"(\d+(?:[.,]\d+)?\s*[xX×]\s*\d+(?:[.,]\d+)?\s*(?:cm|mm)?)", full_text)
        if dm:
            d["dimensoes"] = dm.group(1).strip()

    # ── Complementa ano se não encontrado ──────────────────────────
    if not d["ano"]:
        d["ano"] = extract_year(d["artista"])
    if not d["ano"]:
        d["ano"] = extract_year(" ".join(parts[:50]))

    # Limpa strings
    for f in ("artista", "titulo", "tecnica", "dimensoes", "ano", "data_leilao", "casa"):
        d[f] = str(d.get(f, "")).strip()[:200]

    return d


def fetch_api_v89(session: requests.Session, house_url: str, item_id: str) -> dict | None:
    """
    Chama a API JSON da plataforma LeilõesBR V8.9.
    Endpoint: {house_url}/templates/peca/asp/peca-content2.asp?ID={id}&remote=1
    Retorna dict com os campos ou None se falhar.
    """
    api_url = urljoin(house_url, "templates/peca/asp/peca-content2.asp")
    hdrs = {
        "X-Requested-With": "XMLHttpRequest",
        "Referer": urljoin(house_url, f"peca.asp?Id={item_id}"),
        "Accept": "application/json, text/javascript, */*; q=0.01",
    }
    for attempt in range(1, MAX_RETRY + 1):
        try:
            r = session.get(api_url, params={"ID": item_id, "remote": "1"},
                            headers=hdrs, timeout=20)
            if r.status_code == 404:
                return None
            # Tenta decodificar JSON
            data = json.loads(r.content.decode("utf-8", errors="replace"))
            row = data.get("data", [{}])[0] if "data" in data else data
            return row
        except Exception as e:
            if attempt < MAX_RETRY:
                time.sleep(2)
    return None


def extract_foto_url(html: str, item_id: str, final_url: str) -> str:
    """
    Extrai a URL da foto principal do lote.
    Padrão CloudFront: img_g (grande) ou img_m (médio), contendo o item_id no nome.
    """
    pat = r'(?:https?:)?//[^\s"\'<>]+/img_[gm]/\d+/' + re.escape(item_id) + r'\.(?:jpg|jpeg|png|webp)'
    matches = re.findall(pat, html, re.I)
    if not matches:
        pat2 = r'(?:https?:)?//[^\s"\'<>]+' + re.escape(item_id) + r'\.(?:jpg|jpeg|png|webp)'
        matches = re.findall(pat2, html, re.I)
    if not matches:
        return ""
    full = [("https:" + m if m.startswith("//") else m) for m in matches]
    grandes = [u for u in full if "/img_g/" in u]
    return grandes[0] if grandes else full[0]


def scrape_lot_detail(session: requests.Session, card: dict) -> dict | None:
    # Resolve a URL final (segue o redirect do leiloesbr → casa de leilão)
    r = get(session, card["url"])
    if not r:
        return None

    final_url = r.url
    parsed_url = urlparse(final_url)

    # Extrai item_id da URL final (ex: peca.asp?Id=29727106)
    id_match = re.search(r"[?&][Ii]d=(\d+)", final_url)
    item_id = id_match.group(1) if id_match else None
    house_url = f"{parsed_url.scheme}://{parsed_url.netloc}/"

    detail = {
        "artista":     "",
        "titulo":      "",
        "tecnica":     "",
        "dimensoes":   "",
        "ano":         "",
        "assinatura":  "",
        "foto_url":    "",
        "lance_base":  0.0,
        "maior_lance": 0.0,
        "num_lances":  0,
        "data_leilao": "",
        "status":      "",
        "casa":        parsed_url.netloc.replace("www.", ""),
        "url_detalhe": final_url,
    }

    # Extrai foto da página HTML
    if item_id:
        detail["foto_url"] = extract_foto_url(r.text, item_id, final_url)

    # ── Tenta a API V8.9 primeiro (retorna dados ricos em JSON) ──────────────
    api_data = None
    if item_id:
        api_data = fetch_api_v89(session, house_url, item_id)

    if api_data:
        # Descrição completa do lote
        descricao = api_data.get("PECA") or api_data.get("DESCRICAO") or ""
        parsed = parse_card_text(descricao) if descricao else {}

        detail["artista"]     = parsed.get("artista", "")
        detail["titulo"]      = parsed.get("titulo", "")
        detail["tecnica"]     = parsed.get("tecnica", "")
        detail["dimensoes"]   = parsed.get("dimensoes", "")
        detail["ano"]         = extract_year(descricao)
        detail["assinatura"]  = detect_assinatura(descricao)
        detail["data_leilao"] = api_data.get("DATADIA", "")
        detail["status"]      = api_data.get("MOSTRABTN_STATUS", "")

        # Preços: VALOR_CONTRATADO = base (int); NOVO_VALOR = lance atual (int/str)
        base_raw  = api_data.get("VALOR_CONTRATADO", 0)
        novo_raw  = api_data.get("NOVO_VALOR", 0)
        venda_raw = api_data.get("VALOR_VENDA", 0)

        detail["lance_base"]  = clean_price(str(base_raw))
        detail["num_lances"]  = int(api_data.get("QTDLANCE", 0) or 0)

        # Maior lance: se há lances, NOVO_VALOR > VALOR_CONTRATADO
        novo_val  = clean_price(str(novo_raw))
        venda_val = clean_price(str(venda_raw))
        base_val  = detail["lance_base"]

        if venda_val > 0:
            detail["maior_lance"] = venda_val   # lote vendido
        elif novo_val > base_val:
            detail["maior_lance"] = novo_val    # há lance acima da base
        elif detail["num_lances"] > 0:
            detail["maior_lance"] = novo_val    # tem lance, mesmo que igual à base

    else:
        # ── Fallback: parse HTML da página ───────────────────────────────────
        detail = parse_detail(r.text, final_url)

    # ── Complementa campos vazios com o card_text ────────────────────────────
    if card.get("card_text"):
        parsed_card = parse_card_text(card["card_text"])
        for field in ("artista", "titulo", "tecnica", "dimensoes"):
            if not detail.get(field) and parsed_card.get(field):
                detail[field] = parsed_card[field]
        if not detail.get("assinatura"):
            detail["assinatura"] = detect_assinatura(card["card_text"])

    # Tenta assinatura no HTML completo se ainda não detectou
    if not detail.get("assinatura"):
        detail["assinatura"] = detect_assinatura(r.text)

    # Lance base fallback
    if not detail["lance_base"] and card.get("preco_card"):
        detail["lance_base"] = card["preco_card"]

    # ── Descarta maior_lance suspeito: valor muito pequeno com 0 lances ──────
    # (falso positivo de parse HTML em casas sem API)
    if detail["maior_lance"] > 0 and detail["num_lances"] == 0:
        if detail["maior_lance"] < 100:
            detail["maior_lance"] = 0.0

    # ── Limpa artista: se a descrição inteira foi capturada como artista,
    # tenta extrair só o nome (antes do primeiro " - " ou parênteses com local) ─
    artista = detail.get("artista", "")
    if len(artista) > 60 and (" - " in artista or "Obra em" in artista):
        # Pega só a parte antes do primeiro " - "
        nome = artista.split(" - ")[0].strip()
        if len(nome) > 3:
            detail["artista"] = nome[:100]
            # O que sobrou pode ser o título
            resto = artista[len(nome):].lstrip(" -").strip()
            if not detail.get("titulo") and resto:
                detail["titulo"] = resto[:150]

    return detail


# ── Banco de dados local ──────────────────────────────────────────────────────

def load_db() -> dict:
    if os.path.exists(DB_FILE):
        with open(DB_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_db(db: dict):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)


def fechar_lotes_passados(db: dict) -> int:
    """Marca como em_leilao=False todos os lotes cuja data_leilao já passou."""
    from datetime import date as _date_cls
    hoje = _date_cls.today()
    fechados = 0
    for v in db.values():
        if not isinstance(v, dict): continue
        if v.get("em_leilao") is not True: continue
        data_str = v.get("data_leilao", "")
        if not data_str: continue
        try:
            partes = str(data_str).split("/")
            if len(partes) == 3:
                dt = _date_cls(int(partes[2]), int(partes[1]), int(partes[0]))
                if dt < hoje:
                    v["em_leilao"] = False
                    fechados += 1
        except Exception:
            pass
    return fechados


# ── Exportação Excel ──────────────────────────────────────────────────────────

def _load_tableau_rows():
    """Carrega e normaliza lotes do Tableau para o padrão da planilha."""
    tab_path = os.path.join(_DIR, "tableau_db.json")
    if not os.path.exists(tab_path):
        return []
    with open(tab_path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    if not isinstance(raw, list):
        return []
    out = []
    for r in raw:
        out.append({
            "artista":     r.get("artista", ""),
            "titulo":      r.get("titulo", ""),
            "tecnica":     r.get("tecnica", ""),
            "dimensoes":   r.get("medidas", ""),
            "ano":         r.get("data_obra", ""),
            "assinatura":  r.get("assinado", ""),
            "lance_base":  r.get("valor_base", 0) or 0,
            "maior_lance": r.get("lance_atual", 0) or 0,
            "num_lances":  0,
            "status":      r.get("tipo_lance", "Em leilão"),
            "data_leilao": r.get("data_leilao", ""),
            "casa":        "Tableau Arte & Leilões",
            "data_coleta": r.get("coletado_em", ""),
            "url_detalhe": r.get("url_lote", ""),
            "foto_url":    r.get("img_grande", "") or r.get("img_thumb", ""),
        })
    return out


def _parse_data_leilao(s):
    """Normaliza qualquer formato de data_leilao.
    Retorna (data_DD_MM_YYYY, horario_str, date_obj_ou_None).
    Exemplos:
      '1/4/2026'                              → ('01/04/2026', '', date(2026,4,1))
      'TERÇA FEIRA (24/03/2026) a partir de 20h' → ('24/03/2026', '20h', date(2026,3,24))
      'Ao vivo'                               → ('Ao vivo', '', None)
    """
    from datetime import date as _date_cls
    s = str(s or "").strip()
    if not s:
        return ("", "", None)
    sl = s.lower()
    if sl in ("ao vivo", "ao vivo"):
        return (s, "", None)
    horario = ""
    data_raw = s
    # Tableau: "TERÇA FEIRA (24/03/2026) a partir de 20h"
    m = re.search(r'\((\d{1,2}/\d{2}/\d{4})\)', s)
    if m:
        data_raw = m.group(1)
        hm = re.search(r'(\d{1,2}h(?:\d{2})?)', s, re.I)
        if hm:
            horario = hm.group(1).lower()
    # Normaliza D/M/YYYY → DD/MM/YYYY
    parts = data_raw.split('/')
    if len(parts) == 3:
        try:
            d, mo, y = int(parts[0]), int(parts[1]), int(parts[2])
            dt = _date_cls(y, mo, d)
            return (f"{d:02d}/{mo:02d}/{y}", horario, dt)
        except Exception:
            pass
    return (s, horario, None)


def _build_historico():
    """Carrega histórico de preços de todas as bases. Retorna dict artista_norm → [preços]."""
    import unicodedata as _ud, re as _re
    def _norm(s):
        s = _re.sub(r'\(.*?\)', '', s or '').strip()
        nfkd = _ud.normalize('NFKD', s.upper())
        s = ''.join(c for c in nfkd if not _ud.combining(c))
        return _re.sub(r'\s+', ' ', _re.sub(r'[^A-Z\s]', '', s)).strip()

    hist = {}   # norm → [preços]
    count = {}  # norm → n_lotes (para liquidez)

    for arq in ["bolsadearte_db.json", "cda_db.json"]:
        p = os.path.join(_DIR, arq)
        if not os.path.exists(p): continue
        with open(p, "r", encoding="utf-8") as f:
            d = json.load(f)
        for k, v in d.items():
            if not isinstance(v, dict): continue
            art = _norm(v.get("artista", ""))
            if not art: continue
            count[art] = count.get(art, 0) + 1
            p_ = v.get("maior_lance") or v.get("lance_atual") or 0
            try: p_ = float(p_)
            except: p_ = 0
            if p_ > 0:
                hist.setdefault(art, []).append(p_)

    return hist, count, _norm


def _calcular_scores(row, hist, count, norm_fn):
    """Calcula Liquidez (0-10) e Fator de Avaliação (0-10) para um lote."""
    art      = norm_fn(str(row.get("artista", "") or ""))
    lance    = float(row.get("maior_lance") or 0)
    base     = float(row.get("lance_base") or 0)
    n_lances = int(row.get("num_lances") or 0)
    foto     = bool(str(row.get("foto_url", "") or "").strip())
    ass      = str(row.get("assinatura", "") or "").lower()
    assinado = ass and "não" not in ass and "sem" not in ass and len(ass) > 2

    precos   = hist.get(art, [])
    n_hist   = count.get(art, 0)
    mediana  = float(sorted(precos)[len(precos)//2]) if precos else 0

    # ── Liquidez (0–10): demanda real + presença no mercado ──────────────────
    liq = 0
    if n_lances >= 10: liq += 4
    elif n_lances >= 5: liq += 3
    elif n_lances >= 2: liq += 2
    elif n_lances == 1: liq += 1
    if lance > base * 1.5: liq += 2
    elif lance > base:     liq += 1
    if n_hist >= 20: liq += 2
    elif n_hist >= 5: liq += 1
    if foto:     liq += 1
    if assinado: liq += 1
    liq = min(10, liq)

    # ── Fator de Avaliação (0–10): potencial de valor ────────────────────────
    fav = 0
    # Prêmio sobre a base
    if base > 0:
        mult = lance / base if lance > 0 else 0
        if mult >= 3:    fav += 3
        elif mult >= 2:  fav += 2
        elif mult >= 1.2: fav += 1
    # Relação lance atual vs mediana histórica
    ref = lance if lance > 0 else base
    if mediana > 0 and ref > 0:
        ratio = mediana / ref
        if ratio >= 5:   fav += 4
        elif ratio >= 3: fav += 3
        elif ratio >= 2: fav += 2
        elif ratio >= 1: fav += 1
    # Popularidade do artista no histórico
    if n_hist >= 20: fav += 2
    elif n_hist >= 5: fav += 1
    if assinado: fav += 1
    fav = min(10, fav)

    return liq, fav, mediana, n_hist


_EXCLUIR_TECNICA = {
    # Reproduções / múltiplos
    "gravura", "serigrafia", "litografia", "xilogravura", "monotipo",
    "agua-forte", "água-forte", "etching", "offset", "multiplo", "múltiplo",
    "silk", "screen print", "linoleogravura", "calcogravura",
    # Esculturas / objetos 3D
    "escultura", "bronze", "ceramica", "cerâmica", "porcelana", "vidro",
    "cristal", "marmore", "mármore", "mosaico", "tapeçaria", "tapecaria",
    # Fotografia / digital
    "fotografia", "foto", "digital", "impressão", "inkjet",
}

def _e_excluido(row: dict) -> bool:
    """Retorna True se o lote deve ser excluído (gravura, escultura, etc.)."""
    tec = (row.get("tecnica") or "").lower()
    tit = (row.get("titulo") or "").lower()
    texto = tec + " " + tit
    return any(p in texto for p in _EXCLUIR_TECNICA)


def save_excel(db: dict):
    # Filtra apenas pinturas (descarta lotes ignorados e técnicas não-pintura)
    rows = [v for v in db.values() if not v.get("_ignorado") and not _e_excluido(v)]

    # Adiciona lotes do Tableau (também filtrando)
    tab_rows = [r for r in _load_tableau_rows() if not _e_excluido(r)]
    rows = rows + tab_rows

    # Normaliza datas e extrai horário
    from datetime import date as _date_cls
    _today = _date_cls.today()

    def _norm_row(row):
        data_str, horario, dt = _parse_data_leilao(row.get("data_leilao", ""))
        r = dict(row)
        r["data_leilao"] = data_str
        r["horario"]     = horario
        r["_dt_leilao"]  = dt
        return r

    rows = [_norm_row(r) for r in rows]

    # Remove lotes cujo leilão já ocorreu
    antes = len(rows)
    rows = [r for r in rows if r["_dt_leilao"] is None or r["_dt_leilao"] >= _today]
    removidos = antes - len(rows)
    if removidos:
        print(f"  Removidos {removidos} lotes com data passada.")

    if not rows:
        print("  Nenhuma pintura para exportar.")
        return

    # Carrega histórico para scores
    hist, count, norm_fn = _build_historico()

    # ── Paleta tema escuro ─────────────────────────────────────────────────────
    C_DARK  = '0F0F1A'; C_CARD  = '1A1A2E'; C_ALT   = '161628'
    C_GOLD  = 'FFD700'; C_GREEN = '22C55E'; C_BLUE  = '60A5FA'
    C_AMBER = 'F59E0B'; C_GRAY  = '94A3B8'; C_TEXT  = 'E2E8F0'
    C_BORD  = '2D2D4E'; C_RED   = 'EF4444'

    def fill(c):  return PatternFill('solid', fgColor=c)
    def brd():
        s = Side(style='thin', color=C_BORD)
        return Border(bottom=s, right=s)
    def aln(h='left'): return Alignment(horizontal=h, vertical='center', wrap_text=True)

    # ── Preparar dados ─────────────────────────────────────────────────────────
    df = pd.DataFrame(rows)
    for c in ["artista","titulo","tecnica","dimensoes","ano","assinatura",
              "lance_base","maior_lance","num_lances","status",
              "data_leilao","horario","casa","data_coleta","url_detalhe","foto_url","em_leilao"]:
        if c not in df.columns: df[c] = ""

    num_cols = ["lance_base","maior_lance","num_lances"]
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)

    df_todos = df.copy()
    df_todos.sort_values(["artista","maior_lance"], ascending=[True,False], inplace=True)
    df_todos.reset_index(drop=True, inplace=True)

    df_lance = df_todos[df_todos["maior_lance"] > 0].copy().reset_index(drop=True)

    # Calcula scores por lote para depois agregar por artista
    scores_rows = []
    for _, row in df_todos.iterrows():
        liq, fav, mediana, _ = _calcular_scores(row.to_dict(), hist, count, norm_fn)
        scores_rows.append({"artista": row["artista"], "_liq": liq, "_fav": fav, "_med": mediana})
    df_scores = pd.DataFrame(scores_rows)

    grp = df_todos.groupby("artista").agg(
        lotes=("titulo","count"),
        base_min=("lance_base","min"),
        lance_max=("maior_lance","max"),
        lance_med=("maior_lance", lambda x: round(x[x>0].mean(),2) if (x>0).any() else 0),
    ).reset_index()
    grp_scores = df_scores.groupby("artista").agg(
        liq_med=("_liq","mean"),
        fav_med=("_fav","mean"),
        med_hist=("_med","max"),
    ).reset_index()
    grp = grp.merge(grp_scores, on="artista", how="left")
    grp.sort_values("lance_max", ascending=False, inplace=True)
    grp.reset_index(drop=True, inplace=True)

    # ── Workbook ───────────────────────────────────────────────────────────────
    wb = Workbook()

    def make_sheet(wb, name, dados, tab_color):
        """Cria aba formatada com tema escuro."""
        ws = wb.create_sheet(title=name)
        ws.sheet_properties.tabColor = tab_color

        HDRS = ["Liquidez","Fator","Artista","Título","Técnica","Dimensões","Ano","Assinatura",
                "Base R$","Lance R$","Med.Hist R$","Avaliação R$","Lances","Status","Data Leilão","Horário",
                "Casa","Coleta","Ver Lote"]
        WIDS = [10,8,28,32,22,14,8,18,12,12,13,13,8,12,14,8,26,14,35]
        N = len(HDRS)

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
        tc = ws.cell(row=1, column=1,
            value=f'CATÁLOGO DE PINTURAS — LeilõesBR + Tableau  |  {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  {len(dados)} lotes')
        tc.font = Font(name='Calibri', size=14, bold=True, color=C_GOLD)
        tc.fill = fill(C_DARK); tc.alignment = aln('center')
        ws.row_dimensions[1].height = 30

        # Cabeçalhos
        for col, h in enumerate(HDRS, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = Font(name='Calibri', size=9, bold=True, color=C_GOLD)
            c.fill = fill(C_CARD); c.alignment = aln('center')
            c.border = Border(bottom=Side(style='medium', color=C_GOLD))
        ws.row_dimensions[2].height = 24

        for i, w in enumerate(WIDS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'C3'

        # Cores para scores 0-10
        def score_color(v, high='22C55E', mid='F59E0B', low='EF4444'):
            if v >= 7: return high
            if v >= 4: return mid
            return low

        for ri, (_, row) in enumerate(dados.iterrows(), 3):
            bg = C_ALT if ri % 2 == 0 else C_CARD
            lance = float(row.get("maior_lance") or 0)
            base  = float(row.get("lance_base") or 0)

            # Scores
            liq, fav, mediana, n_hist = _calcular_scores(row.to_dict(), hist, count, norm_fn)

            def wc(col, val, fmt=None, h='left', bold=False, color=C_TEXT):
                cell = ws.cell(row=ri, column=col, value=val if val != "" else None)
                cell.font = Font(name='Calibri', size=9, bold=bold, color=color)
                cell.fill = fill(bg); cell.alignment = aln(h); cell.border = brd()
                if fmt: cell.number_format = fmt
                return cell

            # Liquidez — barra visual + número
            lc = ws.cell(row=ri, column=1, value=liq)
            lc.font = Font(name='Calibri', size=11, bold=True, color=score_color(liq))
            lc.fill = fill(bg); lc.alignment = aln('center'); lc.border = brd()

            # Fator = mediana histórica / max(lance, base)
            # Quantas vezes a avaliação de mercado supera o melhor preço atual
            ref   = max(lance, base)
            fator = round(mediana / ref, 1) if mediana > 0 and ref > 0 else None
            if fator and fator >= 50:   f_color = C_GREEN
            elif fator and fator >= 10: f_color = C_GREEN
            elif fator and fator >= 3:  f_color = C_AMBER
            elif fator and fator >= 1:  f_color = C_GRAY
            else:                       f_color = C_GRAY
            fc = ws.cell(row=ri, column=2, value=fator)
            fc.font = Font(name='Calibri', size=10, bold=bool(fator and fator >= 3), color=f_color)
            fc.fill = fill(bg); fc.alignment = aln('center'); fc.border = brd()
            if fator: fc.number_format = '0.0"x"'

            wc(3,  row.get("artista",""),      bold=True, color=C_BLUE)
            wc(4,  row.get("titulo",""))
            wc(5,  row.get("tecnica",""))
            wc(6,  row.get("dimensoes",""),    h='center')
            wc(7,  row.get("ano",""),          h='center')
            wc(8,  row.get("assinatura",""),   h='center')
            wc(9,  base  if base  > 0 else None, fmt='R$ #,##0', h='center', color=C_TEXT)
            l_color = C_GREEN if lance > 0 else C_GRAY
            wc(10, lance if lance > 0 else None, fmt='R$ #,##0', h='center', bold=lance>0, color=l_color)
            # Mediana histórica
            med_color = C_AMBER if mediana > 0 else C_GRAY
            wc(11, mediana if mediana > 0 else None, fmt='R$ #,##0', h='center', color=med_color)
            # Avaliação R$ — média histórica (average, não mediana)
            art_norm = norm_fn(str(row.get("artista","") or ""))
            precos_h = hist.get(art_norm, [])
            media_hist = round(sum(precos_h) / len(precos_h)) if precos_h else 0
            av_color = C_GREEN if media_hist > 0 else C_GRAY
            wc(12, media_hist if media_hist > 0 else None, fmt='R$ #,##0', h='center', bold=media_hist>0, color=av_color)
            wc(13, int(row.get("num_lances") or 0) or None, h='center',
               bold=bool(row.get("num_lances",0)), color=C_AMBER if row.get("num_lances",0) else C_GRAY)
            status = str(row.get("status","")).lower()
            s_color = C_GREEN if "aberto" in status else (C_AMBER if "agendado" in status else C_GRAY)
            wc(14, row.get("status",""),       h='center', color=s_color)
            wc(15, row.get("data_leilao",""),  h='center')
            wc(16, row.get("horario",""),      h='center', color=C_AMBER)
            wc(17, row.get("casa",""),         color=C_AMBER)
            wc(18, row.get("data_coleta",""),  h='center')

            url = str(row.get("url_detalhe",""))
            uc = ws.cell(row=ri, column=19, value='Ver lote' if url else '')
            if url:
                uc.hyperlink = url
                uc.font = Font(name='Calibri', size=9, color=C_BLUE, underline='single')
            else:
                uc.font = Font(name='Calibri', size=9, color=C_GRAY)
            uc.fill = fill(bg); uc.alignment = aln('center'); uc.border = brd()

            ws.row_dimensions[ri].height = 36

        ultima = 2 + len(dados)
        ws.auto_filter.ref = f'A2:{get_column_letter(N)}{ultima}'
        return ws

    def make_resumo(wb, grp):
        """Aba Resumo por Artista."""
        ws = wb.create_sheet(title='Resumo por Artista')
        ws.sheet_properties.tabColor = C_GOLD

        HDRS2 = ["Artista","Lotes","Liq. Média","Aval. Média","Base Mín R$","Lance Máx R$","Lance Médio R$","Med.Histórica R$"]
        WIDS2 = [30, 8, 12, 12, 14, 14, 14, 16]
        N2 = len(HDRS2)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N2)
        tc = ws.cell(row=1, column=1, value='RESUMO POR ARTISTA')
        tc.font = Font(name='Calibri', size=13, bold=True, color=C_GOLD)
        tc.fill = fill(C_DARK); tc.alignment = aln('center')
        ws.row_dimensions[1].height = 28

        for col, h in enumerate(HDRS2, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = Font(name='Calibri', size=9, bold=True, color=C_GOLD)
            c.fill = fill(C_CARD); c.alignment = aln('center')
            c.border = Border(bottom=Side(style='medium', color=C_GOLD))
        ws.row_dimensions[2].height = 22
        for i, w in enumerate(WIDS2, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A3'

        for ri, (_, row) in enumerate(grp.iterrows(), 3):
            bg = C_ALT if ri % 2 == 0 else C_CARD
            lance_max = float(row.get("lance_max") or 0)
            l_color = C_GREEN if lance_max > 0 else C_GRAY

            def wc2(col, val, fmt=None, h='left', bold=False, color=C_TEXT):
                cell = ws.cell(row=ri, column=col, value=val if val not in ("",0,None) else None)
                cell.font = Font(name='Calibri', size=9, bold=bold, color=color)
                cell.fill = fill(bg); cell.alignment = aln(h); cell.border = brd()
                if fmt: cell.number_format = fmt

            liq_m = round(float(row.get("liq_med") or 0), 1)
            fav_m = round(float(row.get("fav_med") or 0), 1)
            med_h = float(row.get("med_hist") or 0)

            def sc(v):
                if v >= 7: return C_GREEN
                if v >= 4: return C_AMBER
                return C_RED

            wc2(1, row.get("artista",""),         bold=True, color=C_BLUE)
            wc2(2, int(row.get("lotes",0)),       h='center', bold=True, color=C_TEXT)
            wc2(3, liq_m or None,                 fmt='0.0', h='center', bold=True, color=sc(liq_m))
            wc2(4, fav_m or None,                 fmt='0.0', h='center', bold=True, color=sc(fav_m))
            wc2(5, float(row.get("base_min",0)) or None, fmt='R$ #,##0', h='center')
            wc2(6, lance_max or None,             fmt='R$ #,##0', h='center', bold=lance_max>0, color=l_color)
            wc2(7, float(row.get("lance_med",0)) or None, fmt='R$ #,##0', h='center', color=C_AMBER)
            wc2(8, med_h or None,                 fmt='R$ #,##0', h='center', color=C_AMBER if med_h else C_GRAY)
            ws.row_dimensions[ri].height = 22

        ws.auto_filter.ref = f'A2:{get_column_letter(N2)}{2+len(grp)}'
        return ws

    def make_tableau(wb):
        """Aba dedicada ao Tableau Arte & Leilões com análise completa de oportunidades."""
        ws = wb.create_sheet(title='Tableau — Análise')
        ws.sheet_properties.tabColor = 'A855F7'  # roxo

        C_PURPLE = 'A855F7'; C_CYAN = '22D3EE'

        tab_path = os.path.join(_DIR, "tableau_db.json")
        if not os.path.exists(tab_path):
            return ws
        with open(tab_path, "r", encoding="utf-8") as f:
            raw = json.load(f)

        # Filtra gravuras e datas passadas
        from datetime import date as _date_cls
        _today = _date_cls.today()
        pinturas_tab_raw = [r for r in raw if not _e_excluido({
            "tecnica": r.get("tecnica",""), "titulo": r.get("titulo","")
        })]
        pinturas_tab = []
        for r in pinturas_tab_raw:
            data_str, horario_str, dt_obj = _parse_data_leilao(r.get("data_leilao",""))
            if dt_obj is not None and dt_obj < _today:
                continue  # leilão já encerrado
            pinturas_tab.append({**r, "_data_norm": data_str, "_horario": horario_str})

        # Calcula análise para cada lote
        analisados = []
        for r in pinturas_tab:
            art    = norm_fn(r.get("artista",""))
            precos = hist.get(art, [])
            mediana = float(sorted(precos)[len(precos)//2]) if precos else 0
            n_hist  = count.get(art, 0)
            lance   = float(r.get("lance_atual") or 0)
            base    = float(r.get("valor_base") or 0)
            ref     = max(lance, base)
            fator   = round(mediana / ref, 1) if mediana > 0 and ref > 0 else 0
            ganho   = round(mediana - ref) if mediana > ref else 0

            # Liquidez tableau: baseada em lance e histórico
            liq = 0
            if lance > base * 1.5: liq += 3
            elif lance > base:     liq += 2
            elif lance > 0:        liq += 1
            if n_hist >= 20: liq += 3
            elif n_hist >= 5: liq += 2
            elif n_hist >= 1: liq += 1
            if r.get("medidas"): liq += 1
            if r.get("data_obra"): liq += 1
            if r.get("assinado"): liq += 1
            liq = min(10, liq)

            # Classificação de oportunidade
            if fator >= 20:   classificacao = "🔥 EXCEPCIONAL"
            elif fator >= 10: classificacao = "⭐ EXCELENTE"
            elif fator >= 5:  classificacao = "✅ MUITO BOM"
            elif fator >= 2:  classificacao = "👍 BOM"
            elif fator >= 1:  classificacao = "〰 JUSTO"
            elif fator > 0:   classificacao = "⚠ ACIMA MED."
            else:             classificacao = "— SEM HIST."

            analisados.append({
                **r,
                "_mediana": mediana, "_n_hist": n_hist,
                "_fator": fator, "_ganho": ganho, "_liq": liq,
                "_class": classificacao,
            })

        analisados.sort(key=lambda x: (-x["_fator"], -x["_liq"]))

        HDRS = ["Liq","Fator","Classificação","Lote","Artista","Título","Técnica",
                "Medidas","Ano","Assinatura","Base R$","Lance Atual R$",
                "Med.Hist R$","Ganho Pot. R$","N.Hist","Data Leilão","Horário","Ver Lote"]
        WIDS = [6,8,18,6,30,32,22,14,8,16,12,14,13,14,8,14,8,35]
        N    = len(HDRS)

        # ── Cabeçalho título ──────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
        tc = ws.cell(row=1, column=1,
            value=f'TABLEAU ARTE & LEILÕES  |  Análise de Oportunidades  |  {datetime.now().strftime("%d/%m/%Y %H:%M")}')
        tc.font = Font(name='Calibri', size=15, bold=True, color='A855F7')
        tc.fill = fill(C_DARK); tc.alignment = aln('center')
        ws.row_dimensions[1].height = 32

        # Subtítulo — datas únicas dos lotes filtrados
        datas_unicas = sorted(set(r["_data_norm"] for r in pinturas_tab if r["_data_norm"]))
        datas_txt = "  |  ".join(datas_unicas) if datas_unicas else "sem data"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
        sc = ws.cell(row=2, column=1,
            value=f'{datas_txt}  |  {len(analisados)} pinturas  |  {sum(1 for r in analisados if r["_fator"]>0)} com histórico de preços')
        sc.font = Font(name='Calibri', size=10, color=C_GRAY)
        sc.fill = fill(C_DARK); sc.alignment = aln('center')
        ws.row_dimensions[2].height = 16

        # ── Cabeçalhos ────────────────────────────────────────────────────────
        for col, h in enumerate(HDRS, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(name='Calibri', size=9, bold=True, color='A855F7')
            c.fill = fill(C_CARD); c.alignment = aln('center')
            c.border = Border(bottom=Side(style='medium', color='A855F7'))
        ws.row_dimensions[3].height = 24
        for i, w in enumerate(WIDS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'E4'

        # Faixas de cor para classificação
        CLASS_COLOR = {
            "🔥 EXCEPCIONAL": C_GREEN,
            "⭐ EXCELENTE":   C_GREEN,
            "✅ MUITO BOM":   C_AMBER,
            "👍 BOM":         C_AMBER,
            "〰 JUSTO":       C_GRAY,
            "⚠ ACIMA MED.":  C_RED,
            "— SEM HIST.":    C_GRAY,
        }

        for ri, r in enumerate(analisados, 4):
            bg = C_ALT if ri % 2 == 0 else C_CARD
            fator   = r["_fator"]
            ganho   = r["_ganho"]
            mediana = r["_mediana"]
            liq     = r["_liq"]
            lance   = float(r.get("lance_atual") or 0)
            base    = float(r.get("valor_base") or 0)
            cls     = r["_class"]

            def wt(col, val, fmt=None, h='left', bold=False, color=C_TEXT):
                cell = ws.cell(row=ri, column=col, value=val if val not in ("","",None) else None)
                cell.font = Font(name='Calibri', size=9, bold=bold, color=color)
                cell.fill = fill(bg); cell.alignment = aln(h); cell.border = brd()
                if fmt: cell.number_format = fmt

            # Liquidez
            lc = ws.cell(row=ri, column=1, value=liq)
            l_c = C_GREEN if liq>=7 else (C_AMBER if liq>=4 else C_GRAY)
            lc.font = Font(name='Calibri', size=11, bold=True, color=l_c)
            lc.fill = fill(bg); lc.alignment = aln('center'); lc.border = brd()

            # Fator
            fc = ws.cell(row=ri, column=2, value=fator if fator > 0 else None)
            f_c = C_GREEN if fator>=10 else (C_AMBER if fator>=2 else (C_GRAY if fator>0 else C_GRAY))
            fc.font = Font(name='Calibri', size=10, bold=fator>=2, color=f_c)
            fc.fill = fill(bg); fc.alignment = aln('center'); fc.border = brd()
            if fator > 0: fc.number_format = '0.0"x"'

            # Classificação — destaque especial
            cc = ws.cell(row=ri, column=3, value=cls)
            cc.font = Font(name='Calibri', size=9, bold="EXCEPCIONAL" in cls or "EXCELENTE" in cls,
                           color=CLASS_COLOR.get(cls, C_GRAY))
            cc.fill = fill(bg); cc.alignment = aln('center'); cc.border = brd()

            wt(4,  r.get("lote_num"),           h='center', color=C_GRAY)
            wt(5,  r.get("artista",""),          bold=True, color=C_CYAN)
            wt(6,  r.get("titulo",""))
            wt(7,  r.get("tecnica",""))
            wt(8,  r.get("medidas",""),          h='center')
            wt(9,  r.get("data_obra",""),        h='center')
            wt(10, r.get("assinado",""),         h='center')
            wt(11, base  if base>0  else None,   fmt='R$ #,##0', h='center')
            # Lance — verde se subiu acima da base
            l_color2 = C_GREEN if lance > base else (C_AMBER if lance > 0 else C_GRAY)
            lt = ws.cell(row=ri, column=12, value=lance if lance>0 else None)
            lt.font = Font(name='Calibri', size=9, bold=lance>base, color=l_color2)
            lt.fill = fill(bg); lt.alignment = aln('center'); lt.border = brd()
            if lance > 0: lt.number_format = 'R$ #,##0'

            # Mediana histórica
            mt = ws.cell(row=ri, column=13, value=mediana if mediana>0 else None)
            mt.font = Font(name='Calibri', size=9, bold=mediana>0, color=C_PURPLE if mediana>0 else C_GRAY)
            mt.fill = fill(bg); mt.alignment = aln('center'); mt.border = brd()
            if mediana > 0: mt.number_format = 'R$ #,##0'

            # Ganho potencial — destaque forte
            gt = ws.cell(row=ri, column=14, value=ganho if ganho>0 else None)
            g_c = C_GREEN if ganho > 10000 else (C_AMBER if ganho > 1000 else C_GRAY)
            gt.font = Font(name='Calibri', size=9, bold=ganho>5000, color=g_c)
            gt.fill = fill(bg); gt.alignment = aln('center'); gt.border = brd()
            if ganho > 0: gt.number_format = 'R$ #,##0'

            wt(15, r["_n_hist"] if r["_n_hist"] > 0 else None, h='center', color=C_GRAY)
            wt(16, r.get("_data_norm","") or r.get("data_leilao",""), h='center', color=C_AMBER)
            wt(17, r.get("_horario",""), h='center', color=C_CYAN)

            # URL hyperlink
            url = r.get("url_lote","")
            uc = ws.cell(row=ri, column=18, value='Ver lote' if url else '')
            if url:
                uc.hyperlink = url
                uc.font = Font(name='Calibri', size=9, color=C_BLUE, underline='single')
            else:
                uc.font = Font(name='Calibri', size=9, color=C_GRAY)
            uc.fill = fill(bg); uc.alignment = aln('center'); uc.border = brd()

            ws.row_dimensions[ri].height = 38

        ultima = 3 + len(analisados)
        ws.auto_filter.ref = f'A3:{get_column_letter(N)}{ultima}'
        return ws

    def make_hoje(wb):
        """Aba com todos os lotes de hoje — todas as plataformas via leiloesbr_db.
        Data calculada dinamicamente. Incorpora referências de preço para artistas sem histórico."""
        from datetime import date as _date_cls
        _hoje = _date_cls.today()
        _hoje_lbr  = f"{_hoje.day}/{_hoje.month}/{_hoje.year}"       # "25/3/2026"
        _hoje_lbr2 = f"{_hoje.day:02d}/{_hoje.month:02d}/{_hoje.year}"  # "25/03/2026"
        _hoje_label = _hoje.strftime("%d-%m")                         # "25-03"

        # ── Referências de mercado pesquisadas (artistas sem registro de preço no DB) ──
        REF_MANUAL = {
            # Robertohaddad / MundoEmArtes 25/03
            "JOAO BATISTA DA COSTA":        (85000, "Impressionismo BR; record Sotheby's ≈ R$180k; acervo em museus nacionais"),
            "JOÃO BATISTA DA COSTA":        (85000, "Impressionismo BR; record Sotheby's ≈ R$180k; acervo em museus nacionais"),
            "JACQUEZ DOUCHEZ":              (18000, "Pintor francês (1921-2012); leilões Drouot/Tajan; estimativa MutualArt"),
            "GLAUCO RODRIGUES":             (15000, "Pop art BR; record Sotheby's ≈ R$25k; obra '6 de abril de 1500' icônica"),
            "MAURICE BOMPARD":              (12000, "Orientalista francês (1857-1936); obras Christie's/Sotheby's"),
            "MARCIER":                      (5000,  "Emeric Marcier (1916-1990); artista afro-venezuelano; coleções públicas BR"),
            "MIRABEAU MENEZES":             (4500,  "Artista BR contemporâneo (1964); estimativa mercado"),
            "HARRY ELSAS":                  (4500,  "Artista BR (1925-1994); poucos registros públicos; estimativa mercado"),
            "STEF WENIJS":                  (5000,  "Artista holandês; estimativa mercado"),
            "NEWTON REZENDE":               (2800,  "Artista BR (1912); poucos registros públicos; estimativa mercado"),
            "SYLVIO PINTO":                 (3500,  "Artista BR (1918); estimativa mercado"),
            "ANNE MARIE NIVOULI":           (3500,  "Pintora francesa (1879); estimativa mercado"),
            "ANNE MARIE NIVOULIÈS":         (3500,  "Pintora francesa (1879); estimativa mercado"),
            "ANNE MARIE NIVOULIES":         (3500,  "Pintora francesa (1879); estimativa mercado"),
            "ANNE MARIE NIVOULIÈS DE PIERREFORT": (3500, "Pintora francesa (1879); estimativa mercado"),
            "EDUARDO LEON GARRIDO":         (4000,  "Escola espanhola (1856-1949); obras Christie's; estimativa"),
            "CHICO DA SILVA":               (34000, "Francisco da Silva (1910-1985); record Sotheby's 2023 USD 330k"),
            "FRANCISCO DA SILVA":           (34000, "Record Sotheby's 2023 USD 330k; média 12m USD 6.680 (~R$38k)"),
            # Tableau 25/03 — artistas sem histórico DB
            "EMILIANO DI CAVALCANTI":       (45000, "Modernismo BR; record R$1.2M; lanceado R$21k Tableau hoje"),
            "DI CAVALCANTI":                (45000, "Emiliano di Cavalcanti; modernismo BR; record R$1.2M"),
            "HEITOR DOS PRAZERES":          (55000, "Modernismo BR (1898-1966); record ≈ R$800k; lanceado R$34.500 hoje"),
            "HENRI MATISSE":                (250000,"Mestre impressionismo FR (1869-1954); lanceado R$8.600 hoje"),
            "ISMAEL NERY":                  (30000, "Surrealismo BR (1900-1934); record ≈ R$180k"),
            "ALBERTO DA VEIGA GUIGNARD":    (40000, "Modernismo BR (1896-1962); record ≈ R$350k"),
            "GUIGNARD":                     (40000, "Alberto da Veiga Guignard; modernismo BR; record ≈ R$350k"),
            "ALDO BONADEI":                 (15000, "Modernismo BR (1906-1974); lanceado R$4.800 Tableau; record ≈ R$80k"),
            "DJANIRA DA MOTTA E SILVA":     (20000, "Djanira (1914-1979); record ≈ R$180k"),
            "CARYBÉ":                       (12000, "Carybé (1911-1997); modernismo BR; record ≈ R$80k"),
            "CARYB":                        (12000, "Carybé (1911-1997); modernismo BR; record ≈ R$80k"),
            "ALEJANDRO OTERO":              (8000,  "Artista venezuelano (1921-1990); obras MoMA; estimativa"),
            "BUSTAMANTE SA":                (4000,  "Artista BR (1907-1988); estimativa baseada em lances"),
            "TADASHI KAMINAGAI":            (5000,  "Artista japonês-BR (1899-1982); estimativa mercado"),
            "DARCY PENTEADO":               (5000,  "Artista BR (1926-1987); estimativa mercado"),
            "FERREIRA GULLAR":              (8000,  "Poeta/artista BR (1930-2016); estimativa mercado"),
            "GERSON DE SOUZA":              (8000,  "Artista BR (1926-2008); estimativa baseada em lances Tableau"),
            "GERARDO DE SOUSA":             (6000,  "Artista; estimativa baseada em lances Tableau"),
            "LEONOR FINI":                  (15000, "Surrealismo (1907-1996); obras Christie's; estimativa"),
            "ANTONIO POTEIRO":              (5000,  "Artista BR (1925-2010); estimativa mercado"),
            "ALDEMIR MARTINS":              (4500,  "Artista BR (1922-2006); record ≈ R$30k; presença frequente em leilão"),
            "INOS CORRADIN":                (6500,  "Faleceu set/2025 → alta de preços; record USD 2.036"),
            "DURVAL PEREIRA":               (3500,  "Record USD 2.062; 80×120cm ≈ R$5.800; premiado internacionalmente"),
        }

        ws = wb.create_sheet(title='Em Leilão Agora')
        ws.sheet_properties.tabColor = 'F97316'  # laranja
        C_ORANGE = 'F97316'; C_CYAN = '22D3EE'; C_PURPLE = 'A855F7'

        # Lê do leiloesbr_db.json (agrega todas as plataformas)
        lbr_path = os.path.join(_DIR, "leiloesbr_db.json")
        if not os.path.exists(lbr_path): return ws
        with open(lbr_path, "r", encoding="utf-8") as f:
            raw_lbr = json.load(f)
        raw_lbr_list = list(raw_lbr.values()) if isinstance(raw_lbr, dict) else raw_lbr

        THUMB_W, THUMB_H = 90, 90

        def _make_thumb(url):
            if not url or not _PIL_OK: return None
            try:
                resp = requests.get(url, timeout=6)
                if resp.status_code != 200: return None
                img = PILImage.open(io.BytesIO(resp.content)).convert('RGB')
                img.thumbnail((THUMB_W, THUMB_H), PILImage.LANCZOS)
                buf = io.BytesIO()
                img.save(buf, format='PNG')
                buf.seek(0)
                return buf
            except: return None

        # Normaliza campos LBR → nomes internos usados na análise
        def _norm_rec(r):
            data_raw = r.get("data_leilao", "")
            data_norm, horario_norm, dt_obj = _parse_data_leilao(data_raw)
            return {
                "artista":    r.get("artista", ""),
                "titulo":     r.get("titulo", ""),
                "tecnica":    r.get("tecnica", ""),
                "medidas":    r.get("dimensoes", ""),
                "data_obra":  r.get("ano", ""),
                "assinado":   "",
                "valor_base": float(r.get("lance_base") or 0),
                "lance_atual":float(r.get("maior_lance") or 0),
                "url_lote":   r.get("url_detalhe", ""),
                "foto_url":   r.get("foto_url", ""),
                "casa":       r.get("casa", ""),
                "lote_num":   None,
                "data_leilao": data_norm,
                "_dt":        dt_obj,
            }

        # Mostra todos os lotes ativos (em_leilao=True), não só os de hoje
        from datetime import date as _date_lbr
        lotes_hoje = [_norm_rec(r) for r in raw_lbr_list
                      if r.get('em_leilao', True)
                      and not _e_excluido({"tecnica": r.get("tecnica",""),
                                           "titulo": r.get("titulo","")})]
        # Remove lotes com data já passada
        lotes_hoje = [r for r in lotes_hoje
                      if r["_dt"] is None or r["_dt"] >= _date_lbr.today()]

        # Calcula análise com enriquecimento manual
        def _get_mediana(art_norm):
            precos = hist.get(art_norm, [])
            if precos:
                return float(sorted(precos)[len(precos)//2]), count.get(art_norm, 0), "DB"
            for nome_ref, (preco_ref, _) in REF_MANUAL.items():
                if norm_fn(nome_ref) == art_norm:
                    return float(preco_ref), -1, "PESQ"
            return 0.0, count.get(art_norm, 0), "—"

        analisados = []
        for r in lotes_hoje:
            art_n  = norm_fn(r.get("artista",""))
            med, n_hist, fonte = _get_mediana(art_n)
            lance  = float(r.get("lance_atual") or 0)
            base   = float(r.get("valor_base") or 0)
            ref    = max(lance, base)
            fator  = round(med / ref, 1) if med > 0 and ref > 0 else (99.9 if med > 0 else 0)
            ganho  = round(med - ref) if med > ref else 0

            liq = 0
            if lance > base * 1.5: liq += 3
            elif lance > base:     liq += 2
            elif lance > 0:        liq += 1
            if n_hist >= 20:   liq += 3
            elif n_hist >= 5:  liq += 2
            elif n_hist == -1: liq += 2   # referência pesquisada
            elif n_hist >= 1:  liq += 1
            if r.get("medidas"):   liq += 1
            if r.get("data_obra"): liq += 1
            liq = min(10, liq)

            if fator >= 20:            cls = "🔥 EXCEPCIONAL"
            elif fator >= 10:          cls = "⭐ EXCELENTE"
            elif fator >= 5:           cls = "✅ MUITO BOM"
            elif fator >= 2:           cls = "👍 BOM"
            elif fator >= 1:           cls = "〰 JUSTO"
            elif fator > 0:            cls = "⚠ ACIMA MED."
            elif med > 0 and ref == 0: cls = "💡 SEM LANCE"
            else:                      cls = "— SEM REF."

            nota_ref = ""
            for nome_ref, (_, nota) in REF_MANUAL.items():
                if norm_fn(nome_ref) == art_n:
                    nota_ref = nota; break

            precos_h = hist.get(art_n, [])
            media_hist = round(sum(precos_h) / len(precos_h)) if precos_h else 0

            analisados.append({**r,
                "_med": med, "_n_hist": n_hist, "_fonte": fonte,
                "_fator": fator, "_ganho": ganho, "_liq": liq,
                "_cls": cls, "_nota_ref": nota_ref,
                "_media_hist": media_hist,
            })

        from datetime import date as _date_sort
        analisados.sort(key=lambda x: (
            x["_dt"] or _date_sort(9999, 12, 31),
            -(x["_fator"] if x["_fator"] < 99 else 200),
            -x["_liq"],
        ))

        HDRS = ["Foto","Liq","Fator","Classificação","Data Leilão","Casa","Artista","Título","Técnica",
                "Medidas","Ano","Base R$","Lance Atual","Avaliação R$","Ref.Mercado R$",
                "Ganho Pot. R$","Fonte","N.Hist","Nota Referência","Ver Lote"]
        WIDS = [14,6,8,18,14,24,30,30,22,14,7,11,11,14,14,14,7,7,42,32]
        N    = len(HDRS)

        # ── Título ────────────────────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
        tc = ws.cell(row=1, column=1,
            value=f'EM LEILÃO AGORA  |  Todas as datas  |  gerado {_hoje.strftime("%d/%m/%Y")}  |  {len(analisados)} pinturas/obras analisadas')
        tc.font = Font(name='Calibri', size=15, bold=True, color=C_ORANGE)
        tc.fill = fill(C_DARK); tc.alignment = aln('center')
        ws.row_dimensions[1].height = 34

        c_exc = sum(1 for r in analisados if "EXCEPCIONAL" in r["_cls"] or "EXCELENTE" in r["_cls"])
        c_bom = sum(1 for r in analisados if "MUITO BOM" in r["_cls"] or "BOM" in r["_cls"])
        c_sem = sum(1 for r in analisados if "SEM REF" in r["_cls"])
        c_pesq= sum(1 for r in analisados if r["_fonte"] == "PESQ")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
        sc = ws.cell(row=2, column=1,
            value=f'🔥⭐ {c_exc} excepcionais/excelentes  |  👍✅ {c_bom} boas oportunidades  |  🔍 {c_pesq} preços pesquisados  |  — {c_sem} sem referência')
        sc.font = Font(name='Calibri', size=10, bold=True, color=C_ORANGE)
        sc.fill = fill(C_DARK); sc.alignment = aln('center')
        ws.row_dimensions[2].height = 18

        # ── Cabeçalhos ────────────────────────────────────────────────────────────
        for col, h in enumerate(HDRS, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font = Font(name='Calibri', size=9, bold=True, color=C_ORANGE)
            c.fill = fill(C_CARD); c.alignment = aln('center')
            c.border = Border(bottom=Side(style='medium', color=C_ORANGE))
        ws.row_dimensions[3].height = 22
        for i, w in enumerate(WIDS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'G4'

        CLASS_CLR = {
            "🔥 EXCEPCIONAL": C_GREEN, "⭐ EXCELENTE": C_GREEN,
            "✅ MUITO BOM":   C_AMBER, "👍 BOM":        C_AMBER,
            "〰 JUSTO":       C_GRAY,  "⚠ ACIMA MED.": C_RED,
            "💡 SEM LANCE":   C_CYAN,  "— SEM REF.":    C_GRAY,
        }

        for ri, r in enumerate(analisados, 4):
            bg = C_ALT if ri % 2 == 0 else C_CARD
            fator  = r["_fator"]
            ganho  = r["_ganho"]
            med    = r["_med"]
            liq    = r["_liq"]
            lance  = float(r.get("lance_atual") or 0)
            base   = float(r.get("valor_base") or 0)
            cls    = r["_cls"]
            fonte  = r["_fonte"]
            is_star = "EXCEPCIONAL" in cls or "EXCELENTE" in cls

            def wh(col, val, fmt=None, h='left', bold=False, color=C_TEXT):
                cell = ws.cell(row=ri, column=col, value=val if val not in ("", None) else None)
                cell.font = Font(name='Calibri', size=9, bold=bold, color=color)
                cell.fill = fill(bg); cell.alignment = aln(h); cell.border = brd()
                if fmt: cell.number_format = fmt

            # Col 1 — Miniatura
            ws.cell(row=ri, column=1).fill = fill(bg)
            ws.cell(row=ri, column=1).border = brd()
            thumb = _make_thumb(r.get("foto_url",""))
            if thumb:
                xl_img = XLImage(thumb)
                xl_img.width = THUMB_W; xl_img.height = THUMB_H
                xl_img.anchor = f'A{ri}'
                ws.add_image(xl_img)

            # Col 2 — Liquidez
            lc = ws.cell(row=ri, column=2, value=liq)
            l_c = C_GREEN if liq>=7 else (C_AMBER if liq>=4 else C_GRAY)
            lc.font = Font(name='Calibri', size=11, bold=True, color=l_c)
            lc.fill = fill(bg); lc.alignment = aln('center'); lc.border = brd()

            # Col 3 — Fator
            fator_disp = fator if fator < 99 else None
            f_txt = "∞" if fator >= 99 and med > 0 else None
            fc = ws.cell(row=ri, column=3, value=fator_disp if not f_txt else f_txt)
            f_c = C_GREEN if fator>=10 or fator>=99 else (C_AMBER if fator>=2 else C_GRAY)
            fc.font = Font(name='Calibri', size=10, bold=fator>=2 or fator>=99, color=f_c)
            fc.fill = fill(bg); fc.alignment = aln('center'); fc.border = brd()
            if fator_disp and fator_disp > 0: fc.number_format = '0.0"x"'

            # Col 4 — Classificação
            cc = ws.cell(row=ri, column=4, value=cls)
            cc.font = Font(name='Calibri', size=9, bold=is_star, color=CLASS_CLR.get(cls, C_GRAY))
            cc.fill = fill(bg); cc.alignment = aln('center'); cc.border = brd()

            # Col 5 — Data Leilão (filtrável)
            dt_c = C_AMBER if r.get("data_leilao") else C_GRAY
            wh(5,  r.get("data_leilao",""),     h='center', bold=True, color=dt_c)

            # Col 6 — Casa
            casa_str = r.get("casa","")
            casa_str = casa_str.replace(".com.br","").replace(".lel.br","").replace(".br","")
            wh(6,  casa_str,                    h='center', color=C_PURPLE)
            wh(7,  r.get("artista",""),         bold=True, color=C_CYAN if not is_star else C_GREEN)
            wh(8,  r.get("titulo",""))
            wh(9,  r.get("tecnica",""))
            wh(10, r.get("medidas",""),         h='center')
            wh(11, r.get("data_obra",""),       h='center')
            wh(12, base  if base  > 0 else None, fmt='R$ #,##0', h='center')

            # Col 13 — Lance atual
            l2c = C_GREEN if lance > base else (C_AMBER if lance > 0 else C_GRAY)
            lt = ws.cell(row=ri, column=13, value=lance if lance>0 else None)
            lt.font = Font(name='Calibri', size=9, bold=lance>base, color=l2c)
            lt.fill = fill(bg); lt.alignment = aln('center'); lt.border = brd()
            if lance>0: lt.number_format = 'R$ #,##0'

            # Col 14 — Avaliação R$ (média histórica do artista)
            media_h = r.get("_media_hist", 0)
            av_c = C_GREEN if media_h > 0 else C_GRAY
            avt = ws.cell(row=ri, column=14, value=media_h if media_h > 0 else None)
            avt.font = Font(name='Calibri', size=9, bold=media_h>0, color=av_c)
            avt.fill = fill(bg); avt.alignment = aln('center'); avt.border = brd()
            if media_h>0: avt.number_format = 'R$ #,##0'

            # Col 15 — Ref. Mercado (mediana/pesquisa)
            m_c = C_GREEN if fonte=="DB" else (C_CYAN if fonte=="PESQ" else C_GRAY)
            mt = ws.cell(row=ri, column=15, value=med if med>0 else None)
            mt.font = Font(name='Calibri', size=9, bold=med>0, color=m_c)
            mt.fill = fill(bg); mt.alignment = aln('center'); mt.border = brd()
            if med>0: mt.number_format = 'R$ #,##0'

            # Col 16 — Ganho potencial
            gt = ws.cell(row=ri, column=16, value=ganho if ganho>0 else None)
            g_c = C_GREEN if ganho>10000 else (C_AMBER if ganho>1000 else C_GRAY)
            gt.font = Font(name='Calibri', size=9, bold=ganho>5000, color=g_c)
            gt.fill = fill(bg); gt.alignment = aln('center'); gt.border = brd()
            if ganho>0: gt.number_format = 'R$ #,##0'

            src_c = C_GREEN if fonte=="DB" else (C_CYAN if fonte=="PESQ" else C_GRAY)
            wh(17, fonte, h='center', bold=True, color=src_c)
            wh(18, r["_n_hist"] if r["_n_hist"] > 0 else None, h='center', color=C_GRAY)
            wh(19, r["_nota_ref"] if r["_nota_ref"] else None, color=C_AMBER if fonte=="PESQ" else C_GRAY)

            # Col 20 — URL
            url = r.get("url_lote","")
            uc = ws.cell(row=ri, column=20, value='Ver lote' if url else '')
            if url:
                uc.hyperlink = url
                uc.font = Font(name='Calibri', size=9, color=C_BLUE, underline='single')
            else:
                uc.font = Font(name='Calibri', size=9, color=C_GRAY)
            uc.fill = fill(bg); uc.alignment = aln('center'); uc.border = brd()

            ws.row_dimensions[ri].height = THUMB_H + 8

        ultima = 3 + len(analisados)
        ws.auto_filter.ref = f'A3:{get_column_letter(N)}{ultima}'
        return ws

    # Uma única aba com tudo
    del wb['Sheet']
    make_sheet(wb, 'Catálogo', df_todos, C_GOLD)
    make_tableau(wb)
    make_hoje(wb)

    print(f"\n  Salvando Excel: {OUTPUT_XLSX}")
    wb.save(OUTPUT_XLSX)
    print(f"  [OK] {len(df_todos)} pinturas salvas (com lance: {len(df_todos[df_todos['maior_lance']>0])}) → {OUTPUT_XLSX}")


# ── Tableau Arte & Leilões ────────────────────────────────────────────────────

TABLEAU_BASE = "https://www.tableau.com.br"
TABLEAU_LOT  = f"{TABLEAU_BASE}/leilao/lote.php"
TABLEAU_IMG  = f"{TABLEAU_BASE}/leilao/"          # + {work_id}g.jpg
TABLEAU_MAX  = 600   # limite superior de lotes por varredura


def _norm_tab(s: str) -> str:
    """Normaliza string para comparar labels no HTML do Tableau."""
    import unicodedata
    nfkd = unicodedata.normalize("NFKD", s.lower())
    return "".join(c for c in nfkd if not unicodedata.combining(c)).rstrip(":").strip()


_TAB_LABEL_MAP = {
    "titulo":     "titulo",
    "tecnica":    "tecnica",
    "medidas":    "dimensoes",
    "tiragem":    "edicao",
    "edicao":     "edicao",
    "data/local": "ano",
    "data":       "ano",
}


def parse_tableau_lot(html: str, lot_num: int) -> dict | None:
    """
    Extrai dados de um lote do Tableau Arte & Leilões.
    Retorna None se a página não contiver lote válido.
    """
    soup = BeautifulSoup(html, "lxml")

    # Verifica se há conteúdo real de lote
    full_text = soup.get_text(separator="|", strip=True)
    if "lote" not in full_text.lower() or len(full_text) < 100:
        return None

    data = {
        "artista":    "",
        "titulo":     "",
        "tecnica":    "",
        "dimensoes":  "",
        "ano":        "",
        "edicao":     "",
        "assinatura": "",
        "foto_url":   "",
        "lance_base": 0.0,
        "maior_lance": 0.0,
        "num_lances": 0,
        "data_leilao": "",
        "status":     "",
        "casa":       "tableau.com.br",
        "url_detalhe": f"{TABLEAU_LOT}?lote={lot_num}",
        "lot_id":     f"TAB-{lot_num}",
    }

    # ── Foto: extrai work_id do link moldura.php?obr=XXXXX ────────────────
    m_obr = re.search(r"moldura\.php\?obr=([\w\-]+)", html, re.I)
    if m_obr:
        work_id = m_obr.group(1)
        data["foto_url"] = f"{TABLEAU_IMG}{work_id}g.jpg"

    # ── Parse por partes ─────────────────────────────────────────────────
    parts = [p.strip() for p in full_text.split("|") if p.strip()]

    i = 0
    while i < len(parts):
        seg   = parts[i]
        seg_n = _norm_tab(seg)

        # Lance base / tipo
        if re.search(r"Lote\s*N", seg, re.I):
            m_base = re.search(r"Base\s+R\$\s*([\d.,]+)", seg, re.I)
            if m_base:
                data["lance_base"] = clean_price(m_base.group(1))
            elif "livre" in seg.lower():
                data["lance_base"] = 100.0
            i += 1; continue

        # Lance atual / oferta
        m_oferta = re.search(r"oferta\s+de:\s*R\$\s*([\d.,]+)", seg, re.I)
        if m_oferta:
            data["maior_lance"] = clean_price(m_oferta.group(1))
            i += 1; continue

        # Data do leilão
        if re.search(r"terça|segunda|quarta|quinta|sexta|sábado|domingo", seg, re.I):
            if re.search(r"\d{4}", seg):
                data["data_leilao"] = seg.strip()[:60]
                i += 1; continue

        # Artista: maiúsculas + parênteses com anos
        if not data["artista"] and "(" in seg and ")" in seg and len(seg) < 100:
            before = seg[:seg.index("(")].strip()
            if before and len(before) > 3:
                if not re.search(r"lote|leil|lance|zoom|oferta|moldura", seg.lower()):
                    data["artista"] = seg.strip()
                    i += 1; continue

        # Labels conhecidos
        field = _TAB_LABEL_MAP.get(seg_n)
        if field and i + 1 < len(parts):
            val = parts[i + 1].strip()
            if val and len(val) < 300 and not data.get(field):
                data[field] = val
            i += 2; continue

        # Assinatura
        if re.search(r"assinad|canto|c\.i\.|c\.s\.", seg, re.I) and not data["assinatura"]:
            data["assinatura"] = detect_assinatura(seg) or detect_assinatura(full_text)

        i += 1

    # Complementa assinatura com texto completo se ainda vazio
    if not data["assinatura"]:
        data["assinatura"] = detect_assinatura(full_text)

    # Complementa ano a partir do artista
    if not data["ano"]:
        data["ano"] = extract_year(data["artista"])

    # Complementa dimensões
    if not data["dimensoes"]:
        dm = re.search(r"(\d+(?:[.,]\d+)?\s*[xX×]\s*\d+(?:[.,]\d+)?\s*(?:cm|mm)?)", full_text)
        if dm:
            data["dimensoes"] = dm.group(1).strip()

    # Limpa
    for f in ("artista", "titulo", "tecnica", "dimensoes", "ano", "data_leilao"):
        data[f] = str(data.get(f, "")).strip()[:200]

    # Expande abreviações na técnica
    data["tecnica"] = expand_abbreviations(data["tecnica"])

    # Retorna None se não conseguiu artista nem técnica
    if not data["artista"] and not data["tecnica"]:
        return None

    return data


def collect_tableau(session: requests.Session, db: dict) -> tuple[int, int]:
    """
    Varre todos os lotes ativos do Tableau Arte & Leilões.
    Adiciona diretamente ao DB. Retorna (novos, ignorados).
    """
    print(f"\n{'─' * 68}")
    print("  TABLEAU — Verificando disponibilidade do site...")
    print(f"{'─' * 68}\n")

    # Teste rápido de conectividade (3s) — pula tudo se o site estiver fora
    try:
        _ping = session.get(TABLEAU_LOT, params={"lote": 1}, timeout=5)
        if _ping.status_code != 200:
            print("  [TABLEAU] Site indisponível (status != 200) — pulando varredura.")
            return 0, 0
    except Exception as _e:
        print(f"  [TABLEAU] Site inacessível ({type(_e).__name__}) — pulando varredura.")
        return 0, 0

    print("  TABLEAU — Site OK. Varrendo lotes...")

    novos = 0
    ignorados = 0
    vazios_seguidos = 0

    for lot_num in range(1, TABLEAU_MAX + 1):
        lot_id = f"TAB-{lot_num}"
        if lot_id in db:
            continue  # já processado

        r = get(session, TABLEAU_LOT, params={"lote": lot_num})
        if not r or r.status_code != 200:
            vazios_seguidos += 1
            if vazios_seguidos >= 5:
                break
            continue

        detail = parse_tableau_lot(r.text, lot_num)

        if detail is None:
            vazios_seguidos += 1
            if vazios_seguidos >= 5:
                break
            continue

        vazios_seguidos = 0  # reset ao encontrar lote válido

        # Filtro: só pinturas
        tecnica_full = detail.get("tecnica", "") + " " + detail.get("titulo", "")
        if not is_pintura(tecnica_full):
            print(f"  TAB-{lot_num:03d}  (ignor. — {detail.get('tecnica','?')[:35]})")
            detail["_ignorado"] = True
            db[lot_id] = detail
            ignorados += 1
            save_db(db)
            time.sleep(DELAY * 0.5)
            continue

        detail["data_coleta"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        detail["em_leilao"] = True
        db[lot_id] = detail
        novos += 1

        artista  = (detail.get("artista") or "?")[:28]
        base     = detail.get("lance_base", 0)
        lance    = detail.get("maior_lance", 0)
        tecn     = (detail.get("tecnica") or "")[:30]
        lance_str = f"R${lance:>8,.0f}" if lance > 0 else "sem lance"
        print(f"  TAB-{lot_num:03d}  {artista:<28}  base R${base:>8,.0f}  {lance_str}  [{tecn}]")

        if novos % 15 == 0:
            save_db(db)

        time.sleep(DELAY * 0.5)

    save_db(db)
    print(f"\n  Tableau — novos: {novos}  |  ignorados: {ignorados}")
    return novos, ignorados


# ── Principal ─────────────────────────────────────────────────────────────────

def main():
    print("=" * 68)
    print("  CATÁLOGO DE PINTURAS — LeilõesBR")
    print(f"  Execução: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 68)

    db = load_db()
    _fechados = fechar_lotes_passados(db)
    if _fechados:
        save_db(db)
        print(f"  [limpeza] {_fechados} lote(s) com data passada marcados como encerrados")
    pinturas_db = {k: v for k, v in db.items() if not v.get("_ignorado")}
    print(f"  Base atual: {len(pinturas_db)} pinturas  |  {len(db)} lotes processados total")

    session = requests.Session()
    session.headers.update(HEADERS)

    # ── Fase 1: coleta IDs dos lotes ──────────────────────────────────────────
    print(f"\n{'─' * 68}")
    print("  FASE 1 — Coletando lotes do site...")
    print(f"{'─' * 68}")

    all_cards = collect_lot_ids(session)
    print(f"\n  Lotes encontrados no site : {len(all_cards)}")

    # Atualiza em_leilao de lotes já existentes no DB
    # Se um lote passou de ativo → finalizado, re-busca o detalhe para capturar o preço final
    atualizados = 0
    recoletar   = []   # lotes que enceraram e precisam de re-scraping
    card_map    = {c["lot_id"]: c for c in all_cards}

    for c in all_cards:
        lid = c["lot_id"]
        if lid in db and isinstance(db[lid], dict):
            era_ativo   = db[lid].get("em_leilao", True)
            agora_ativo = c["em_leilao"]
            if era_ativo != agora_ativo:
                db[lid]["em_leilao"] = agora_ativo
                atualizados += 1
                # Lote encerrou: agenda re-scraping para capturar lance final
                if era_ativo and not agora_ativo:
                    recoletar.append(c)

    if atualizados:
        print(f"  Status em_leilao atualizado em {atualizados} lotes existentes")
        save_db(db)

    if recoletar:
        print(f"\n{'─' * 68}")
        print(f"  FASE 1b — Re-coletando preço final de {len(recoletar)} lotes encerrados...")
        print(f"{'─' * 68}\n")
        atualizados_preco = 0
        for idx, card in enumerate(recoletar, 1):
            lid = card["lot_id"]
            print(f"  [{idx:03d}/{len(recoletar)}] ", end="", flush=True)
            detail = scrape_lot_detail(session, card)
            if detail is None:
                print("(erro)")
                time.sleep(DELAY)
                continue
            # Preserva em_leilao=False e dados já existentes, atualiza preços
            existing = db.get(lid, {})
            lance_novo = detail.get("maior_lance", 0)
            lance_ant  = existing.get("maior_lance", 0)
            if lance_novo > lance_ant:
                existing["maior_lance"] = lance_novo
                existing["num_lances"]  = detail.get("num_lances", existing.get("num_lances", 0))
                atualizados_preco += 1
                print(f"{existing.get('artista','?')[:28]}  lance final: R${lance_novo:,.0f}")
            else:
                print(f"{existing.get('artista','?')[:28]}  sem lance novo ({lance_ant:,.0f})")
            existing["em_leilao"] = False
            db[lid] = existing
            time.sleep(DELAY)
        if atualizados_preco:
            save_db(db)
        print(f"\n  Preços finais capturados: {atualizados_preco}/{len(recoletar)}")

    pending = [c for c in all_cards if c["lot_id"] not in db]
    print(f"  Já na base               : {len(all_cards) - len(pending)}")
    print(f"  Novos a processar        : {len(pending)}")

    if not pending:
        print("\n  Nenhum lote novo encontrado.")
    else:
        # ── Fase 2: scraping dos detalhes ─────────────────────────────────────
        print(f"\n{'─' * 68}")
        print("  FASE 2 — Obtendo detalhes...")
        print(f"{'─' * 68}\n")

        novos = 0
        ignorados = 0

        for idx, card in enumerate(pending, 1):
            pct = idx / len(pending) * 100
            print(f"  [{idx:04d}/{len(pending)}] {pct:5.1f}%  ", end="", flush=True)

            detail = scrape_lot_detail(session, card)

            if detail is None:
                print("(erro de conexão)")
                time.sleep(DELAY)
                continue

            # ── Filtro: só pinturas ────────────────────────────────────────
            tecnica_full = detail.get("tecnica", "") + " " + detail.get("titulo", "") + " " + card.get("card_text", "")
            if not is_pintura(tecnica_full):
                print(f"(ignor. — {detail.get('tecnica','?')[:30]})")
                detail["_ignorado"] = True
                db[card["lot_id"]] = detail
                ignorados += 1
                save_db(db)
                time.sleep(DELAY)
                continue

            detail["data_coleta"] = datetime.now().strftime("%d/%m/%Y %H:%M")
            detail["lot_id"] = card["lot_id"]
            detail["em_leilao"] = card.get("em_leilao", True)
            detail.pop("_ignorado", None)

            # Fallback foto: constrói URL CloudFront a partir do lot_id
            if not detail.get("foto_url"):
                _parts = card["lot_id"].split("|")
                if len(_parts) == 4:
                    _cat_id, _peca_id = _parts[2], _parts[3]
                    detail["foto_url"] = (
                        f"https://d1o6h00a1h5k7q.cloudfront.net"
                        f"/imagens/img_g/{_cat_id}/{_peca_id}.jpg"
                    )

            db[card["lot_id"]] = detail
            novos += 1

            # Log
            artista = (detail.get("artista") or "?")[:28]
            base    = detail.get("lance_base", 0)
            lance   = detail.get("maior_lance", 0)
            nlances = detail.get("num_lances", 0)
            tecn    = (detail.get("tecnica") or "")[:30]
            lance_str = f"R${lance:>8,.0f} ({nlances}L)" if lance > 0 else "sem lance"
            print(f"{artista:<28}  base R${base:>8,.0f}  {lance_str}  [{tecn}]")

            # Salva a cada 15 novos
            if novos % 15 == 0:
                save_db(db)
                print(f"\n  >> DB salvo — {novos} novos nesta execução\n")

            time.sleep(DELAY)

        save_db(db)
        print(f"\n  Novos lotes adicionados : {novos}")
        print(f"  Lotes ignorados         : {ignorados}")

    # ── Fase 2b: Tableau Arte & Leilões ───────────────────────────────────────
    collect_tableau(session, db)

    # ── Fase 3: exporta Excel ─────────────────────────────────────────────────
    print(f"\n{'─' * 68}")
    print("  FASE 3 — Gerando planilha...")
    save_excel(db)

    # Resumo final
    pinturas_final = {k: v for k, v in db.items() if not v.get("_ignorado")}
    com_lance = sum(1 for v in pinturas_final.values() if v.get("maior_lance", 0) > 0)

    print(f"\n{'=' * 68}")
    print("  RESUMO FINAL")
    print(f"  Total de pinturas na base : {len(pinturas_final)}")
    print(f"  Lotes com lance registrado: {com_lance}")
    print(f"  Arquivo: {OUTPUT_XLSX}")
    print("=" * 68)

    # ── Sincroniza com Supabase (se SUPABASE_KEY estiver definida) ─────────────
    try:
        import supabase_sync
        if supabase_sync.enabled():
            print("\n  Sincronizando com Supabase...")
            supabase_sync.sync_leiloesbr(db)
    except Exception as _e:
        print(f"  [supabase] aviso: {_e}")


def pausar():
    if sys.stdin.isatty():
        print("\n  Pressione ENTER para fechar...")
        input()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n  [!] Interrompido. DB foi salvo.")
    except Exception as e:
        import traceback
        print("\n\n  [ERRO INESPERADO]")
        traceback.print_exc()
    finally:
        pausar()
