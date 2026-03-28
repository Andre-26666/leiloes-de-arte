#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json, re, unicodedata
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def norm_artista(s):
    s = re.sub(r'\(.*?\)', '', s or '').strip()
    nfkd = unicodedata.normalize('NFKD', s.upper())
    s = ''.join(c for c in nfkd if not unicodedata.combining(c))
    s = re.sub(r'[^A-Z\s]', '', s).strip()
    return re.sub(r'\s+', ' ', s)

def vb(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def parse_dims(s: str) -> float:
    """'50x70cm', '50,5 x 30 cm', '100X80' → área em cm². Retorna 0 se não parsear."""
    if not s: return 0.0
    s = str(s).lower().replace(' ', '').replace('cm', '').replace('mm', '')
    m = re.match(r'([\d,\.]+)[x×]([\d,\.]+)', s)
    if m:
        try:
            w = float(m.group(1).replace(',', '.'))
            h = float(m.group(2).replace(',', '.'))
            if 0 < w < 1000 and 0 < h < 1000:
                return round(w * h, 1)
        except: pass
    return 0.0

# ── Histórico de preços ────────────────────────────────────────────────────
historico     = {}   # artista → [preços absolutos]
historico_cm2 = {}   # artista → [R$/cm²]

def add_preco(artista, preco, dims=''):
    k = norm_artista(artista)
    if not k or preco <= 0: return
    historico.setdefault(k, []).append(preco)
    area = parse_dims(dims)
    if area > 0:
        historico_cm2.setdefault(k, []).append(round(preco / area, 4))

with open('bolsadearte_db.json', 'r', encoding='utf-8') as f:
    bda = json.load(f)
for v in bda.values():
    if isinstance(v, dict):
        add_preco(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

with open('arrematearte_db.json', 'r', encoding='utf-8') as f:
    arr = json.load(f)
for k, v in arr.items():
    if not k.startswith('__meta') and isinstance(v, dict):
        add_preco(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

with open('leiloesbr_db.json', 'r', encoding='utf-8') as f:
    lbr = json.load(f)
for v in lbr.values():
    if isinstance(v, dict) and not v.get('_ignorado'):
        add_preco(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

print(f'Artistas com historico: {len(historico)}')
print(f'Precos totais: {sum(len(v) for v in historico.values())}')

# ── Catálogo Tableau ───────────────────────────────────────────────────────
with open('tableau_db.json', 'r', encoding='utf-8') as f:
    tableau = json.load(f)

PINTURA_KW = ['oleo','acrilica','acrilico','aquarela','guache','gouache',
              'pastel','tempera','tecnica mista','nanquim','tinta','encaustica']
GRAVURA_KW = ['serigrafia','litografia','xilogravura','gravura','offset',
              'impressao','multiplo','serigraph']

def tipo_obra(tec):
    t = unicodedata.normalize('NFKD', tec.lower())
    t = ''.join(c for c in t if not unicodedata.combining(c))
    if any(k in t for k in GRAVURA_KW): return 'Grafica'
    if any(k in t for k in PINTURA_KW): return 'Pintura'
    if any(k in t for k in ['escultura','bronze','terracota','ceramica']): return 'Escultura'
    if 'desenho' in t: return 'Desenho'
    return 'Outro'

TIER1 = {norm_artista(a) for a in [
    'TARSILA DO AMARAL','CANDIDO PORTINARI','EMILIANO DI CAVALCANTI',
    'ALFREDO VOLPI','ALBERTO DA VEIGA GUIGNARD','JOSE PANCETTI',
    'ISMAEL NERY','ANITA MALFATTI','DJANIRA DA MOTTA E SILVA',
    'ARTHUR TIMOTEO DA COSTA','ROBERTO BURLE MARX','HELIO OITICICA',
    'LASAR SEGALL','VICTOR BRECHERET','MARIO ZANINI','RUBEM VALENTIM',
    'ALDO BONADEI','FULVIO PENNACCHI','GEORGINA DE ALBUQUERQUE',
    'HENRI MATISSE','SALVADOR DALI',
]}
TIER2 = {norm_artista(a) for a in [
    'ALDEMIR MARTINS','CARYIBE','ATHOS BULCAO','CARLOS SCLIAR',
    'RUBENS GERCHMAN','CLAUDIO TOZZI','FERREIRA GULLAR','THOMAZ IANELLI',
    'ANTONIO POTEIRO','RENINA KATZ','LOTHAR CHAROUX','DARIO MECATTI',
    'INOS CORRADIN','HANSEN BAHIA','LIVIO ABRAMO','MILTON DACOSTA',
    'DURVAL PEREIRA','RUTH SCHLOSS','JOSE ANTONIO DA SILVA',
    'IWAO NAKAJIMA','CAROL KOSSAK','EDUARDO SUED','ALICE BRILL',
    'FRANCISCO BRENNAND','VICENTE DO REGO MONTEIRO','FRANCISCO DA SILVA',
    'SERGIO BERTONI','ISMAEL NERY','MARCELO GRASSMANN',
]}

rows = []
for lote in tableau:
    base = vb(lote.get('valor_base'))
    if base <= 0 or base > 20000:
        continue

    artista_raw = lote.get('artista','')
    artista_norm = norm_artista(artista_raw)
    artista_display = re.sub(r'\s*\(.*?\)\s*$', '', artista_raw).strip().title()

    tec = lote.get('tecnica','') or ''
    tipo = tipo_obra(tec)

    hist = historico.get(artista_norm, [])
    n_hist = len(hist)
    med_hist = float(np.median(hist)) if hist else 0.0
    max_hist = float(max(hist)) if hist else 0.0

    multiplo = round(med_hist / base, 1) if (base > 0 and med_hist > 0) else 0.0

    # ── Ajuste por tamanho (R$/cm²) ────────────────────────────────────────
    dims_raw       = lote.get('medidas','') or ''
    area_lote      = parse_dims(dims_raw)
    preco_cm2_lote = round(base / area_lote, 4) if area_lote > 0 else 0.0
    hist_cm2       = historico_cm2.get(artista_norm, [])
    n_hist_cm2     = len(hist_cm2)
    med_cm2_hist   = float(np.median(hist_cm2)) if hist_cm2 else 0.0
    mult_cm2       = round(med_cm2_hist / preco_cm2_lote, 1) if (preco_cm2_lote > 0 and med_cm2_hist > 0) else 0.0

    # Múltiplo efetivo: usa ajustado por tamanho se disponível, senão absoluto
    multiplo_efetivo = mult_cm2 if mult_cm2 > 0 else multiplo

    if artista_norm in TIER1:
        tier = 'A - Mestre Nacional'
    elif artista_norm in TIER2:
        tier = 'B - Moderno Relevante'
    elif n_hist >= 5:
        tier = 'C - Mercado Ativo'
    elif n_hist >= 1:
        tier = 'D - Referencia Limitada'
    else:
        tier = 'E - Sem Historico'

    liq = min(10.0, n_hist * 1.5)
    if artista_norm in TIER1: liq = min(10.0, liq + 3.0)
    elif artista_norm in TIER2: liq = min(10.0, liq + 1.5)
    liq = round(liq, 1)

    oport = 0.0
    if multiplo_efetivo > 0:
        oport += min(40.0, multiplo_efetivo * 8)
    oport += liq * 3
    if tipo == 'Pintura':   oport += 20
    elif tipo == 'Desenho': oport += 10
    elif tipo == 'Grafica' and artista_norm in TIER1: oport += 8
    if artista_norm in TIER1:   oport += 15
    elif artista_norm in TIER2: oport += 8
    # Bônus/penalidade por formato (pintura)
    if tipo == 'Pintura' and area_lote > 0:
        if area_lote >= 3000:   oport += 5   # grande formato
        elif area_lote < 400:   oport -= 3   # obra muito pequena
    if n_hist == 0: oport *= 0.3
    oport = round(min(100.0, oport), 1)

    # Recomendação simplificada
    if oport >= 75:
        rec = 'FORTE'
    elif oport >= 55:
        rec = 'BOA'
    elif oport >= 35:
        rec = 'MODERADA'
    else:
        rec = 'BAIXA'

    rows.append({
        'Score':            oport,
        'Rec':              rec,
        'Lote':             lote.get('lote_num',''),
        'Artista':          artista_display,
        'Titulo':           lote.get('titulo','') or '',
        'Tipo':             tipo,
        'Tecnica':          tec,
        'Medidas':          dims_raw,
        'Area_cm2':         round(area_lote) if area_lote > 0 else None,
        'Ano':              lote.get('data_obra','') or '',
        'Assinado':         lote.get('assinado','') or '',
        'Valor_Base':       base,
        'R_cm2':            round(preco_cm2_lote, 2) if preco_cm2_lote > 0 else None,
        'Med_R_cm2':        round(med_cm2_hist, 2) if med_cm2_hist > 0 else None,
        'Mult_Tam':         mult_cm2 if mult_cm2 > 0 else None,
        'Mediana_Hist':     med_hist if med_hist > 0 else None,
        'Max_Hist':         max_hist if max_hist > 0 else None,
        'Multiplo':         multiplo if multiplo > 0 else None,
        'N_Hist':           n_hist,
        'Liquidez':         liq,
        'Tier':             tier,
        'URL':              lote.get('url_lote',''),
        'Foto':             lote.get('img_grande',''),
    })

df = pd.DataFrame(rows).sort_values('Score', ascending=False)
top = df[df['Score'] >= 30].head(50).copy()

print(f'\nLotes analisados: {len(df)}')
print(f'Score >= 30 (tabela principal): {len(top)}')
print('\nTop 20 oportunidades:')
for _, r in top.head(20).iterrows():
    mult = f'{r["Multiplo"]}x' if r['Multiplo'] else 'sem ref'
    print(f'  Lote {str(r["Lote"]):>3} | {r["Artista"]:<30} | R${r["Valor_Base"]:>7,.0f} | {mult:>7} | Liq:{r["Liquidez"]:>4} | Score:{r["Score"]:>5} | {r["Tipo"]} | {r["Rec"]}')

# ── Excel ──────────────────────────────────────────────────────────────────
wb = Workbook()
ws1 = wb.active
ws1.title = 'Top Oportunidades'
ws2 = wb.create_sheet('Catalogo Completo')
ws3 = wb.create_sheet('Metodologia')

C_DARK  = '0F0F1A'
C_CARD  = '1A1A2E'
C_ALT   = '161628'
C_GOLD  = 'FFD700'
C_GREEN = '22C55E'
C_BLUE  = '60A5FA'
C_RED   = 'EF4444'
C_AMBER = 'F59E0B'
C_GRAY  = '94A3B8'
C_TEXT  = 'E2E8F0'
C_BORD  = '2D2D4E'

TIER_STYLE = {
    'A - Mestre Nacional':    ('1E3A5F', 'FFD700'),
    'B - Moderno Relevante':  ('1A3A2A', '22C55E'),
    'C - Mercado Ativo':      ('1A2A3A', '60A5FA'),
    'D - Referencia Limitada':('2A2A1A', 'F59E0B'),
    'E - Sem Historico':      ('2A1A1A', '94A3B8'),
}
REC_COLOR = {'FORTE':'22C55E','BOA':'60A5FA','MODERADA':'F59E0B','BAIXA':'94A3B8'}

def fill(c): return PatternFill('solid', fgColor=c)
def font(color=C_TEXT, size=10, bold=False, italic=False):
    return Font(name='Calibri', size=size, bold=bold, italic=italic, color=color)
def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border():
    s = Side(style='thin', color=C_BORD)
    return Border(bottom=s, right=s)

def write_header(ws, row, ncols, title, subtitle=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name='Calibri', size=18, bold=True, color=C_GOLD)
    c.fill = fill(C_DARK)
    c.alignment = align('center')
    ws.row_dimensions[row].height = 38
    if subtitle:
        r2 = row + 1
        ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=ncols)
        c2 = ws.cell(row=r2, column=1, value=subtitle)
        c2.font = Font(name='Calibri', size=10, color=C_GRAY)
        c2.fill = fill(C_DARK)
        c2.alignment = align('center')
        ws.row_dimensions[r2].height = 18

def write_col_headers(ws, row, headers, bg=C_CARD):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name='Calibri', size=10, bold=True, color=C_GOLD)
        c.fill = fill(bg)
        c.alignment = align('center', wrap=True)
        c.border = Border(bottom=Side(style='medium', color=C_GOLD))
    ws.row_dimensions[row].height = 30

# ────── ABA 1: TOP OPORTUNIDADES ────────────────────────────────────────────
HDR1 = ['Rec','Score','Lote','Artista','Titulo','Tipo','Tecnica',
        'Medidas','Area cm²','Ano','Assinado','Valor Base R$',
        'R$/cm²','Med R$/cm²','Mult Tam','Mediana Hist R$',
        'Max Hist R$','Multiplo','Liquidez','Tier','Ver Lote']
N1 = len(HDR1)
write_header(ws1, 1, N1,
    'TABLEAU ARTE & LEILOES  |  TOP OPORTUNIDADES ATE R$ 20.000',
    f'Leilao 24-26/03/2026   |  Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}   |  Score >= 30 de 100')
write_col_headers(ws1, 3, HDR1)

widths1 = [9,7,6,28,32,9,22,13,9,6,16,13,10,10,10,14,13,10,9,22,35]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A4'

for r_idx, (_, row) in enumerate(top.iterrows(), 4):
    bg = C_ALT if r_idx % 2 == 0 else C_CARD
    score = row['Score']
    tier  = row['Tier']
    tb, tf = TIER_STYLE.get(tier, (C_CARD, C_TEXT))

    def wc(col, val, fmt=None, h='left', bold=False, color=C_TEXT, bg_=None):
        cell = ws1.cell(row=r_idx, column=col, value=val)
        cell.font = Font(name='Calibri', size=10, bold=bold, color=color)
        cell.fill = fill(bg_ if bg_ is not None else bg)
        cell.alignment = align(h)
        cell.border = border()
        if fmt: cell.number_format = fmt
        return cell

    rec = row['Rec']
    rc = ws1.cell(row=r_idx, column=1, value=rec)
    rc.font = Font(name='Calibri', size=10, bold=True, color=REC_COLOR.get(rec, C_TEXT))
    rc.fill = fill(bg); rc.alignment = align('center'); rc.border = border()

    sc = ws1.cell(row=r_idx, column=2, value=score)
    sc_color = C_GREEN if score >= 75 else (C_BLUE if score >= 55 else (C_AMBER if score >= 35 else C_GRAY))
    sc.font = Font(name='Calibri', size=11, bold=True, color=sc_color)
    sc.fill = fill(bg); sc.alignment = align('center'); sc.border = border()

    wc(3,  row['Lote'],     h='center')
    ac = ws1.cell(row=r_idx, column=4, value=row['Artista'])
    ac.font = Font(name='Calibri', size=10, bold=True, color=tf)
    ac.fill = fill(tb); ac.alignment = align(); ac.border = border()

    wc(5,  row['Titulo'])
    wc(6,  row['Tipo'],     h='center')
    wc(7,  row['Tecnica'])
    wc(8,  row['Medidas'],  h='center')
    wc(9,  row['Area_cm2'], h='center', fmt='#,##0')
    wc(10, row['Ano'],      h='center')
    wc(11, row['Assinado'])

    vc = ws1.cell(row=r_idx, column=12, value=row['Valor_Base'])
    vc.number_format = 'R$ #,##0'; vc.font = Font(name='Calibri', size=10, bold=True, color=C_GOLD)
    vc.fill = fill(bg); vc.alignment = align('center'); vc.border = border()

    # R$/cm² do lote
    rcm = row['R_cm2']
    rcmc = ws1.cell(row=r_idx, column=13, value=rcm)
    rcmc.number_format = 'R$ #,##0.00'
    rcmc.font = Font(name='Calibri', size=10, color=C_TEXT if rcm else C_GRAY)
    rcmc.fill = fill(bg); rcmc.alignment = align('center'); rcmc.border = border()

    # Mediana histórica R$/cm²
    mcm = row['Med_R_cm2']
    mcmc = ws1.cell(row=r_idx, column=14, value=mcm)
    mcmc.number_format = 'R$ #,##0.00'
    mcmc.font = Font(name='Calibri', size=10, color=C_GREEN if mcm else C_GRAY)
    mcmc.fill = fill(bg); mcmc.alignment = align('center'); mcmc.border = border()

    # Múltiplo ajustado por tamanho (número com sufixo visual "x" — filtrável)
    mt = row['Mult_Tam']
    mtc = ws1.cell(row=r_idx, column=15, value=mt)
    mtc.number_format = '0.00'
    mtc.font = Font(name='Calibri', size=10, bold=bool(mt and mt >= 2),
                    color=C_GREEN if (mt and mt >= 2) else (C_AMBER if mt else C_GRAY))
    mtc.fill = fill(bg); mtc.alignment = align('center'); mtc.border = border()

    med = row['Mediana_Hist']
    mc = ws1.cell(row=r_idx, column=16, value=med)
    mc.number_format = 'R$ #,##0'
    mc.font = Font(name='Calibri', size=10, color=C_GREEN if med else C_GRAY)
    mc.fill = fill(bg); mc.alignment = align('center'); mc.border = border()

    mx = row['Max_Hist']
    mxc = ws1.cell(row=r_idx, column=17, value=mx)
    mxc.number_format = 'R$ #,##0'
    mxc.font = Font(name='Calibri', size=10, color=C_AMBER if mx else C_GRAY)
    mxc.fill = fill(bg); mxc.alignment = align('center'); mxc.border = border()

    # Múltiplo absoluto (número com sufixo "x" — filtrável)
    mult = row['Multiplo']
    multc = ws1.cell(row=r_idx, column=18, value=mult)
    multc.number_format = '0.00'
    multc.font = Font(name='Calibri', size=10, bold=bool(mult and mult >= 2),
                      color=C_GREEN if (mult and mult >= 2) else C_TEXT)
    multc.fill = fill(bg); multc.alignment = align('center'); multc.border = border()

    liq = row['Liquidez']
    lc = ws1.cell(row=r_idx, column=19, value=liq)
    lc.number_format = '0.0'
    lc.font = Font(name='Calibri', size=10,
                   color=C_GREEN if liq >= 7 else (C_AMBER if liq >= 4 else C_GRAY))
    lc.fill = fill(bg); lc.alignment = align('center'); lc.border = border()

    tc = ws1.cell(row=r_idx, column=20, value=tier)
    tc.font = Font(name='Calibri', size=9, bold=True, color=tf)
    tc.fill = fill(tb); tc.alignment = align('center', wrap=True); tc.border = border()

    url = row['URL']
    uc = ws1.cell(row=r_idx, column=21, value='Ver lote' if url else '')
    if url:
        uc.hyperlink = url
        uc.font = Font(name='Calibri', size=9, color=C_BLUE, underline='single')
    uc.fill = fill(bg); uc.alignment = align('center'); uc.border = border()

    ws1.row_dimensions[r_idx].height = 26

ultima_linha_1 = 3 + len(top)
ws1.auto_filter.ref = f'A3:{get_column_letter(N1)}{ultima_linha_1}'

# ────── ABA 2: CATÁLOGO COMPLETO ────────────────────────────────────────────
HDR2 = ['Score','Rec','Lote','Artista','Titulo','Tipo','Tecnica','Medidas',
        'Area cm²','Valor Base','R$/cm²','Med R$/cm²','Mult Tam',
        'Mediana Hist','Multiplo','Liquidez','N Hist','Tier','URL']
N2 = len(HDR2)
write_header(ws2, 1, N2, 'CATALOGO COMPLETO  |  TODOS OS LOTES ATE R$ 20.000')
write_col_headers(ws2, 2, HDR2)
w2 = [7,8,6,26,28,9,20,13,9,12,10,10,10,13,9,9,7,22,35]
for i,w in enumerate(w2,1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A3'

for r_idx, (_, row) in enumerate(df.iterrows(), 3):
    bg = C_ALT if r_idx % 2 == 0 else C_CARD
    for col, key, fmt, h in [
        (1,'Score','0','center'), (2,'Rec',None,'center'),
        (3,'Lote',None,'center'), (4,'Artista',None,'left'),
        (5,'Titulo',None,'left'), (6,'Tipo',None,'center'),
        (7,'Tecnica',None,'left'), (8,'Medidas',None,'center'),
        (9,'Area_cm2','#,##0','center'),
        (10,'Valor_Base','R$ #,##0','center'),
        (11,'R_cm2','R$ #,##0.00','center'),
        (12,'Med_R_cm2','R$ #,##0.00','center'),
        (13,'Mult_Tam','0.00','center'),
        (14,'Mediana_Hist','R$ #,##0','center'),
        (15,'Multiplo','0.00','center'),
        (16,'Liquidez','0.0','center'),
        (17,'N_Hist',None,'center'),
        (18,'Tier',None,'center'),
        (19,'URL',None,'left'),
    ]:
        val = row.get(key)
        c = ws2.cell(row=r_idx, column=col, value=val if pd.notna(val) and val != 0 and val is not None else None)
        c.font = Font(name='Calibri', size=9, color=C_TEXT)
        c.fill = fill(bg)
        c.alignment = align(h)
        c.border = border()
        if fmt: c.number_format = fmt
        if col == 19 and val:
            c.hyperlink = str(val)
            c.font = Font(name='Calibri', size=9, color=C_BLUE, underline='single')
    ws2.row_dimensions[r_idx].height = 16

ultima_linha_2 = 2 + len(df)
ws2.auto_filter.ref = f'A2:{get_column_letter(N2)}{ultima_linha_2}'

# ────── ABA 3: METODOLOGIA ────────────────────────────────────────────────
ws3.column_dimensions['A'].width = 90
n_bda = sum(1 for v in bda.values() if isinstance(v,dict) and vb(v.get('maior_lance'))>0)
n_arr = sum(1 for k,v in arr.items() if not k.startswith('__meta') and isinstance(v,dict) and vb(v.get('maior_lance'))>0)

linhas = [
    ('METODOLOGIA DE ANALISE', True, 16, C_GOLD),
    ('', False, 11, C_TEXT),
    ('FONTES DE DADOS', True, 13, C_BLUE),
    (f'Bolsa de Arte (bolsadearte.com): {n_bda:,} precos historicos', False, 11, C_TEXT),
    (f'Arremate Arte (TNT Arte + Alexei): {n_arr:,} lances registrados', False, 11, C_TEXT),
    ('LeiloesBR: historico de lotes finalizados', False, 11, C_TEXT),
    ('', False, 11, C_TEXT),
    ('COMO O SCORE E CALCULADO (0-100)', True, 13, C_BLUE),
    ('Multiplo (ate 40 pts): usa R$/cm2 ajustado por tamanho se disponivel, senao mediana absoluta / base', False, 11, C_TEXT),
    ('Ajuste tamanho: compara preco por cm2 do lote vs mediana historica R$/cm2 do artista', False, 11, C_TEXT),
    ('Grande formato (>=3000cm2) +5pts | Obra minuscula (<400cm2) -3pts', False, 11, C_TEXT),
    ('Liquidez (ate 30 pts): frequencia em leiloes historicos (0-10) x 3pts', False, 11, C_TEXT),
    ('Tipo de Obra: Pintura original +20pts | Desenho +10pts | Grafica de Mestre +8pts', False, 11, C_TEXT),
    ('Tier do Artista: Mestre Nacional +15pts | Moderno Relevante +8pts', False, 11, C_TEXT),
    ('Penalidade: artista sem qualquer referencia historica -> score x 0.30', False, 11, C_TEXT),
    ('', False, 11, C_TEXT),
    ('INTERPRETACAO DA RECOMENDACAO', True, 13, C_BLUE),
    ('FORTE  (score 75-100): oportunidade clara, multiplo alto e artista liquido', False, 11, C_GREEN),
    ('BOA    (score 55-74):  boa relacao risco/retorno, historico confirma valor', False, 11, C_BLUE),
    ('MODERADA (35-54):      potencial mas com incertezas de liquidez ou historico', False, 11, C_AMBER),
    ('BAIXA  (0-34):         risco elevado, pouco historico ou multiplicador baixo', False, 11, C_GRAY),
    ('', False, 11, C_TEXT),
    ('TIER DE ARTISTAS', True, 13, C_BLUE),
    ('A - Mestre Nacional: canonicos da arte brasileira moderna e contemporanea', False, 11, C_GOLD),
    ('B - Moderno Relevante: presenca consolidada em leiloes, mercado ativo', False, 11, C_GREEN),
    ('C - Mercado Ativo: 5+ aparicoes em leiloes historicos nossa base', False, 11, C_BLUE),
    ('D - Referencia Limitada: 1-4 aparicoes historicas', False, 11, C_AMBER),
    ('E - Sem Historico: nenhuma referencia de preco encontrada nas nossas bases', False, 11, C_GRAY),
    ('', False, 11, C_TEXT),
    ('AVISOS IMPORTANTES', True, 13, C_RED),
    ('Este relatorio e auxiliar - nao constitui recomendacao de investimento financeiro', False, 11, C_TEXT),
    ('Graficas/serigrafias: menor liquidez que pinturas originais, mesmo de grandes mestres', False, 11, C_TEXT),
    ('Considerar custos adicionais: comissao leilao ~15-20%, seguro e transporte', False, 11, C_TEXT),
    ('Obras com assinatura verificavel tem premio de liquidez no mercado secundario', False, 11, C_TEXT),
    ('', False, 11, C_TEXT),
    (f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', False, 10, C_GRAY),
]
for r_idx, (txt, bold, size, color) in enumerate(linhas, 1):
    c = ws3.cell(row=r_idx, column=1, value=txt)
    c.font = Font(name='Calibri', size=size, bold=bold, color=color)
    c.fill = fill(C_DARK)
    c.alignment = align('left', wrap=True)
    ws3.row_dimensions[r_idx].height = 22 if bold else 17

ws1.sheet_properties.tabColor = C_GOLD
ws2.sheet_properties.tabColor = '3B82F6'
ws3.sheet_properties.tabColor = '64748B'

OUT = 'tableau_oportunidades.xlsx'
wb.save(OUT)
print(f'\nPlanilha salva: {OUT}')
