#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Relatório Diário Unificado — Todas as Casas
Uso:
  python relatorio_diario.py            # leilões dos próximos 3 dias
  python relatorio_diario.py 7          # próximos 7 dias
  python relatorio_diario.py 23/03/2026 # data específica
  python relatorio_diario.py all        # todos os lotes ativos (sem filtro de data)
"""

import json, re, sys, unicodedata, os
from datetime import datetime, timedelta, date
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Argumento de data ──────────────────────────────────────────────────────
arg = sys.argv[1] if len(sys.argv) > 1 else '3'
HOJE = date.today()

if arg.lower() == 'all':
    data_inicio = None
    data_fim    = None
    label_data  = 'Todos os lotes ativos'
elif re.match(r'\d{2}/\d{2}/\d{4}', arg):
    d = datetime.strptime(arg, '%d/%m/%Y').date()
    data_inicio = data_fim = d
    label_data  = arg
else:
    dias = int(arg)
    data_inicio = HOJE
    data_fim    = HOJE + timedelta(days=dias - 1)
    if dias == 1:
        label_data = f'Hoje ({HOJE.strftime("%d/%m/%Y")})'
    else:
        label_data = f'{HOJE.strftime("%d/%m/%Y")} ate {data_fim.strftime("%d/%m/%Y")} ({dias} dias)'

print('=' * 65)
print(f'RELATÓRIO DIÁRIO  |  {label_data}')
print('=' * 65)


# ── Helpers ────────────────────────────────────────────────────────────────
def norm_artista(s):
    s = re.sub(r'\(.*?\)', '', s or '').strip()
    nfkd = unicodedata.normalize('NFKD', s.upper())
    s = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', re.sub(r'[^A-Z\s]', '', s)).strip()

def vb(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def parse_dims(s: str) -> float:
    if not s: return 0.0
    s = str(s).lower().replace(' ', '').replace('cm', '').replace('mm', '')
    m = re.match(r'([\d,\.]+)[x×]([\d,\.]+)', s)
    if m:
        try:
            w = float(m.group(1).replace(',', '.'))
            h = float(m.group(2).replace(',', '.'))
            if 0 < w < 1000 and 0 < h < 1000:
                return round(w * h, 1)
        except: pass
    return 0.0

def parse_date(s: str):
    """Tenta parsear data nos formatos D/M/YYYY, DD/MM/YYYY ou 'DD de mês de YYYY'."""
    if not s: return None
    s = s.strip()
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        try: return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except: pass
    meses = {'janeiro':1,'fevereiro':2,'março':3,'marco':3,'abril':4,
             'maio':5,'junho':6,'julho':7,'agosto':8,'setembro':9,
             'outubro':10,'novembro':11,'dezembro':12}
    m2 = re.search(r'(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})', s, re.I)
    if m2:
        mes = meses.get(m2.group(2).lower())
        if mes:
            try: return date(int(m2.group(3)), mes, int(m2.group(1)))
            except: pass
    # só ano/mês (ex: "março 2026")
    m3 = re.search(r'(\w+)\s+(\d{4})', s, re.I)
    if m3:
        mes = meses.get(m3.group(1).lower())
        if mes:
            try: return date(int(m3.group(2)), mes, 1)
            except: pass
    return None

def data_ok(d_leilao: str) -> bool:
    """True se a data do leilão está dentro da janela solicitada."""
    if data_inicio is None:
        return True
    d = parse_date(d_leilao)
    if d is None:
        return False
    return data_inicio <= d <= data_fim


# ── Histórico de preços (todas as fontes) ─────────────────────────────────
historico     = {}
historico_cm2 = {}

def add_preco(artista, preco, dims=''):
    k = norm_artista(artista)
    if not k or preco <= 0: return
    historico.setdefault(k, []).append(preco)
    area = parse_dims(dims)
    if area > 0:
        historico_cm2.setdefault(k, []).append(round(preco / area, 4))

def load_json(fname):
    p = os.path.join(_DIR, fname)
    if os.path.exists(p):
        with open(p, 'r', encoding='utf-8') as f:
            return json.load(f)
    print(f'  [AVISO] {fname} não encontrado — pulando')
    return {}

bda = load_json('bolsadearte_db.json')
arr = load_json('arrematearte_db.json')
lbr = load_json('leiloesbr_db.json')
cda = load_json('cda_db.json')

for v in bda.values():
    if isinstance(v, dict):
        add_preco(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

for k, v in arr.items():
    if not k.startswith('__meta') and isinstance(v, dict):
        add_preco(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

for v in lbr.values():
    if isinstance(v, dict) and not v.get('_ignorado') and v.get('em_leilao') is False:
        add_preco(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

print(f'  Histórico: {len(historico):,} artistas | {sum(len(v) for v in historico.values()):,} preços')
print(f'  Histórico R$/cm²: {len(historico_cm2):,} artistas com dimensões')


# ── Tiers de artistas ─────────────────────────────────────────────────────
TIER1 = {norm_artista(a) for a in [
    'TARSILA DO AMARAL','CANDIDO PORTINARI','EMILIANO DI CAVALCANTI',
    'ALFREDO VOLPI','ALBERTO DA VEIGA GUIGNARD','JOSE PANCETTI',
    'ISMAEL NERY','ANITA MALFATTI','DJANIRA DA MOTTA E SILVA',
    'ARTHUR TIMOTEO DA COSTA','ROBERTO BURLE MARX','HELIO OITICICA',
    'LASAR SEGALL','VICTOR BRECHERET','MARIO ZANINI','RUBEM VALENTIM',
    'ALDO BONADEI','FULVIO PENNACCHI','GEORGINA DE ALBUQUERQUE',
    'TARSILA','DI CAVALCANTI','HENRI MATISSE','SALVADOR DALI',
]}
TIER2 = {norm_artista(a) for a in [
    'ALDEMIR MARTINS','CARYIBE','CARYBE','ATHOS BULCAO','CARLOS SCLIAR',
    'RUBENS GERCHMAN','CLAUDIO TOZZI','FERREIRA GULLAR','THOMAZ IANELLI',
    'ANTONIO POTEIRO','RENINA KATZ','LOTHAR CHAROUX','DARIO MECATTI',
    'INOS CORRADIN','HANSEN BAHIA','LIVIO ABRAMO','MILTON DACOSTA',
    'DURVAL PEREIRA','RUTH SCHLOSS','JOSE ANTONIO DA SILVA',
    'IWAO NAKAJIMA','CAROL KOSSAK','EDUARDO SUED','ALICE BRILL',
    'FRANCISCO BRENNAND','VICENTE DO REGO MONTEIRO','FRANCISCO DA SILVA',
    'SERGIO BERTONI','MARCELO GRASSMANN','GERSON DE SOUZA',
    'MANOEL SANTIAGO','SYLVIO PINTO','MANEZINHO ARAUJO',
    'NEWTON REZENDE','ESCOLA CUSQUENHA',
]}

PINTURA_KW = ['oleo','acrilica','acrilico','aquarela','guache','gouache',
              'pastel','tempera','tecnica mista','nanquim','tinta','encaustica']
GRAVURA_KW = ['serigrafia','litografia','xilogravura','gravura','offset',
              'impressao','multiplo','serigraph']

def tipo_obra(tec):
    t = unicodedata.normalize('NFKD', (tec or '').lower())
    t = ''.join(c for c in t if not unicodedata.combining(c))
    if any(k in t for k in GRAVURA_KW): return 'Gráfica'
    if any(k in t for k in PINTURA_KW): return 'Pintura'
    if any(k in t for k in ['escultura','bronze','terracota','ceramica']): return 'Escultura'
    if 'desenho' in t: return 'Desenho'
    return 'Outro'


# ── Score ──────────────────────────────────────────────────────────────────
def calcular_score(artista_raw, tec, dims, preco_ref, data_leilao=''):
    """Retorna dict com todos os campos calculados."""
    artista_norm = norm_artista(artista_raw)
    if not artista_norm or artista_raw.lower() in ('autor desconhecido','?','','desconhecido'):
        return None

    hist     = historico.get(artista_norm, [])
    n_hist   = len(hist)
    med_hist = float(np.median(hist)) if hist else 0.0
    max_hist = float(max(hist))       if hist else 0.0
    multiplo = round(med_hist / preco_ref, 2) if (preco_ref > 0 and med_hist > 0) else 0.0

    area           = parse_dims(dims)
    preco_cm2      = round(preco_ref / area, 4) if area > 0 else 0.0
    hist_cm2       = historico_cm2.get(artista_norm, [])
    med_cm2        = float(np.median(hist_cm2)) if hist_cm2 else 0.0
    mult_cm2       = round(med_cm2 / preco_cm2, 2) if (preco_cm2 > 0 and med_cm2 > 0) else 0.0
    mult_efetivo   = mult_cm2 if mult_cm2 > 0 else multiplo

    if artista_norm in TIER1:   tier = 'A - Mestre'
    elif artista_norm in TIER2: tier = 'B - Moderno'
    elif n_hist >= 5:           tier = 'C - Ativo'
    elif n_hist >= 1:           tier = 'D - Limitado'
    else:                       tier = 'E - Sem Hist'

    liq = min(10.0, n_hist * 1.5)
    if artista_norm in TIER1:   liq = min(10.0, liq + 3.0)
    elif artista_norm in TIER2: liq = min(10.0, liq + 1.5)
    liq = round(liq, 1)

    tipo = tipo_obra(tec)
    oport = 0.0
    if mult_efetivo > 0:
        oport += min(40.0, mult_efetivo * 8)
    oport += liq * 3
    if tipo == 'Pintura':    oport += 20
    elif tipo == 'Desenho':  oport += 10
    elif tipo == 'Gráfica' and artista_norm in TIER1: oport += 8
    if artista_norm in TIER1:   oport += 15
    elif artista_norm in TIER2: oport += 8
    if area > 0:
        if   area >= 3000: oport += 5
        elif area < 400:   oport -= 3
    if n_hist == 0: oport *= 0.3
    oport = round(min(100.0, oport), 1)

    if oport >= 75:   rec = 'FORTE'
    elif oport >= 55: rec = 'BOA'
    elif oport >= 35: rec = 'MODERADA'
    else:             rec = 'BAIXA'

    d = parse_date(data_leilao)
    data_fmt = d.strftime('%d/%m/%Y') if d else data_leilao

    return dict(
        Score=oport, Rec=rec, Tier=tier,
        Data=data_fmt, Artista=artista_raw.strip().title(),
        Tipo=tipo, Tecnica=tec or '', Dimensoes=dims or '',
        Area_cm2=round(area) if area > 0 else None,
        Preco_Ref=preco_ref,
        R_cm2=round(preco_cm2, 2) if preco_cm2 > 0 else None,
        Med_R_cm2=round(med_cm2, 2) if med_cm2 > 0 else None,
        Mult_Tam=mult_cm2 if mult_cm2 > 0 else None,
        Mediana_Hist=med_hist if med_hist > 0 else None,
        Max_Hist=max_hist if max_hist > 0 else None,
        Multiplo=multiplo if multiplo > 0 else None,
        Liquidez=liq, N_Hist=n_hist,
        artista_norm=artista_norm,
    )


# ── Coleta lotes de todas as fontes ────────────────────────────────────────
rows = []
contagem = {}

def add_lote(artista, tec, dims, base, lance, data_leilao,
             casa, url, foto, titulo='', ano='', ass='', lote_num=''):
    preco_ref = lance if lance > 0 else base
    if preco_ref <= 0: return
    if not data_ok(data_leilao): return

    sc = calcular_score(artista, tec, dims, preco_ref, data_leilao)
    if sc is None: return

    sc.update(dict(
        Titulo=titulo, Ano=ano, Assinatura=ass, Lote=lote_num,
        Lance_Base=base, Lance_Atual=lance if lance > 0 else None,
        Casa=casa, URL=url, Foto=foto,
    ))
    rows.append(sc)
    contagem[casa] = contagem.get(casa, 0) + 1

# ── LeiloesBR (ativos) ─────────────────────────────────────────────────────
n_lbr = 0
for v in lbr.values():
    if not isinstance(v, dict) or v.get('_ignorado'): continue
    if not v.get('em_leilao', True): continue
    add_lote(
        artista    = v.get('artista',''),
        tec        = v.get('tecnica',''),
        dims       = v.get('dimensoes',''),
        base       = vb(v.get('lance_base')),
        lance      = vb(v.get('maior_lance')),
        data_leilao= v.get('data_leilao','') or v.get('data_coleta',''),
        casa       = v.get('casa','leiloesbr'),
        url        = v.get('url_detalhe',''),
        foto       = v.get('foto_url',''),
        titulo     = v.get('titulo',''),
        ano        = v.get('ano',''),
        ass        = v.get('assinatura',''),
        lote_num   = v.get('lote_id',''),
    )

# ── Arremate Arte: TNT e Alexei (ativos) ──────────────────────────────────
for k, v in arr.items():
    if k.startswith('__meta') or not isinstance(v, dict): continue
    if not v.get('em_leilao', True): continue
    add_lote(
        artista    = v.get('artista',''),
        tec        = v.get('tecnica',''),
        dims       = v.get('dimensoes',''),
        base       = vb(v.get('lance_base')),
        lance      = vb(v.get('maior_lance')),
        data_leilao= v.get('data_leilao','') or v.get('data_coleta',''),
        casa       = v.get('casa','arrematearte'),
        url        = v.get('url_detalhe',''),
        foto       = v.get('foto_url',''),
        titulo     = v.get('titulo',''),
        ano        = v.get('ano',''),
    )

print(f'\n  Lotes coletados por casa:')
for casa, n in sorted(contagem.items()):
    print(f'    {casa:<40} {n:>4} lotes')
print(f'  Total: {len(rows)} lotes com artista identificado na janela {label_data}')

if not rows:
    print('\n  Nenhum lote encontrado para o período. Tente "all" para ver todos os ativos.')
    sys.exit(0)

df = pd.DataFrame(rows).sort_values('Score', ascending=False)


# ── Excel ──────────────────────────────────────────────────────────────────
C_DARK  = '0F0F1A'; C_CARD  = '1A1A2E'; C_ALT   = '161628'
C_GOLD  = 'FFD700'; C_GREEN = '22C55E'; C_BLUE  = '60A5FA'
C_RED   = 'EF4444'; C_AMBER = 'F59E0B'; C_GRAY  = '94A3B8'
C_TEXT  = 'E2E8F0'; C_BORD  = '2D2D4E'

TIER_STYLE = {
    'A - Mestre':    ('1E3A5F','FFD700'),
    'B - Moderno':   ('1A3A2A','22C55E'),
    'C - Ativo':     ('1A2A3A','60A5FA'),
    'D - Limitado':  ('2A2A1A','F59E0B'),
    'E - Sem Hist':  ('2A1A1A','94A3B8'),
}
REC_COLOR = {'FORTE':'22C55E','BOA':'60A5FA','MODERADA':'F59E0B','BAIXA':'94A3B8'}

def fill(c): return PatternFill('solid', fgColor=c)
def brd():
    s = Side(style='thin', color=C_BORD)
    return Border(bottom=s, right=s)
def aln(h='left'): return Alignment(horizontal=h, vertical='center', wrap_text=False)

wb = Workbook()
ws = wb.active
ws.title = 'Leilões'
ws2 = wb.create_sheet('Resumo por Casa')

# ── Cabeçalho ─────────────────────────────────────────────────────────────
HDRS = [
    'Rec','Score','Data','Casa','Artista','Titulo','Tipo','Tecnica',
    'Dim','Area cm²','Ano','Ass','Base R$','Lance R$',
    'R$/cm²','Med R$/cm²','Fator Tam','Med Hist R$','Fator Abs',
    'Liquidez','Tier','Ver Lote'
]
N = len(HDRS)
WIDTHS = [9,7,12,28,28,26,9,22,13,9,6,10,12,12,10,10,10,13,10,9,14,35]

# Título
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N)
tc = ws.cell(row=1, column=1,
    value=f'LEILÕES  |  {label_data.upper()}  |  Gerado {datetime.now().strftime("%d/%m/%Y %H:%M")}')
tc.font = Font(name='Calibri', size=16, bold=True, color=C_GOLD)
tc.fill = fill(C_DARK); tc.alignment = aln('center')
ws.row_dimensions[1].height = 34

# Subtítulo
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N)
sc2 = ws.cell(row=2, column=1,
    value=f'{len(df)} lotes  |  {len(contagem)} casas  |  Fator Tam = R$/cm² ajustado por tamanho')
sc2.font = Font(name='Calibri', size=10, color=C_GRAY)
sc2.fill = fill(C_DARK); sc2.alignment = aln('center')
ws.row_dimensions[2].height = 16

# Cabeçalhos de coluna
for col, h in enumerate(HDRS, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = Font(name='Calibri', size=10, bold=True, color=C_GOLD)
    c.fill = fill(C_CARD)
    c.alignment = aln('center')
    c.border = Border(bottom=Side(style='medium', color=C_GOLD))
ws.row_dimensions[3].height = 28

for i, w in enumerate(WIDTHS, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A4'

# ── Dados ──────────────────────────────────────────────────────────────────
for ri, (_, row) in enumerate(df.iterrows(), 4):
    bg  = C_ALT if ri % 2 == 0 else C_CARD
    sc  = row['Score']
    tier = row['Tier']
    tb, tf = TIER_STYLE.get(tier, (C_CARD, C_TEXT))
    rec  = row['Rec']

    def wc(col, val, fmt=None, h='left', bold=False, color=C_TEXT, bg_=None):
        cell = ws.cell(row=ri, column=col, value=val)
        cell.font = Font(name='Calibri', size=10, bold=bold, color=color)
        cell.fill = fill(bg_ if bg_ is not None else bg)
        cell.alignment = aln(h)
        cell.border = brd()
        if fmt: cell.number_format = fmt
        return cell

    # Col 1 — Rec
    rc = ws.cell(row=ri, column=1, value=rec)
    rc.font = Font(name='Calibri', size=10, bold=True, color=REC_COLOR.get(rec, C_TEXT))
    rc.fill = fill(bg); rc.alignment = aln('center'); rc.border = brd()

    # Col 2 — Score
    sc_color = C_GREEN if sc >= 75 else (C_BLUE if sc >= 55 else (C_AMBER if sc >= 35 else C_GRAY))
    scc = ws.cell(row=ri, column=2, value=sc)
    scc.font = Font(name='Calibri', size=11, bold=True, color=sc_color)
    scc.fill = fill(bg); scc.alignment = aln('center'); scc.border = brd()

    wc(3,  row['Data'],       h='center')

    # Col 4 — Casa (com cor por casa)
    wc(4,  row['Casa'])

    # Col 5 — Artista (com cor tier)
    ac = ws.cell(row=ri, column=5, value=row['Artista'])
    ac.font = Font(name='Calibri', size=10, bold=True, color=tf)
    ac.fill = fill(tb); ac.alignment = aln(); ac.border = brd()

    wc(6,  row['Titulo'])
    wc(7,  row['Tipo'],       h='center')
    wc(8,  row['Tecnica'])
    wc(9,  row['Dimensoes'],  h='center')
    wc(10, row['Area_cm2'],   h='center', fmt='#,##0')
    wc(11, row['Ano'],        h='center')
    wc(12, row['Assinatura'], h='center')

    # Col 13 — Base
    bc = ws.cell(row=ri, column=13, value=row['Lance_Base'])
    bc.number_format = 'R$ #,##0'
    bc.font = Font(name='Calibri', size=10, bold=True, color=C_GOLD)
    bc.fill = fill(bg); bc.alignment = aln('center'); bc.border = brd()

    # Col 14 — Lance atual
    la = row['Lance_Atual']
    lac = ws.cell(row=ri, column=14, value=la)
    lac.number_format = 'R$ #,##0'
    lac.font = Font(name='Calibri', size=10, bold=bool(la), color=C_RED if la else C_GRAY)
    lac.fill = fill(bg); lac.alignment = aln('center'); lac.border = brd()

    # Col 15 — R$/cm²
    rcm = row['R_cm2']
    rcmc = ws.cell(row=ri, column=15, value=rcm)
    rcmc.number_format = 'R$ #,##0.00'
    rcmc.font = Font(name='Calibri', size=10, color=C_TEXT if rcm else C_GRAY)
    rcmc.fill = fill(bg); rcmc.alignment = aln('center'); rcmc.border = brd()

    # Col 16 — Med R$/cm²
    mcm = row['Med_R_cm2']
    mcmc = ws.cell(row=ri, column=16, value=mcm)
    mcmc.number_format = 'R$ #,##0.00'
    mcmc.font = Font(name='Calibri', size=10, color=C_GREEN if mcm else C_GRAY)
    mcmc.fill = fill(bg); mcmc.alignment = aln('center'); mcmc.border = brd()

    # Col 17 — Fator Tam (número, mostra "3.5x", filtrável)
    mt = row['Mult_Tam']
    mtc = ws.cell(row=ri, column=17, value=mt)
    mtc.number_format = '0.00'
    mtc.font = Font(name='Calibri', size=10, bold=bool(mt and mt >= 2),
                    color=C_GREEN if (mt and mt >= 2) else (C_AMBER if mt else C_GRAY))
    mtc.fill = fill(bg); mtc.alignment = aln('center'); mtc.border = brd()

    # Col 18 — Med Hist absoluta
    med = row['Mediana_Hist']
    mc = ws.cell(row=ri, column=18, value=med)
    mc.number_format = 'R$ #,##0'
    mc.font = Font(name='Calibri', size=10, color=C_GREEN if med else C_GRAY)
    mc.fill = fill(bg); mc.alignment = aln('center'); mc.border = brd()

    # Col 19 — Fator Abs (número, mostra "3.5x", filtrável)
    mult = row['Multiplo']
    multc = ws.cell(row=ri, column=19, value=mult)
    multc.number_format = '0.00'
    multc.font = Font(name='Calibri', size=10, bold=bool(mult and mult >= 2),
                      color=C_GREEN if (mult and mult >= 2) else C_TEXT)
    multc.fill = fill(bg); multc.alignment = aln('center'); multc.border = brd()

    # Col 20 — Liquidez
    liq = row['Liquidez']
    lc = ws.cell(row=ri, column=20, value=liq)
    lc.number_format = '0.0'
    lc.font = Font(name='Calibri', size=10,
                   color=C_GREEN if liq >= 7 else (C_AMBER if liq >= 4 else C_GRAY))
    lc.fill = fill(bg); lc.alignment = aln('center'); lc.border = brd()

    # Col 21 — Tier
    tc2 = ws.cell(row=ri, column=21, value=tier)
    tc2.font = Font(name='Calibri', size=9, bold=True, color=tf)
    tc2.fill = fill(tb); tc2.alignment = aln('center'); tc2.border = brd()

    # Col 22 — URL
    url = row['URL']
    uc = ws.cell(row=ri, column=22, value='Ver lote' if url else '')
    if url:
        uc.hyperlink = url
        uc.font = Font(name='Calibri', size=9, color=C_BLUE, underline='single')
    else:
        uc.font = Font(name='Calibri', size=9, color=C_GRAY)
    uc.fill = fill(bg); uc.alignment = aln('center'); uc.border = brd()

    ws.row_dimensions[ri].height = 22

# Autofilter após escrita dos dados (range completo = header + dados)
ultima_linha = 3 + len(df)
ws.auto_filter.ref = f'A3:{get_column_letter(N)}{ultima_linha}'

# ── ABA 2: Resumo por Casa ────────────────────────────────────────────────
df_casa = df.groupby('Casa').agg(
    Lotes=('Score','count'),
    Score_Medio=('Score','mean'),
    FORTE=('Rec', lambda x: (x=='FORTE').sum()),
    BOA=('Rec', lambda x: (x=='BOA').sum()),
    MODERADA=('Rec', lambda x: (x=='MODERADA').sum()),
    Base_Min=('Lance_Base','min'),
    Base_Max=('Lance_Base','max'),
    Base_Medio=('Lance_Base','mean'),
).reset_index().sort_values('Score_Medio', ascending=False)

ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
th = ws2.cell(row=1, column=1, value=f'RESUMO POR CASA  |  {label_data}')
th.font = Font(name='Calibri', size=14, bold=True, color=C_GOLD)
th.fill = fill(C_DARK); th.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

h2 = ['Casa','Lotes','Score Médio','FORTE','BOA','MODERADA','Base Min','Base Max','Base Médio']
for col, h in enumerate(h2, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.font = Font(name='Calibri', size=10, bold=True, color=C_GOLD)
    c.fill = fill(C_CARD); c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = Border(bottom=Side(style='medium', color=C_GOLD))
ws2.row_dimensions[2].height = 24

for ri2, (_, r) in enumerate(df_casa.iterrows(), 3):
    bg = C_ALT if ri2 % 2 == 0 else C_CARD
    vals = [
        (r['Casa'],None,'left'),
        (r['Lotes'],None,'center'),
        (round(r['Score_Medio'],1),None,'center'),
        (r['FORTE'],None,'center'),
        (r['BOA'],None,'center'),
        (r['MODERADA'],None,'center'),
        (r['Base_Min'],'R$ #,##0','center'),
        (r['Base_Max'],'R$ #,##0','center'),
        (r['Base_Medio'],'R$ #,##0','center'),
    ]
    for col, (val, fmt, h) in enumerate(vals, 1):
        c = ws2.cell(row=ri2, column=col, value=val)
        c.font = Font(name='Calibri', size=10, color=C_TEXT)
        c.fill = fill(bg)
        c.alignment = Alignment(horizontal=h, vertical='center')
        c.border = Border(bottom=Side(style='thin', color=C_BORD), right=Side(style='thin', color=C_BORD))
        if fmt: c.number_format = fmt
    ws2.row_dimensions[ri2].height = 18

for i, w in enumerate([32,8,12,8,8,10,13,13,13], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws.sheet_properties.tabColor  = C_GOLD
ws2.sheet_properties.tabColor = '3B82F6'

# ── Salva ──────────────────────────────────────────────────────────────────
nome_arquivo = f'relatorio_{HOJE.strftime("%Y%m%d")}.xlsx'
out = os.path.join(_DIR, nome_arquivo)
wb.save(out)

print(f'\n  Score >= 75 (FORTE):   {(df["Rec"]=="FORTE").sum()}')
print(f'  Score >= 55 (BOA):     {(df["Rec"]=="BOA").sum()}')
print(f'  Score >= 35 (MODERADA):{(df["Rec"]=="MODERADA").sum()}')
print(f'\nPlanilha salva: {nome_arquivo}')
print(f'Fim: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}')
