#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json, re, unicodedata
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def norm_artista(s):
    s = re.sub(r'\(.*?\)', '', s or '').strip()
    nfkd = unicodedata.normalize('NFKD', s.upper())
    s = ''.join(c for c in nfkd if not unicodedata.combining(c))
    s = re.sub(r'[^A-Z\s]', '', s).strip()
    return re.sub(r'\s+', ' ', s)

def vb(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def parse_dims(s: str) -> float:
    """'50x70cm', '50,5 x 30 cm', '100X80' → área em cm². Retorna 0 se não parsear."""
    if not s: return 0.0
    s = str(s).lower().replace(' ', '').replace('cm', '').replace('mm', '')
    m = re.match(r'([\d,\.]+)[x×]([\d,\.]+)', s)
    if m:
        try:
            w = float(m.group(1).replace(',', '.'))
            h = float(m.group(2).replace(',', '.'))
            if 0 < w < 1000 and 0 < h < 1000:
                return round(w * h, 1)
        except: pass
    return 0.0

# ── Histórico de preços ────────────────────────────────────────────────────
historico    = {}   # artista → [preços absolutos]
historico_cm2 = {}  # artista → [R$/cm²] (apenas obras com dimensões)

def add_preco(artista, preco, dims=''):
    k = norm_artista(artista)
    if not k or preco <= 0: return
    historico.setdefault(k, []).append(preco)
    area = parse_dims(dims)
    if area > 0:
        historico_cm2.setdefault(k, []).append(round(preco / area, 4))

with open('bolsadearte_db.json', 'r', encoding='utf-8') as f:
    bda = json.load(f)
for v in bda.values():
    if isinstance(v, dict):
        add_preco(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

with open('arrematearte_db.json', 'r', encoding='utf-8') as f:
    arr = json.load(f)
for k, v in arr.items():
    if not k.startswith('__meta') and isinstance(v, dict):
        add_preco(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

# Histórico dos finalizados do próprio LeiloesBR
with open('leiloesbr_db.json', 'r', encoding='utf-8') as f:
    db = json.load(f)
for v in db.values():
    if isinstance(v, dict) and not v.get('_ignorado') and v.get('em_leilao') is False:
        add_preco(v.get('artista',''), vb(v.get('maior_lance')), v.get('dimensoes',''))

print(f'Base historica: {len(historico)} artistas | {sum(len(v) for v in historico.values())} precos')

# ── Lotes ativos ───────────────────────────────────────────────────────────
ativos = [(k, v) for k, v in db.items()
          if isinstance(v, dict) and not v.get('_ignorado') and v.get('em_leilao', True)]
print(f'Lotes ativos no DB: {len(ativos)}')

TIER1 = {norm_artista(a) for a in [
    'TARSILA DO AMARAL','CANDIDO PORTINARI','EMILIANO DI CAVALCANTI',
    'ALFREDO VOLPI','ALBERTO DA VEIGA GUIGNARD','JOSE PANCETTI',
    'ISMAEL NERY','ANITA MALFATTI','DJANIRA DA MOTTA E SILVA',
    'ARTHUR TIMOTEO DA COSTA','ROBERTO BURLE MARX','HELIO OITICICA',
    'LASAR SEGALL','VICTOR BRECHERET','MARIO ZANINI','RUBEM VALENTIM',
    'ALDO BONADEI','FULVIO PENNACCHI','GEORGINA DE ALBUQUERQUE',
    'TARSILA','DI CAVALCANTI',
]}
TIER2 = {norm_artista(a) for a in [
    'ALDEMIR MARTINS','CARYIBE','CARYBE','ATHOS BULCAO','CARLOS SCLIAR',
    'RUBENS GERCHMAN','CLAUDIO TOZZI','FERREIRA GULLAR','THOMAZ IANELLI',
    'ANTONIO POTEIRO','RENINA KATZ','LOTHAR CHAROUX','DARIO MECATTI',
    'INOS CORRADIN','HANSEN BAHIA','LIVIO ABRAMO','MILTON DACOSTA',
    'DURVAL PEREIRA','RUTH SCHLOSS','JOSE ANTONIO DA SILVA',
    'IWAO NAKAJIMA','CAROL KOSSAK','EDUARDO SUED','ALICE BRILL',
    'FRANCISCO BRENNAND','VICENTE DO REGO MONTEIRO','FRANCISCO DA SILVA',
    'SERGIO BERTONI','MARCELO GRASSMANN','GERSON DE SOUZA',
    'MANOEL SANTIAGO','SYLVIO PINTO','MANEZINHO ARAUJO',
]}

rows = []
for key, v in ativos:
    base  = vb(v.get('lance_base'))
    lance = vb(v.get('maior_lance'))
    preco_ref = lance if lance > 0 else base
    if preco_ref <= 0:
        continue

    artista_raw  = v.get('artista','') or ''
    if not artista_raw or artista_raw.lower() in ('autor desconhecido','?','','desconhecido'):
        continue

    artista_norm    = norm_artista(artista_raw)
    artista_display = artista_raw.strip().title()

    tec  = v.get('tecnica','')  or ''
    dims = v.get('dimensoes','') or ''
    ano  = v.get('ano','')      or ''
    ass  = v.get('assinatura','') or ''
    casa = v.get('casa','')     or ''
    data = v.get('data_leilao','') or v.get('data_coleta','')
    url  = v.get('url_detalhe','') or ''
    foto = v.get('foto_url','') or ''

    hist   = historico.get(artista_norm, [])
    n_hist = len(hist)
    med_hist = float(np.median(hist)) if hist else 0.0
    max_hist = float(max(hist))       if hist else 0.0

    multiplo = round(med_hist / preco_ref, 1) if (preco_ref > 0 and med_hist > 0) else 0.0

    # ── Ajuste por tamanho (R$/cm²) ────────────────────────────────────────
    area_lote      = parse_dims(dims)
    preco_cm2_lote = round(preco_ref / area_lote, 4) if area_lote > 0 else 0.0
    hist_cm2       = historico_cm2.get(artista_norm, [])
    n_hist_cm2     = len(hist_cm2)
    med_cm2_hist   = float(np.median(hist_cm2)) if hist_cm2 else 0.0
    mult_cm2       = round(med_cm2_hist / preco_cm2_lote, 1) if (preco_cm2_lote > 0 and med_cm2_hist > 0) else 0.0

    # Múltiplo efetivo: usa ajustado por tamanho se disponível, senão absoluto
    multiplo_efetivo = mult_cm2 if mult_cm2 > 0 else multiplo

    if artista_norm in TIER1:
        tier = 'A - Mestre Nacional'
    elif artista_norm in TIER2:
        tier = 'B - Moderno Relevante'
    elif n_hist >= 5:
        tier = 'C - Mercado Ativo'
    elif n_hist >= 1:
        tier = 'D - Referencia Limitada'
    else:
        tier = 'E - Sem Historico'

    liq = min(10.0, n_hist * 1.5)
    if artista_norm in TIER1:  liq = min(10.0, liq + 3.0)
    elif artista_norm in TIER2: liq = min(10.0, liq + 1.5)
    liq = round(liq, 1)

    oport = 0.0
    if multiplo_efetivo > 0:
        oport += min(40.0, multiplo_efetivo * 8)
    oport += liq * 3
    tec_l = unicodedata.normalize('NFKD', tec.lower())
    tec_l = ''.join(c for c in tec_l if not unicodedata.combining(c))
    is_pintura = any(k in tec_l for k in ['oleo','acrilica','acrilico','aquarela',
                                           'guache','gouache','pastel','tempera',
                                           'tecnica mista','nanquim','tinta'])
    is_grafica = any(k in tec_l for k in ['serigrafia','litografia','xilogravura',
                                           'gravura','offset','multiplo'])
    if is_pintura:   oport += 20
    elif is_grafica and artista_norm in TIER1: oport += 8
    elif 'desenho' in tec_l: oport += 10
    if artista_norm in TIER1:   oport += 15
    elif artista_norm in TIER2: oport += 8
    if ass in ('Assinado',): oport += 5
    # Bônus/penalidade por formato (pintura)
    if is_pintura and area_lote > 0:
        if area_lote >= 3000:   oport += 5   # grande formato (ex. 60x50+)
        elif area_lote < 400:   oport -= 3   # obra muito pequena
    if n_hist == 0: oport *= 0.3
    oport = round(min(100.0, oport), 1)

    if oport >= 75:   rec = 'FORTE'
    elif oport >= 55: rec = 'BOA'
    elif oport >= 35: rec = 'MODERADA'
    else:             rec = 'BAIXA'

    rows.append({
        'Score':        oport,
        'Rec':          rec,
        'Data Leilao':  data,
        'Casa':         casa,
        'Artista':      artista_display,
        'Tecnica':      tec,
        'Dimensoes':    dims,
        'Area cm2':     round(area_lote) if area_lote > 0 else None,
        'Ano':          ano,
        'Assinatura':   ass,
        'Valor Base':   base,
        'Lance Atual':  lance if lance > 0 else None,
        'Preco Ref':    preco_ref,
        'R$/cm2':       round(preco_cm2_lote, 2) if preco_cm2_lote > 0 else None,
        'Med R$/cm2':   round(med_cm2_hist, 2) if med_cm2_hist > 0 else None,
        'Mult Tam':     mult_cm2 if mult_cm2 > 0 else None,
        'Mediana Hist': med_hist if med_hist > 0 else None,
        'Max Hist':     max_hist if max_hist > 0 else None,
        'Multiplo':     multiplo if multiplo > 0 else None,
        'Liquidez':     liq,
        'N Hist':       n_hist,
        'Tier':         tier,
        'URL':          url,
        'Foto':         foto,
    })

df = pd.DataFrame(rows).sort_values('Score', ascending=False)
top = df[df['Score'] >= 35].copy()

print(f'\nLotes com artista identificado: {len(df)}')
print(f'Score >= 35: {len(top)}')
print('\nTop 20:')
for _, r in top.head(20).iterrows():
    mult = f'{r["Multiplo"]}x' if r['Multiplo'] else 'sem ref'
    print(f'  {str(r["Data Leilao"]):>12} | {r["Artista"]:<28} | R${r["Valor Base"]:>7,.0f} | {mult:>7} | Liq:{r["Liquidez"]:>4} | {r["Score"]:>5} | {r["Rec"]} | {r["Casa"][:25]}')

# ── Excel ──────────────────────────────────────────────────────────────────
wb = Workbook()
ws1 = wb.active; ws1.title = 'Top Oportunidades'
ws2 = wb.create_sheet('Todos os Lotes')
ws3 = wb.create_sheet('Metodologia')

C_DARK = '0F0F1A'; C_CARD = '1A1A2E'; C_ALT = '161628'
C_GOLD = 'FFD700'; C_GREEN = '22C55E'; C_BLUE = '60A5FA'
C_RED = 'EF4444'; C_AMBER = 'F59E0B'; C_GRAY = '94A3B8'; C_TEXT = 'E2E8F0'
C_BORD = '2D2D4E'
TIER_STYLE = {
    'A - Mestre Nacional':    ('1E3A5F','FFD700'),
    'B - Moderno Relevante':  ('1A3A2A','22C55E'),
    'C - Mercado Ativo':      ('1A2A3A','60A5FA'),
    'D - Referencia Limitada':('2A2A1A','F59E0B'),
    'E - Sem Historico':      ('2A1A1A','94A3B8'),
}
REC_COLOR = {'FORTE':'22C55E','BOA':'60A5FA','MODERADA':'F59E0B','BAIXA':'94A3B8'}

def fill(c): return PatternFill('solid', fgColor=c)
def brd():
    s = Side(style='thin', color=C_BORD)
    return Border(bottom=s, right=s)
def aln(h='left'): return Alignment(horizontal=h, vertical='center', wrap_text=True)

def write_header(ws, r, n, title, sub=None):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
    c = ws.cell(row=r, column=1, value=title)
    c.font = Font(name='Calibri', size=17, bold=True, color=C_GOLD)
    c.fill = fill(C_DARK); c.alignment = aln('center')
    ws.row_dimensions[r].height = 36
    if sub:
        ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=n)
        c2 = ws.cell(row=r+1, column=1, value=sub)
        c2.font = Font(name='Calibri', size=10, color=C_GRAY)
        c2.fill = fill(C_DARK); c2.alignment = aln('center')
        ws.row_dimensions[r+1].height = 17

def write_cols(ws, r, hdrs):
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = Font(name='Calibri', size=10, bold=True, color=C_GOLD)
        c.fill = fill(C_CARD); c.alignment = aln('center')
        c.border = Border(bottom=Side(style='medium', color=C_GOLD))
    ws.row_dimensions[r].height = 28

# ABA 1
HDR1 = ['Rec','Score','Data Leilao','Casa','Artista','Tecnica','Dim','Area cm²',
         'Ass','Valor Base R$','Lance Atual R$','R$/cm²','Med R$/cm²','Mult Tam',
         'Mediana Hist R$','Max Hist R$','Multiplo','Liquidez','Tier','Ver Lote']
N1 = len(HDR1)
write_header(ws1, 1, N1,
    'LEILOESBR  |  OPORTUNIDADES EM LEILAO AGORA',
    f'Score >= 35  |  Gerado {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  Cruzado com {len(historico):,} artistas historicos')
write_cols(ws1, 3, HDR1)
ws1.freeze_panes = 'A4'
for i, w in enumerate([9,7,12,26,28,22,12,9,10,13,13,10,10,10,14,13,10,9,22,35], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

for ri, (_, row) in enumerate(top.iterrows(), 4):
    bg = C_ALT if ri % 2 == 0 else C_CARD
    score = row['Score']
    tier  = row['Tier']
    tb, tf = TIER_STYLE.get(tier, (C_CARD, C_TEXT))
    rec = row['Rec']

    def wc(col, val, fmt=None, h='left', bold=False, color=C_TEXT, bg_=None):
        cell = ws1.cell(row=ri, column=col, value=val)
        cell.font = Font(name='Calibri', size=10, bold=bold, color=color)
        cell.fill = fill(bg_ if bg_ is not None else bg)
        cell.alignment = aln(h)
        cell.border = brd()
        if fmt: cell.number_format = fmt
        return cell

    rc = ws1.cell(row=ri, column=1, value=rec)
    rc.font = Font(name='Calibri', size=10, bold=True, color=REC_COLOR.get(rec, C_TEXT))
    rc.fill = fill(bg); rc.alignment = aln('center'); rc.border = brd()

    sc_col = C_GREEN if score >= 75 else (C_BLUE if score >= 55 else (C_AMBER if score >= 35 else C_GRAY))
    sc = ws1.cell(row=ri, column=2, value=score)
    sc.font = Font(name='Calibri', size=11, bold=True, color=sc_col)
    sc.fill = fill(bg); sc.alignment = aln('center'); sc.border = brd()

    wc(3,  row['Data Leilao'], h='center')
    wc(4,  row['Casa'])

    ac = ws1.cell(row=ri, column=5, value=row['Artista'])
    ac.font = Font(name='Calibri', size=10, bold=True, color=tf)
    ac.fill = fill(tb); ac.alignment = aln(); ac.border = brd()

    wc(6,  row['Tecnica'])
    wc(7,  row['Dimensoes'], h='center')
    wc(8,  row['Area cm2'],  h='center', fmt='#,##0')
    wc(9,  row['Assinatura'], h='center')

    vc = ws1.cell(row=ri, column=10, value=row['Valor Base'])
    vc.number_format = 'R$ #,##0'
    vc.font = Font(name='Calibri', size=10, bold=True, color=C_GOLD)
    vc.fill = fill(bg); vc.alignment = aln('center'); vc.border = brd()

    la = row['Lance Atual']
    lac = ws1.cell(row=ri, column=11, value=la)
    lac.number_format = 'R$ #,##0'
    lac.font = Font(name='Calibri', size=10, bold=bool(la), color=C_RED if la else C_GRAY)
    lac.fill = fill(bg); lac.alignment = aln('center'); lac.border = brd()

    # R$/cm² do lote
    rcm = row['R$/cm2']
    rcmc = ws1.cell(row=ri, column=12, value=rcm)
    rcmc.number_format = 'R$ #,##0.00'
    rcmc.font = Font(name='Calibri', size=10, color=C_TEXT if rcm else C_GRAY)
    rcmc.fill = fill(bg); rcmc.alignment = aln('center'); rcmc.border = brd()

    # Mediana histórica R$/cm²
    mcm = row['Med R$/cm2']
    mcmc = ws1.cell(row=ri, column=13, value=mcm)
    mcmc.number_format = 'R$ #,##0.00'
    mcmc.font = Font(name='Calibri', size=10, color=C_GREEN if mcm else C_GRAY)
    mcmc.fill = fill(bg); mcmc.alignment = aln('center'); mcmc.border = brd()

    # Múltiplo ajustado por tamanho (número com sufixo visual "x" — filtrável)
    mt = row['Mult Tam']
    mtc = ws1.cell(row=ri, column=14, value=mt)
    mtc.number_format = '0.00'
    mtc.font = Font(name='Calibri', size=10, bold=bool(mt and mt >= 2),
                    color=C_GREEN if (mt and mt >= 2) else (C_AMBER if mt else C_GRAY))
    mtc.fill = fill(bg); mtc.alignment = aln('center'); mtc.border = brd()

    med = row['Mediana Hist']
    mc = ws1.cell(row=ri, column=15, value=med)
    mc.number_format = 'R$ #,##0'
    mc.font = Font(name='Calibri', size=10, color=C_GREEN if med else C_GRAY)
    mc.fill = fill(bg); mc.alignment = aln('center'); mc.border = brd()

    mx = row['Max Hist']
    mxc = ws1.cell(row=ri, column=16, value=mx)
    mxc.number_format = 'R$ #,##0'
    mxc.font = Font(name='Calibri', size=10, color=C_AMBER if mx else C_GRAY)
    mxc.fill = fill(bg); mxc.alignment = aln('center'); mxc.border = brd()

    # Múltiplo absoluto (número com sufixo "x" — filtrável)
    mult = row['Multiplo']
    multc = ws1.cell(row=ri, column=17, value=mult)
    multc.number_format = '0.00'
    multc.font = Font(name='Calibri', size=10, bold=bool(mult and mult >= 2),
                      color=C_GREEN if (mult and mult >= 2) else C_TEXT)
    multc.fill = fill(bg); multc.alignment = aln('center'); multc.border = brd()

    liq = row['Liquidez']
    lc = ws1.cell(row=ri, column=18, value=liq)
    lc.number_format = '0.0'
    lc.font = Font(name='Calibri', size=10,
                   color=C_GREEN if liq >= 7 else (C_AMBER if liq >= 4 else C_GRAY))
    lc.fill = fill(bg); lc.alignment = aln('center'); lc.border = brd()

    tc = ws1.cell(row=ri, column=19, value=tier)
    tc.font = Font(name='Calibri', size=9, bold=True, color=tf)
    tc.fill = fill(tb); tc.alignment = aln('center'); tc.border = brd()

    url = row['URL']
    uc = ws1.cell(row=ri, column=20, value='Ver lote' if url else '')
    if url:
        uc.hyperlink = url
        uc.font = Font(name='Calibri', size=9, color=C_BLUE, underline='single')
    else:
        uc.font = Font(name='Calibri', size=9, color=C_GRAY)
    uc.fill = fill(bg); uc.alignment = aln('center'); uc.border = brd()

    ws1.row_dimensions[ri].height = 26

ultima_linha_1 = 3 + len(top)
ws1.auto_filter.ref = f'A3:{get_column_letter(N1)}{ultima_linha_1}'

# ABA 2
HDR2 = ['Score','Rec','Data','Casa','Artista','Tecnica','Dim','Area cm²',
        'Valor Base','R$/cm²','Med R$/cm²','Mult Tam','Mediana Hist',
        'Multiplo','Liquidez','N Hist','Tier','URL']
write_header(ws2, 1, len(HDR2), 'TODOS OS LOTES ATIVOS COM ARTISTA IDENTIFICADO')
write_cols(ws2, 2, HDR2)
ws2.freeze_panes = 'A3'
for i, w in enumerate([7,8,12,24,26,20,12,9,12,10,10,10,13,9,9,7,22,35], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

for ri, (_, row) in enumerate(df.iterrows(), 3):
    bg = C_ALT if ri % 2 == 0 else C_CARD
    for col, key, fmt, h in [
        (1,'Score','0','center'), (2,'Rec',None,'center'),
        (3,'Data Leilao',None,'center'), (4,'Casa',None,'left'),
        (5,'Artista',None,'left'), (6,'Tecnica',None,'left'),
        (7,'Dimensoes',None,'center'), (8,'Area cm2','#,##0','center'),
        (9,'Valor Base','R$ #,##0','center'),
        (10,'R$/cm2','R$ #,##0.00','center'),
        (11,'Med R$/cm2','R$ #,##0.00','center'),
        (12,'Mult Tam','0.00','center'),
        (13,'Mediana Hist','R$ #,##0','center'),
        (14,'Multiplo','0.00','center'), (15,'Liquidez','0.0','center'),
        (16,'N Hist',None,'center'), (17,'Tier',None,'center'),
        (18,'URL',None,'left'),
    ]:
        val = row.get(key)
        c = ws2.cell(row=ri, column=col,
                     value=val if (val is not None and val != 0 and pd.notna(val)) else None)
        c.font = Font(name='Calibri', size=9, color=C_TEXT)
        c.fill = fill(bg); c.alignment = aln(h); c.border = brd()
        if fmt: c.number_format = fmt
        if col == 18 and val:
            c.hyperlink = str(val)
            c.font = Font(name='Calibri', size=9, color=C_BLUE, underline='single')
    ws2.row_dimensions[ri].height = 16

ultima_linha_2 = 2 + len(df)
ws2.auto_filter.ref = f'A2:{get_column_letter(len(HDR2))}{ultima_linha_2}'

# ABA 3
ws3.column_dimensions['A'].width = 85
for ri, (txt, bold, size, color) in enumerate([
    ('METODOLOGIA', True, 16, C_GOLD), ('', False, 11, C_TEXT),
    ('FONTES HISTORICAS', True, 13, C_BLUE),
    (f'Bolsa de Arte: {sum(1 for v in bda.values() if isinstance(v,dict) and vb(v.get("maior_lance"))>0):,} precos', False, 11, C_TEXT),
    (f'Arremate Arte: {sum(1 for k,v in arr.items() if not k.startswith("__meta") and isinstance(v,dict) and vb(v.get("maior_lance"))>0):,} lances', False, 11, C_TEXT),
    ('LeiloesBR historico: lotes finalizados cruzados por artista', False, 11, C_TEXT),
    ('', False, 11, C_TEXT),
    ('SCORE (0-100)', True, 13, C_BLUE),
    ('Multiplo (ate 40pts): usa R$/cm2 ajustado por tamanho se disponivel, senao multiplo absoluto', False, 11, C_TEXT),
    ('Liquidez ate 30pts | Pintura +20 | Tier A +15 | Tier B +8 | Assinado +5', False, 11, C_TEXT),
    ('Grande formato (>=3000cm2) +5pts | Obra minuscula (<400cm2) -3pts | Sem historico: x0.3', False, 11, C_TEXT),
    ('', False, 11, C_TEXT),
    ('RECOMENDACAO', True, 13, C_BLUE),
    ('FORTE (75-100): alta confianca, multiplo e liquidez comprovados', False, 11, C_GREEN),
    ('BOA   (55-74):  boa relacao risco/retorno', False, 11, C_BLUE),
    ('MODERADA (35-54): potencial com alguma incerteza', False, 11, C_AMBER),
    ('', False, 11, C_TEXT),
    (f'Gerado: {datetime.now().strftime("%d/%m/%Y %H:%M")}', False, 10, C_GRAY),
], 1):
    c = ws3.cell(row=ri, column=1, value=txt)
    c.font = Font(name='Calibri', size=size, bold=bold, color=color)
    c.fill = fill(C_DARK); c.alignment = aln(); ws3.row_dimensions[ri].height = 20 if bold else 17

ws1.sheet_properties.tabColor = C_GOLD
ws2.sheet_properties.tabColor = '3B82F6'
ws3.sheet_properties.tabColor = '64748B'

OUT = 'leiloesbr_oportunidades.xlsx'
wb.save(OUT)
print(f'\nPlanilha salva: {OUT}')
